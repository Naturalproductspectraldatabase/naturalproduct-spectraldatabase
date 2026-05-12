import base64
import hashlib
import hmac
import html
import io
import json
import mimetypes
import os
import re
import sqlite3
import ssl
import sys
import textwrap
import time
import urllib.error
import urllib.request
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import quote, unquote, urlencode, urlparse

import pandas as pd
import streamlit as st
try:
    import plotly.express as px
except Exception:
    px = None

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

streamlit_ketchersa = None
st_ketcher = None
KETCHER_STATUS = "local Ketcher unavailable"

try:
    from streamlit_ketchersa import streamlit_ketchersa as _local_streamlit_ketchersa

    streamlit_ketchersa = _local_streamlit_ketchersa
    KETCHER_STATUS = "local streamlit_ketchersa loaded"
except Exception:
    streamlit_ketchersa = None

try:
    from streamlit_ketcher import st_ketcher as _local_st_ketcher

    st_ketcher = _local_st_ketcher
    if streamlit_ketchersa is None:
        KETCHER_STATUS = "local streamlit_ketcher fallback loaded"
except Exception:
    st_ketcher = None

try:
    from rdkit import Chem, DataStructs, RDLogger
    from rdkit.Chem import AllChem
    RDLogger.DisableLog("rdApp.*")
except Exception:
    Chem = None
    DataStructs = None
    AllChem = None

try:
    from rdkit.Chem import Draw
except Exception:
    Draw = None

try:
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    Alignment = None
    Font = None
    PatternFill = None

# =========================
# Basic configuration
# =========================
def resolve_project_dir(script_dir: Path) -> Path:
    candidates = [script_dir, script_dir.parent]
    for candidate in candidates:
        if (candidate / "data").exists() and (candidate / "database").exists():
            return candidate
    return script_dir


PROJECT_DIR = resolve_project_dir(SCRIPT_DIR)
DATABASE_DIR = PROJECT_DIR / "database"
DATA_DIR = PROJECT_DIR / "data"
BRANDING_DIR = DATA_DIR / "branding"
BRANDING_OPTIMIZED_DIR = DATA_DIR / "branding_optimized"
STRUCTURES_DIR = DATA_DIR / "structures"
SPECTRA_DIR = DATA_DIR / "spectra"
TEMPLATES_DIR = DATA_DIR / "templates"
SUBMISSIONS_DIR = DATA_DIR / "submissions"
SUBMISSIONS_INBOX_DIR = SUBMISSIONS_DIR / "inbox"
SUBMISSIONS_REVIEWED_DIR = SUBMISSIONS_DIR / "reviewed"
SUBMISSIONS_APPROVED_DIR = SUBMISSIONS_DIR / "approved"
EXPORTS_DIR = DATA_DIR / "exports"
DOCS_DIR = DATA_DIR / "docs"
BACKUPS_DIR = DATABASE_DIR / "backups"
DB_PATH = PROJECT_DIR / "database" / "nmr.db"

MAX_PAGE_ICON_BYTES = 5 * 1024 * 1024
OWNER_CREDIT = "© Trianda Ayuning Tyas_project"
OWNER_EDITOR_USERNAME = "npdb_tyas"
OWNER_EDITOR_ROLES = {"owner", "owner_editor", "admin", "editor"}
PASSWORD_HASH_SCHEME = "pbkdf2_sha256"
SUPABASE_PAGE_SIZE = 1000


def pick_branding_asset(*filenames: str) -> Path:
    for filename in filenames:
        for branding_dir in (BRANDING_OPTIMIZED_DIR, BRANDING_DIR):
            candidate = branding_dir / filename
            if candidate.exists():
                return candidate
    return BRANDING_DIR / filenames[0]


def pick_branding_asset_fuzzy(*keywords: str, fallback: str) -> Path:
    normalized_keywords = [keyword.strip().lower() for keyword in keywords if keyword and keyword.strip()]
    for branding_dir in (BRANDING_OPTIMIZED_DIR, BRANDING_DIR):
        if not branding_dir.exists():
            continue
        candidates = sorted(
            [path for path in branding_dir.iterdir() if path.is_file()],
            key=lambda path: (len(path.name), path.name.lower()),
        )
        for candidate in candidates:
            normalized_name = candidate.name.lower()
            if all(keyword in normalized_name for keyword in normalized_keywords):
                return candidate
    return BRANDING_DIR / fallback


FAVICON_PATH = pick_branding_asset(
    "coral_favicon1.png",
    "Coral_favicon.png",
    "NP_favicon_tab.png",
    "favicon_tab.png",
    "NP_favicon2.png",
    "favicon_circle.png",
    "favicon2.png",
    "favicon.png",
)
SIDEBAR_LOGO_PATH = pick_branding_asset(
    "favicon_circle.png",
    "coral_favicon1.png",
    "favicon2.png",
    "favicon.png",
    "logo_header_web.png",
)
HEADER_LOGO_PATH = pick_branding_asset("logo_header_web.png", "header1_web.png", "logo_header.png", "header1.png", "header.png")
if not HEADER_LOGO_PATH.exists():
    HEADER_LOGO_PATH = pick_branding_asset_fuzzy("logo", "header", fallback="logo_header_web.png")
LOGIN_LOGO_PATH = pick_branding_asset(
    "logo_header_web.png",
    "header_main_web.png",
    "favicon_circle.png",
    "favicon2.png",
    "favicon.png",
)
LOGIN_BADGE_PATH = pick_branding_asset(
    "coral_favicon1.png",
    "Coral_favicon.png",
    "favicon_circle.png",
    "favicon2.png",
)
HERO_BANNER_PATH = pick_branding_asset("background.png", "background1.png", "header_main_web.png", "logo_header_web.png")
if not HERO_BANNER_PATH.exists():
    HERO_BANNER_PATH = pick_branding_asset_fuzzy("background", fallback="background1.png")
LOGIN_BACKGROUND_PATH = pick_branding_asset("background for login.png", "background1.png", "background.png")
if not LOGIN_BACKGROUND_PATH.exists():
    LOGIN_BACKGROUND_PATH = pick_branding_asset_fuzzy("login", "background", fallback="background for login.png")
LOGIN_LEFT_ART_PATH = pick_branding_asset("Onnamide A.png", "NCO cmp.png", "structures.png")
if not LOGIN_LEFT_ART_PATH.exists():
    LOGIN_LEFT_ART_PATH = pick_branding_asset_fuzzy("onnamide", fallback="Onnamide A.png")
LOGIN_RIGHT_ART_PATH = pick_branding_asset("NCO cmp 1.png", "NCO cmp.png", "bioactivity1.png")
if not LOGIN_RIGHT_ART_PATH.exists():
    LOGIN_RIGHT_ART_PATH = pick_branding_asset_fuzzy("nco", fallback="NCO cmp 1.png")


def _normalize_html_block(markup: str) -> str:
    if not isinstance(markup, str):
        return markup
    normalized = textwrap.dedent(markup).strip()
    if any(token in normalized for token in ("<div", "</div>", "<section", "</section>", "<span", "</span>", "<img", "<style")):
        return "\n".join(line.strip() for line in normalized.splitlines() if line.strip())
    return normalized

_STREAMLIT_MARKDOWN = st.markdown


def _safe_markdown(body, *args, **kwargs):
    if isinstance(body, str):
        stripped = body.strip()
        escaped = stripped.replace("&lt;", "<").replace("&gt;", ">")
        standalone_tag_pattern = r"(?:</(?:div|section|span|main|article|header|footer|aside)>\s*)+"
        if escaped in {"</div>", "</section>", "</span>", "</main>", "</article>", "</header>", "</footer>", "</aside>"}:
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
        if escaped and re.fullmatch(standalone_tag_pattern, escaped):
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
        if not kwargs.get("unsafe_allow_html") and escaped != stripped and re.fullmatch(standalone_tag_pattern, escaped):
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
    if kwargs.get("unsafe_allow_html") and isinstance(body, str):
        stripped = body.strip()
        if stripped in {"</div>", "</section>", "</span>", "</main>", "</article>", "</header>", "</footer>", "</aside>"}:
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
        if stripped and re.fullmatch(r"(?:</(?:div|section|span|main|article|header|footer|aside)>\s*)+", stripped):
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
        if stripped and re.fullmatch(r"<(?:div|section|span|main|article|header|footer|aside)(?:\s+[^>]*)?>", stripped):
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
        if stripped and re.fullmatch(r"(?:</(?:div|section|span|main|article|header|footer|aside)>\s*)+<(?:div|section|span|main|article|header|footer|aside)(?:\s+[^>]*)?>", stripped):
            return _STREAMLIT_MARKDOWN("", *args, **kwargs)
    if (
        kwargs.get("unsafe_allow_html")
        and isinstance(body, str)
        and any(token in body for token in ("<div", "<style", "<svg", "<img", "<span", "<section"))
    ):
        body = _normalize_html_block(body)
    return _STREAMLIT_MARKDOWN(body, *args, **kwargs)


st.markdown = _safe_markdown


def render_raw_html(markup: str):
    normalized = _normalize_html_block(markup)
    if hasattr(st, "html"):
        st.html(normalized)
    else:
        _STREAMLIT_MARKDOWN(normalized, unsafe_allow_html=True)


WORKSPACE_ART_PATH = pick_branding_asset("compound workspace.png", "updated.png", "structures.png")
if not WORKSPACE_ART_PATH.exists():
    WORKSPACE_ART_PATH = pick_branding_asset_fuzzy("workspace", fallback="compound workspace.png")
COMPOUNDS_ART_PATH = pick_branding_asset("compounds.png", "structures.png")
if not COMPOUNDS_ART_PATH.exists():
    COMPOUNDS_ART_PATH = pick_branding_asset_fuzzy("compound", fallback="compounds.png")
SPECTRA_ART_PATH = pick_branding_asset("spectra.png", "updated.png")
if not SPECTRA_ART_PATH.exists():
    SPECTRA_ART_PATH = pick_branding_asset_fuzzy("spectra", fallback="spectra.png")
STRUCTURES_ART_PATH = pick_branding_asset("structures.png", "updated.png")
if not STRUCTURES_ART_PATH.exists():
    STRUCTURES_ART_PATH = pick_branding_asset_fuzzy("structure", fallback="structures.png")
BIOACTIVITY_ART_PATH = pick_branding_asset("bioactivity.png", "bioactivity1.png", "updated.png")
if not BIOACTIVITY_ART_PATH.exists():
    BIOACTIVITY_ART_PATH = pick_branding_asset_fuzzy("bioactivity", fallback="bioactivity.png")
UPDATED_ART_PATH = pick_branding_asset("updated.png", "compound workspace.png")
if not UPDATED_ART_PATH.exists():
    UPDATED_ART_PATH = pick_branding_asset_fuzzy("updated", fallback="updated.png")
SEARCH_ART_PATH = pick_branding_asset("Search spectra.png", "spectra.png", "updated.png")
if not SEARCH_ART_PATH.exists():
    SEARCH_ART_PATH = pick_branding_asset_fuzzy("search", "spectra", fallback="Search spectra.png")
SEARCH_BIG_ART_PATH = pick_branding_asset("Search spectra big.png", "Search spectra.png", "spectra.png")
if not SEARCH_BIG_ART_PATH.exists():
    SEARCH_BIG_ART_PATH = pick_branding_asset_fuzzy("search", "spectra", "big", fallback="Search spectra big.png")
DOCUMENTATION_ART_PATH = pick_branding_asset("Documentation.png", "updated.png")
if not DOCUMENTATION_ART_PATH.exists():
    DOCUMENTATION_ART_PATH = pick_branding_asset_fuzzy("documentation", fallback="Documentation.png")

WORKFLOW_CARD_ART_PATHS = {
    "Browse Record": pick_branding_asset("Browse Record.png", "compounds.png"),
    "New Submission": pick_branding_asset("New Submission.png", "updated.png"),
    "Batch Import": pick_branding_asset("Batch Import.png", "spectra.png"),
    "Update Metadata": pick_branding_asset("Update Metadata.png", "updated.png"),
}

SIDEBAR_NAV_LABEL_ICONS = {
    "Home": ":material/home:",
    "Search Spectra": ":material/search:",
    "Browse Dashboard": ":material/dashboard:",
    "Start Submission": ":material/note_add:",
    "New Submission": ":material/note_add:",
    "Batch Import": ":material/upload_file:",
    "Update Metadata": ":material/edit_note:",
    "Delete Record": ":material/delete:",
    "Bioactivity": ":material/biotech:",
    "1H Peaks": ":material/monitoring:",
    "13C Peaks": ":material/analytics:",
    "Spectra Library": ":material/folder_open:",
    "Guide": ":material/menu_book:",
}

SIDEBAR_NAV_ICON_PATHS = {
    "Home": pick_branding_asset("favicon_circle.png", "logo_header_web.png"),
    "Search Spectra": pick_branding_asset("Search spectra.png", "spectra.png"),
    "Browse Dashboard": pick_branding_asset("Browse Dashboard.png", "compounds.png"),
    "Start Submission": pick_branding_asset("Start Submission.png", "New Submission.png"),
    "New Submission": pick_branding_asset("New Submission.png", "Start Submission.png"),
    "Batch Import": pick_branding_asset("Batch Import.png", "spectra.png"),
    "Update Metadata": pick_branding_asset("Update Metadata.png", "updated.png"),
    "Delete Record": pick_branding_asset("Delete Record.png", "updated.png"),
    "Bioactivity": pick_branding_asset("bioactivity.png", "updated.png"),
    "1H Peaks": pick_branding_asset("spectra.png", "updated.png"),
    "13C Peaks": pick_branding_asset("updated.png", "spectra.png"),
    "Spectra Library": pick_branding_asset("spectra.png", "updated.png"),
    "Guide": pick_branding_asset("Documentation.png", "updated.png"),
}

DASHBOARD_WORKFLOW_STEPS = [
    ("Browse Record", "Review compounds and linked spectra."),
    ("New Submission", "Add a new natural product record."),
    ("Batch Import", "Import multiple records from structured files."),
    ("Update Metadata", "Refine structure, source, and reference fields."),
]

DASHBOARD_HIGHLIGHTS = [
    ("Curation-first", "Only npdb_tyas can submit or revise records, so the public database stays tidy and consistent."),
    ("Cloud-ready", "Supabase is treated as the durable home for metadata, while local SQLite remains a working safety copy."),
    ("Searchable", "Keyword lookup, peak matching, and structure-aware workflows keep discovery practical for daily research."),
    ("Visual depth", "Gradient layering, branded artwork, and higher-contrast cards make the workspace feel more premium."),
    ("Research useful", "Compounds, spectra, structures, references, and bioactivity stay connected as one reviewable dossier."),
]

SIDEBAR_NAV_GROUPS = [
    ("Main Menu", [
        {"label": "Home", "section": "Dashboard"},
        {"label": "Search Spectra", "section": "Search & Match"},
        {"label": "Browse Dashboard", "section": "Compound Workspace", "compound_page": "Browse Record"},
        {"label": "Start Submission", "section": "Compound Workspace", "compound_page": "New Submission"},
    ]),
    ("Workflow", [
        {"label": "New Submission", "section": "Compound Workspace", "compound_page": "New Submission"},
        {"label": "Batch Import", "section": "Compound Workspace", "compound_page": "Batch Import"},
        {"label": "Update Metadata", "section": "Compound Workspace", "compound_page": "Update Metadata"},
        {"label": "Delete Record", "section": "Compound Workspace", "compound_page": "Delete Record"},
    ]),
    ("Data Library", [
        {"label": "Bioactivity", "section": "Bioactivity"},
        {"label": "1H Peaks", "section": "1H Peaks"},
        {"label": "13C Peaks", "section": "13C Peaks"},
        {"label": "Spectra Library", "section": "Spectra Library"},
        {"label": "Guide", "section": "Guide"},
    ]),
]

DEFAULT_CLASS_OPTIONS = [
    "Alkaloid",
    "Peptide",
    "Polyketide",
    "Steroid",
    "Terpenoid",
    "Phenolic",
    "Flavonoid",
    "Marine Natural Product",
]
DEFAULT_SOURCE_OPTIONS = [
    "Marine",
    "Sponge",
    "Soft Coral",
    "Hard Coral",
    "Tunicate",
    "Cyanobacteria",
    "Bacteria",
    "Coral",
    "Seaweed",
    "Microorganism",
    "Plant",
    "Fungus",
]
MARINE_SOURCE_ORGANISM_HINTS = [
    "adocidae",
    "haliclona",
    "marine",
    "oscillatoria",
    "stylissa",
]
SOURCE_NORMALIZATION_CACHE_VERSION = "marine-source-v2"
DEFAULT_DATA_SOURCE_OPTIONS = ["Experimental", "Literature", "In-house Archive"]
DEFAULT_SOLVENT_OPTIONS = ["CDCl3", "DMSO-d6", "CD3OD", "Acetone-d6", "Pyridine-d5"]
DEFAULT_SPECTRUM_TYPES = [
    "1H",
    "13C",
    "1H Raw Data",
    "13C Raw Data",
    "JCAMP-DX",
    "MNova",
    "COSY",
    "HSQC",
    "HMBC",
    "NOESY",
    "FTIR",
    "UV",
    "HRMS",
    "Supporting Data",
]
DEFAULT_BIOACTIVITY_CATEGORIES = [
    "Cytotoxicity",
    "Antibacterial",
    "Antifungal",
    "Antiviral",
    "Anti-inflammatory",
    "Antiparasitic",
    "Enzyme Inhibition",
    "Receptor Binding",
    "Antioxidant",
    "Ecological Activity",
]
DEFAULT_TARGET_CATEGORIES = [
    "Cell Line",
    "Bacterium",
    "Fungus",
    "Virus",
    "Parasite",
    "Enzyme",
    "Receptor",
    "In Vivo",
    "General",
]
DEFAULT_POTENCY_TYPES = [
    "IC50",
    "EC50",
    "MIC",
    "GI50",
    "LC50",
    "ED50",
    "% Inhibition",
    "Zone of Inhibition",
]
DEFAULT_POTENCY_UNITS = [
    "uM",
    "nM",
    "ug/mL",
    "mg/mL",
    "%",
    "mm",
]
CURATION_STATUS_OPTIONS = ["curated", "reviewed", "imported"]

SEARCH_FIELD_COLUMN_MAP = {
    "All searchable fields": [
        "trivial_name",
        "iupac_name",
        "molecular_formula",
        "sample_code",
        "source_category",
        "source_organism",
        "source_material",
        "collection_location",
        "compound_class",
        "compound_subclass",
        "journal_name",
        "article_title",
        "doi",
        "inchikey",
        "note",
    ],
    "Trivial Name": ["trivial_name"],
    "IUPAC Name": ["iupac_name"],
    "Molecular Formula": ["molecular_formula"],
    "Sample Code": ["sample_code"],
    "DOI": ["doi"],
    "Journal Name": ["journal_name"],
    "Article Title": ["article_title"],
    "Source Organism": ["source_organism"],
    "Source Category": ["source_category"],
    "InChIKey": ["inchikey"],
    "Compound Class": ["compound_class", "compound_subclass"],
}

PEAK_UPLOAD_TYPES = ["txt", "csv", "tsv", "dx", "jdx", "jcamp"]

NAV_OPTIONS = [
    "Dashboard",
    "Search & Match",
    "Compound Workspace",
    "Bioactivity",
    "1H Peaks",
    "13C Peaks",
    "Spectra Library",
    "Guide",
]

LEGACY_NAV_MAP = {
    "Overview": "Dashboard",
    "Search": "Search & Match",
    "Compound": "Compound Workspace",
    "1H NMR": "1H Peaks",
    "13C NMR": "13C Peaks",
    "Spectra": "Spectra Library",
}

COMPOUND_PAGE_OPTIONS = [
    "Browse Record",
    "New Submission",
    "Batch Import",
    "Update Metadata",
    "Delete Record",
]

LEGACY_COMPOUND_PAGE_MAP = {
    "Compound Detail": "Browse Record",
    "Record Detail": "Browse Record",
    "Add Compound": "New Submission",
    "Add Record": "New Submission",
    "Edit Compound": "Update Metadata",
    "Metadata Editor": "Update Metadata",
    "Delete Compound": "Delete Record",
}

NAV_SECTION_COPY = {
    "Dashboard": {
        "title": "Dashboard",
        "summary": "Overview and backup.",
    },
    "Search & Match": {
        "title": "Search & Match",
        "summary": "Keyword lookup and NMR matching.",
    },
    "Compound Workspace": {
        "title": "Compound Workspace",
        "summary": "Browse, submit, import, and revise records.",
    },
    "Bioactivity": {
        "title": "Bioactivity",
        "summary": "Track assay outcomes, targets, potency values, and literature-reported activity profiles.",
    },
    "1H Peaks": {
        "title": "1H Peaks",
        "summary": "Manage proton peak assignments.",
    },
    "13C Peaks": {
        "title": "13C Peaks",
        "summary": "Manage carbon shift assignments.",
    },
    "Spectra Library": {
        "title": "Spectra Library",
        "summary": "Manage spectra previews, files, and raw-data links.",
    },
    "Guide": {
        "title": "Guide",
        "summary": "Usage guide, submission rules, storage, and access notes.",
    },
}

COMPOUND_IMPORT_COLUMNS = [
    "trivial_name",
    "iupac_name",
    "molecular_formula",
    "molecular_weight",
    "smiles",
    "inchi",
    "inchikey",
    "compound_class",
    "compound_subclass",
    "source_category",
    "source_organism",
    "source_material",
    "sample_code",
    "collection_location",
    "gps_coordinates",
    "depth_m",
    "uv_data",
    "ftir_data",
    "cd_data",
    "optical_rotation",
    "melting_point",
    "crystallization_method",
    "structure_image_path",
    "journal_name",
    "article_title",
    "publication_year",
    "volume",
    "issue",
    "pages",
    "doi",
    "ccdc_number",
    "hrms_data",
    "data_source",
    "curation_status",
    "note",
]

PROTON_IMPORT_COLUMNS = [
    "compound_id",
    "compound_name",
    "delta_ppm",
    "multiplicity",
    "j_value",
    "proton_count",
    "assignment",
    "solvent",
    "instrument_mhz",
    "note",
]

CARBON_IMPORT_COLUMNS = [
    "compound_id",
    "compound_name",
    "delta_ppm",
    "carbon_type",
    "assignment",
    "solvent",
    "instrument_mhz",
    "note",
]

SPECTRA_IMPORT_COLUMNS = [
    "compound_id",
    "compound_name",
    "spectrum_type",
    "file_path",
    "note",
]

if FAVICON_PATH.exists() and FAVICON_PATH.stat().st_size <= MAX_PAGE_ICON_BYTES:
    st.set_page_config(
        page_title="Natural Products Spectral Database",
        page_icon=str(FAVICON_PATH),
        layout="wide"
    )
else:
    st.set_page_config(
        page_title="Natural Products Spectral Database",
        page_icon="🧬",
        layout="wide"
    )


def get_secret_setting(*keys: str) -> str:
    for key in keys:
        value = os.environ.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
        try:
            secret_value = st.secrets.get(key)
        except Exception:
            secret_value = None
        if secret_value is not None and str(secret_value).strip():
            return str(secret_value).strip()
    return ""


def cloud_backend_is_configured() -> bool:
    return bool(
        get_secret_setting("SUPABASE_URL")
        and (
            get_secret_setting("SUPABASE_SERVICE_ROLE_KEY")
            or get_secret_setting("SUPABASE_SECRET_KEY")
            or get_secret_setting("SUPABASE_ANON_KEY")
        )
    )


def should_initialize_sqlite_schema() -> bool:
    backend = get_secret_setting("NPDB_READ_BACKEND", "npdb_read_backend").strip().lower()
    if backend in {"local", "sqlite", "desktop"}:
        return True
    if backend in {"supabase", "cloud", "remote"}:
        return False
    return not cloud_backend_is_configured()


def get_secret_object(*keys: str):
    for key in keys:
        value = os.environ.get(key)
        if value is not None and str(value).strip():
            try:
                return json.loads(str(value))
            except json.JSONDecodeError:
                continue
        try:
            secret_value = st.secrets.get(key)
        except Exception:
            secret_value = None
        if secret_value:
            return secret_value
    return None


def allow_plaintext_password_secrets() -> bool:
    value = get_secret_setting("NPDB_ALLOW_PLAINTEXT_PASSWORD_SECRETS", "allow_plaintext_password_secrets")
    return value.strip().lower() in {"1", "true", "yes", "on"}


def verify_password_secret(submitted_password: str, plain_secret: str = "", hashed_secret: str = "") -> bool:
    submitted_password = "" if submitted_password is None else str(submitted_password)
    hashed_secret = str(hashed_secret or "").strip()
    if hashed_secret:
        try:
            scheme, iterations_text, salt, expected_digest = hashed_secret.split("$", 3)
            if scheme != PASSWORD_HASH_SCHEME:
                return False
            iterations = int(iterations_text)
            digest = hashlib.pbkdf2_hmac(
                "sha256",
                submitted_password.encode("utf-8"),
                salt.encode("utf-8"),
                iterations,
            )
            candidate_digest = base64.urlsafe_b64encode(digest).decode("ascii").rstrip("=")
            return hmac.compare_digest(candidate_digest, expected_digest)
        except Exception:
            return False

    plain_secret = str(plain_secret or "")
    return bool(plain_secret) and allow_plaintext_password_secrets() and hmac.compare_digest(submitted_password, plain_secret)


def load_approved_users() -> list[dict[str, str]]:
    raw_users = get_secret_object("NPDB_APPROVED_USERS", "approved_users")
    if isinstance(raw_users, dict):
        iterable = []
        for username, value in raw_users.items():
            if isinstance(value, dict):
                item = dict(value)
                item.setdefault("username", username)
                iterable.append(item)
            else:
                iterable.append({"username": username, "password": value})
    elif isinstance(raw_users, list):
        iterable = raw_users
    else:
        iterable = []

    users = []
    for item in iterable:
        if not isinstance(item, dict):
            continue
        username = str(item.get("username", "")).strip()
        password = str(item.get("password", "")).strip()
        password_hash = str(item.get("password_hash", "")).strip()
        role = str(item.get("role", "viewer")).strip() or "viewer"
        if username and (password_hash or (password and allow_plaintext_password_secrets())):
            users.append(
                {
                    "username": username,
                    "password": password,
                    "password_hash": password_hash,
                    "role": role,
                }
            )
    return users


def load_approved_names() -> list[str]:
    raw_names = get_secret_object("NPDB_APPROVED_NAMES", "approved_names")
    if not isinstance(raw_names, list):
        return []
    names = []
    for item in raw_names:
        text = str(item).strip() if item is not None else ""
        if text:
            names.append(text)
    return names


def normalize_login_slug(value: str) -> str:
    text = str(value).strip().lower() if value is not None else ""
    return re.sub(r"[^a-z0-9]+", "", text)


@st.cache_data(show_spinner=False)
def optimized_image_data_uri(path_value: str, max_px: int = 1200, quality: int = 82) -> str:
    path = Path(path_value)
    if not path.exists():
        return ""
    mime_type = mimetypes.guess_type(path.name)[0] or "image/png"
    raster_suffixes = {".png", ".jpg", ".jpeg", ".webp"}
    if Image is not None and path.suffix.lower() in raster_suffixes:
        try:
            with Image.open(path) as raw_image:
                image = ImageOps.exif_transpose(raw_image) if ImageOps is not None else raw_image.copy()
                image = image.copy()
                if max_px and max(image.size) > max_px:
                    image.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)

                has_alpha = image.mode in {"RGBA", "LA"} or "transparency" in image.info
                output = io.BytesIO()
                if has_alpha:
                    image.save(output, format="PNG", optimize=True)
                    mime_type = "image/png"
                else:
                    image.convert("RGB").save(
                        output,
                        format="JPEG",
                        quality=quality,
                        optimize=True,
                        progressive=True,
                    )
                    mime_type = "image/jpeg"
                encoded = base64.b64encode(output.getvalue()).decode("ascii")
                return f"data:{mime_type};base64,{encoded}"
        except Exception:
            pass
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def inline_asset_data_uri(path: Path, max_px: int = 1200) -> str:
    if not path.exists():
        return ""
    return optimized_image_data_uri(str(path), max_px=max_px)


def inline_svg_data_uri(svg_markup: str) -> str:
    encoded = base64.b64encode(svg_markup.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{encoded}"


def _normalize_html_block(markup: str) -> str:
    if not isinstance(markup, str):
        return markup
    return textwrap.dedent(markup).strip()

def is_access_gate_enabled() -> bool:
    plaintext_allowed = allow_plaintext_password_secrets()
    return bool(
        (plaintext_allowed and get_secret_setting("NPDB_ACCESS_PASSWORD", "access_password"))
        or get_secret_setting("NPDB_ACCESS_PASSWORD_HASH", "access_password_hash")
        or (plaintext_allowed and get_secret_setting("NPDB_APPROVED_PASSWORD", "approved_password"))
        or get_secret_setting("NPDB_APPROVED_PASSWORD_HASH", "approved_password_hash")
        or load_approved_users()
        or load_approved_names()
    )


def verify_access_gate():
    if not is_access_gate_enabled():
        return

    if st.session_state.get("npdb_authenticated"):
        return

    expected_username = get_secret_setting("NPDB_ACCESS_USERNAME", "access_username")
    expected_password = get_secret_setting("NPDB_ACCESS_PASSWORD", "access_password")
    expected_password_hash = get_secret_setting("NPDB_ACCESS_PASSWORD_HASH", "access_password_hash")
    approved_password = get_secret_setting("NPDB_APPROVED_PASSWORD", "approved_password")
    approved_password_hash = get_secret_setting("NPDB_APPROVED_PASSWORD_HASH", "approved_password_hash")
    approved_users = load_approved_users()
    approved_names = load_approved_names()
    login_background_uri = inline_asset_data_uri(LOGIN_BACKGROUND_PATH, max_px=1600)
    login_logo_uri = inline_asset_data_uri(SIDEBAR_LOGO_PATH if SIDEBAR_LOGO_PATH.exists() else LOGIN_LOGO_PATH, max_px=180)
    login_left_art_uri = inline_asset_data_uri(LOGIN_LEFT_ART_PATH, max_px=560)
    login_right_art_uri = inline_asset_data_uri(LOGIN_RIGHT_ART_PATH, max_px=560)
    background_style = (
        f"background-image: linear-gradient(118deg, rgba(2, 8, 20, 0.90), rgba(3, 7, 19, 0.78)), url('{login_background_uri}');"
        if login_background_uri
        else "background: radial-gradient(circle at 20% 30%, rgba(0, 170, 255, 0.16), transparent 30%), radial-gradient(circle at 80% 32%, rgba(182, 72, 255, 0.16), transparent 28%), linear-gradient(180deg, rgba(3, 8, 22, 1), rgba(4, 8, 18, 1));"
    )
    use_login_brand_lockup = False
    logo_class = "login-brand-logo"
    logo_markup = f'<img class="{logo_class}" src="{login_logo_uri}" alt="NPDB logo" />' if login_logo_uri else ""
    left_art_markup = f'<img class="login-ambient-art login-ambient-art-left" src="{login_left_art_uri}" alt="Natural product structure" />' if login_left_art_uri else ""
    right_art_markup = f'<img class="login-ambient-art login-ambient-art-right" src="{login_right_art_uri}" alt="Natural product structure" />' if login_right_art_uri else ""
    shield_markup = """
        <svg viewBox="0 0 64 64" aria-hidden="true" style="width:42px;height:42px;display:block">
            <defs>
                <linearGradient id="npdbShield" x1="0%" y1="0%" x2="100%" y2="100%">
                    <stop offset="0%" stop-color="#18E5FF"/>
                    <stop offset="55%" stop-color="#4C8EFF"/>
                    <stop offset="100%" stop-color="#B34EFF"/>
                </linearGradient>
            </defs>
            <path d="M32 6l18 6v14c0 12.3-7.1 23.7-18 29-10.9-5.3-18-16.7-18-29V12l18-6z" fill="none" stroke="url(#npdbShield)" stroke-width="2.6"/>
            <path d="M24 31.5l5.8 5.8L41.8 24.8" fill="none" stroke="#DBF8FF" stroke-width="3.2" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
    """

    st.markdown(
        _normalize_html_block(
            f"""
            <style>
            [data-testid="stAppViewContainer"] {{
                {background_style}
                background-size: cover;
                background-position: center;
                background-attachment: fixed;
            }}
            [data-testid="stHeader"] {{
                background: rgba(4, 10, 24, 0.06);
            }}
            [data-testid="stToolbar"] {{
                display: none;
            }}
            [data-testid="stSidebar"] {{
                display: none;
            }}
            .block-container {{
                max-width: none !important;
                padding-top: 0 !important;
                padding-bottom: 0.6rem !important;
            }}
            .login-shell {{
                position: relative;
                min-height: 0;
                height: clamp(132px, 22vh, 190px);
                padding: 0.35rem 0 0 0;
                overflow: visible;
            }}
            .login-shell::before {{
                content: "";
                position: absolute;
                inset: 0;
                background:
                    radial-gradient(circle at 18% 70%, rgba(33, 203, 255, 0.16), transparent 24%),
                    radial-gradient(circle at 82% 70%, rgba(196, 52, 255, 0.18), transparent 24%);
                pointer-events: none;
            }}
            .login-topbar {{
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 1rem;
                max-width: 1060px;
                margin: 0 auto 0.05rem auto;
                padding: 0 1.5rem;
            }}
            .login-brand {{
                display: flex;
                align-items: flex-start;
                gap: 0.9rem;
            }}
            .login-brand-logo {{
                width: 58px;
                height: 58px;
                object-fit: contain;
                filter: drop-shadow(0 12px 24px rgba(0,0,0,0.24));
                flex: 0 0 auto;
            }}
            .login-brand-logo.is-lockup {{
                width: min(40vw, 540px);
                height: auto;
            }}
            .login-brand-copy {{
                padding-top: 0.1rem;
            }}
            .login-brand-title {{
                color: #F7FBFF;
                font-size: clamp(1.02rem, 1.7vw, 1.5rem);
                line-height: 1.05;
                font-weight: 780;
                letter-spacing: -0.05em;
                margin-bottom: 0.12rem;
            }}
            .login-brand-title .accent {{
                background: linear-gradient(90deg, #18E5FF, #3B9BFF 43%, #8F69FF 74%, #DA58FF 100%);
                -webkit-background-clip: text;
                background-clip: text;
                color: transparent;
            }}
            .login-brand-kicker {{
                color: rgba(215, 227, 243, 0.72);
                font-size: 0.58rem;
                letter-spacing: 0.16em;
                text-transform: uppercase;
            }}
            .login-help {{
                display: inline-flex;
                align-items: center;
                gap: 0.6rem;
                color: rgba(241, 246, 255, 0.94);
                font-size: 0.84rem;
                white-space: nowrap;
                padding-top: 0.18rem;
                text-decoration: none;
            }}
            .login-help-badge {{
                width: 1.62rem;
                height: 1.62rem;
                border-radius: 999px;
                display: inline-grid;
                place-items: center;
                border: 1px solid rgba(255,255,255,0.36);
                background: rgba(255,255,255,0.04);
                font-size: 0.88rem;
            }}
            .login-center {{
                position: relative;
                max-width: 1060px;
                min-height: 54px;
                margin: 0 auto;
                padding: 0 1.5rem;
                z-index: 1;
            }}
            .login-ambient-art {{
                position: absolute;
                pointer-events: none;
                opacity: 0.17;
                filter: drop-shadow(0 24px 40px rgba(20, 32, 80, 0.24));
            }}
            .login-ambient-art-left {{
                left: -3rem;
                bottom: 0.2rem;
                width: min(26vw, 360px);
            }}
            .login-ambient-art-right {{
                right: -2rem;
                top: -0.2rem;
                width: min(24vw, 320px);
            }}
            .login-card-anchor {{
                width: min(100%, 560px);
                height: 8px;
                margin: 0 auto;
                border-radius: 28px;
                border: 1px solid rgba(115, 79, 255, 0.44);
                background: linear-gradient(145deg, rgba(6, 20, 38, 0.24), rgba(18, 19, 42, 0.32));
                box-shadow: 0 18px 44px rgba(0, 0, 0, 0.14);
                opacity: 0.12;
            }}
            .login-wave {{
                position: absolute;
                pointer-events: none;
                opacity: 0.66;
                mix-blend-mode: screen;
            }}
            .login-wave-left {{
                left: -4vw;
                bottom: -0.2rem;
                width: min(46vw, 650px);
                height: 120px;
                background:
                    radial-gradient(circle at 22% 72%, rgba(46, 218, 255, 0.85), transparent 8%),
                    radial-gradient(circle at 44% 44%, rgba(45, 162, 255, 0.88), transparent 7%),
                    radial-gradient(circle at 64% 64%, rgba(0, 119, 255, 0.82), transparent 7%);
                filter: blur(1px) drop-shadow(0 0 12px rgba(31, 189, 255, 0.46));
                border-bottom: 2px solid rgba(47, 205, 255, 0.9);
                clip-path: polygon(0 78%, 8% 76%, 12% 68%, 18% 72%, 23% 46%, 29% 78%, 36% 74%, 43% 32%, 49% 86%, 57% 82%, 65% 55%, 72% 88%, 80% 84%, 88% 79%, 100% 80%, 100% 100%, 0 100%);
            }}
            .login-wave-right {{
                right: -3vw;
                bottom: -0.2rem;
                width: min(42vw, 590px);
                height: 120px;
                background:
                    radial-gradient(circle at 22% 68%, rgba(203, 57, 255, 0.86), transparent 8%),
                    radial-gradient(circle at 48% 48%, rgba(255, 116, 244, 0.84), transparent 7%),
                    radial-gradient(circle at 68% 70%, rgba(138, 69, 255, 0.84), transparent 7%);
                filter: blur(1px) drop-shadow(0 0 14px rgba(215, 79, 255, 0.42));
                border-bottom: 2px solid rgba(213, 84, 255, 0.88);
                clip-path: polygon(0 80%, 10% 78%, 16% 48%, 23% 86%, 31% 74%, 39% 84%, 45% 62%, 52% 24%, 58% 84%, 66% 68%, 73% 82%, 80% 62%, 88% 74%, 94% 80%, 100% 82%, 100% 100%, 0 100%);
            }}
            div[data-testid="stForm"] {{
                width: min(100%, 560px) !important;
                margin: -1.05rem auto 0 auto !important;
                padding: 0.84rem 1rem 0.68rem 1rem !important;
                border-radius: 26px !important;
                background: linear-gradient(145deg, rgba(6, 20, 38, 0.82), rgba(18, 19, 42, 0.86)) !important;
                border: 1px solid rgba(94, 222, 255, 0.34) !important;
                box-shadow: 0 22px 54px rgba(0, 0, 0, 0.26) !important;
                backdrop-filter: blur(14px) !important;
            }}
            div[data-testid="stForm"] > div:first-child {{
                border: none !important;
                background: transparent !important;
                box-shadow: none !important;
                padding: 0 !important;
            }}
            .auth-card-badge {{
                width: 58px;
                height: 58px;
                margin: 0 auto 0.34rem auto;
                border-radius: 999px;
                display: grid;
                place-items: center;
                background: linear-gradient(180deg, rgba(18, 44, 82, 0.82), rgba(44, 19, 78, 0.76));
                border: 1px solid rgba(97, 216, 237, 0.34);
                box-shadow: inset 0 1px 0 rgba(255,255,255,0.05);
                overflow: hidden;
            }}
            .auth-badge-logo {{
                width: 42px;
                height: 42px;
                object-fit: contain;
                display: block;
                filter: drop-shadow(0 6px 18px rgba(0, 0, 0, 0.28));
            }}
            .auth-title {{
                color: #F5F8FD;
                font-size: clamp(1.45rem, 2vw, 2rem);
                text-align: center;
                font-weight: 800;
                letter-spacing: -0.04em;
                margin-bottom: 0.06rem;
            }}
            .auth-subtitle {{
                color: rgba(214, 223, 238, 0.86);
                line-height: 1.34;
                text-align: center;
                font-size: 0.82rem;
                margin: 0 auto 0.48rem auto;
                max-width: 19rem;
            }}
            .auth-field-shell {{
                position: relative;
                margin-bottom: 0.18rem;
            }}
            div[data-testid="stForm"] label p {{
                color: #F7FBFF !important;
                font-size: 0.88rem !important;
                font-weight: 620 !important;
            }}
            div[data-testid="stForm"] div[data-testid="stTextInput"] input {{
                min-height: 42px !important;
                border-radius: 14px !important;
                border: 1px solid rgba(150, 168, 206, 0.28) !important;
                background: rgba(19, 27, 48, 0.82) !important;
                color: #F7FBFF !important;
                font-size: 0.9rem !important;
                padding-left: 1rem !important;
            }}
            div[data-testid="stForm"] div[data-testid="stTextInput"] input::placeholder {{
                color: rgba(180, 192, 214, 0.62) !important;
            }}
            div[data-testid="stForm"] div[data-testid="stCheckbox"] label {{
                color: #DCE6F5 !important;
                font-size: 0.84rem !important;
            }}
            div[data-testid="stForm"] div[data-testid="stCheckbox"] input + div {{
                border-radius: 6px !important;
            }}
            div[data-testid="stFormSubmitButton"] button {{
                min-height: 48px !important;
                border-radius: 15px !important;
                border: 1px solid rgba(122, 89, 255, 0.28) !important;
                background: linear-gradient(90deg, #1EAFFF, #2D7FFF 42%, #8246FF 73%, #CB3EFF 100%) !important;
                color: #F8FBFF !important;
                font-size: 0.94rem !important;
                font-weight: 720 !important;
                box-shadow: 0 18px 34px rgba(92, 85, 255, 0.24) !important;
            }}
            .auth-meta {{
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 1rem;
                margin: 0.01rem 0 0.2rem 0;
            }}
            .auth-forgot-link {{
                color: #9FA6FF;
                font-size: 0.82rem;
                text-decoration: none;
            }}
            .auth-meta-note {{
                color: rgba(165, 183, 222, 0.88);
                font-size: 0.78rem;
            }}
            .auth-help-panel {{
                margin: 0.2rem auto 0 auto;
                max-width: 560px;
                padding: 0.58rem 0.82rem;
                border-radius: 18px;
                border: 1px solid rgba(255,255,255,0.08);
                background: linear-gradient(180deg, rgba(9, 16, 34, 0.72), rgba(14, 18, 42, 0.7));
                box-shadow: 0 20px 50px rgba(0, 0, 0, 0.22);
                backdrop-filter: blur(10px);
            }}
            .auth-help-title {{
                color: #F5F8FD;
                font-size: 0.88rem;
                font-weight: 730;
                margin-bottom: 0.18rem;
            }}
            .auth-help-copy {{
                color: rgba(210, 221, 239, 0.84);
                font-size: 0.76rem;
                line-height: 1.34;
            }}
            .auth-help-copy strong {{
                color: #F6FBFF;
            }}
            .auth-footer {{
                margin-top: 0.2rem;
                color: rgba(169, 183, 208, 0.84);
                font-size: 0.72rem;
                text-align: center;
            }}
            @media (max-width: 900px) {{
                .login-topbar {{
                    flex-direction: column;
                }}
                .login-brand-logo {{
                    width: 80px;
                    height: 80px;
                }}
                .login-brand-logo.is-lockup {{
                    width: min(78vw, 460px);
                    height: auto;
                }}
                .login-ambient-art {{
                    display: none;
                }}
                .login-wave {{
                    opacity: 0.52;
                }}
                .login-card-anchor {{
                    height: 24px;
                }}
                div[data-testid="stForm"] {{
                    margin-top: -1.2rem !important;
                }}
            }}
            @media (max-width: 640px) {{
                .login-shell {{
                    padding-top: 0.8rem;
                }}
                .login-topbar,
                .login-center {{
                    padding-left: 1rem;
                    padding-right: 1rem;
                }}
                .login-card-anchor {{
                    height: 20px;
                    margin-top: 0.1rem;
                }}
                div[data-testid="stForm"] {{
                    margin-top: -1rem !important;
                    padding: 0.78rem 0.85rem 0.74rem 0.85rem !important;
                }}
                .auth-meta {{
                    flex-direction: column;
                    align-items: flex-start;
                }}
            }}
            </style>
            <div class="login-shell">
                <div class="login-topbar">
                    <div class="login-brand">
                        {logo_markup}
                {"" if use_login_brand_lockup else '<div class="login-brand-copy"><div class="login-brand-title">Natural Products<br><span class="accent">Spectral Database</span></div><div class="login-brand-kicker">Explore · Analyze · Discover</div></div>'}
                    </div>
                    <a class="login-help" href="#login-access-help">
                        <span class="login-help-badge">?</span>
                        <span>Help</span>
                    </a>
                </div>
                <div class="login-center">
                    {left_art_markup}
                    {right_art_markup}
                    <div class="login-wave login-wave-left"></div>
                    <div class="login-wave login-wave-right"></div>
                    <div class="login-card-anchor"></div>
                </div>
            </div>
            """
        ),
        unsafe_allow_html=True,
    )

    locked_until = float(st.session_state.get("npdb_login_locked_until", 0) or 0)
    login_locked = locked_until > time.time()
    if login_locked:
        remaining_seconds = max(1, int(locked_until - time.time()))
        st.warning(f"Too many failed login attempts. Please try again in {remaining_seconds} second(s).")

    with st.form("npdb_access_gate"):
        badge_uri = inline_asset_data_uri(LOGIN_BADGE_PATH, max_px=120)
        badge_markup = (
            f'<img class="auth-badge-logo" src="{badge_uri}" alt="NPDB secure coral badge" />'
            if badge_uri
            else shield_markup
        )
        st.markdown(f'<div class="auth-card-badge">{badge_markup}</div>', unsafe_allow_html=True)
        st.markdown('<div class="auth-title">Welcome Back</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="auth-subtitle">This workspace is protected.<br>Please enter your credentials to access the database.</div>',
            unsafe_allow_html=True,
        )
        username = st.text_input("Username", value="", placeholder="Enter your username", icon=":material/person:")
        password = st.text_input("Password", value="", type="password", placeholder="Enter your password", icon=":material/lock:")
        remember_col, forgot_col = st.columns([1.1, 1])
        with remember_col:
            remember_me = st.checkbox("Remember me", value=False)
        with forgot_col:
            st.markdown(
                '<div style="text-align:right; padding-top: 1.85rem;"><a class="auth-forgot-link" href="#login-access-help">Forgot password?</a></div>',
                unsafe_allow_html=True,
        )
        st.markdown('<div class="auth-meta"><div class="auth-meta-note">Approved users only. Contact the database owner if you need access.</div></div>', unsafe_allow_html=True)
        submitted = st.form_submit_button("Open Database  →", width="stretch", disabled=login_locked)
        st.markdown(
            '<div class="auth-footer">This access gate protects the curated NPDB workspace.</div>',
            unsafe_allow_html=True,
        )

    st.markdown('<div id="login-access-help"></div>', unsafe_allow_html=True)
    with st.expander("Access Guide", expanded=False):
        st.markdown(
            """
            <div class="auth-help-copy">
                Sign in with an approved NPDB account. If your workspace uses the approved-name pattern, enter the username as
                <strong>npdb_yourname</strong>. Only the owner account can edit or submit records; other approved users can browse,
                search, and download curated data.
            </div>
            """,
            unsafe_allow_html=True,
        )

    if submitted:
        authenticated = False
        matched_role = "viewer"
        matched_auth_mode = ""

        if approved_users:
            for user in approved_users:
                username_ok = hmac.compare_digest(username.strip(), user["username"])
                password_ok = verify_password_secret(
                    password,
                    plain_secret=user.get("password", ""),
                    hashed_secret=user.get("password_hash", ""),
                )
                if username_ok and password_ok:
                    authenticated = True
                    matched_role = user.get("role", "viewer")
                    matched_auth_mode = "approved_users"
                    break
        elif approved_names and approved_password:
            submitted_username = str(username).strip() if username is not None else ""
            if submitted_username.lower().startswith("npdb_"):
                submitted_name = submitted_username[5:]
                submitted_slug = normalize_login_slug(submitted_name)
                allowed_slugs = {normalize_login_slug(name) for name in approved_names}
                if submitted_slug in allowed_slugs and verify_password_secret(
                    password,
                    plain_secret=approved_password,
                    hashed_secret=approved_password_hash,
                ):
                    authenticated = True
                    matched_role = "approved-viewer"
                    matched_auth_mode = "approved_names"
        else:
            username_ok = True if not expected_username else hmac.compare_digest(username.strip(), expected_username)
            password_ok = verify_password_secret(
                password,
                plain_secret=expected_password,
                hashed_secret=expected_password_hash,
            )
            authenticated = username_ok and password_ok
            if authenticated:
                matched_auth_mode = "single_user"

        if authenticated:
            st.session_state["npdb_authenticated"] = True
            st.session_state["npdb_username"] = username.strip()
            st.session_state["npdb_role"] = matched_role
            st.session_state["npdb_auth_mode"] = matched_auth_mode
            st.session_state["npdb_remember_requested"] = bool(remember_me)
            st.session_state["npdb_login_failures"] = 0
            st.session_state["npdb_login_locked_until"] = 0.0
            st.rerun()
        failures = int(st.session_state.get("npdb_login_failures", 0) or 0) + 1
        st.session_state["npdb_login_failures"] = failures
        if failures >= 5:
            lock_seconds = min(300, 30 * (2 ** min(failures - 5, 3)))
            st.session_state["npdb_login_locked_until"] = time.time() + lock_seconds
            st.error(f"Access denied. Too many failed attempts; login is paused for {lock_seconds} seconds.")
        else:
            st.error("Access denied. Please check the approved credentials.")

    st.stop()

verify_access_gate()


def is_owner_editor() -> bool:
    current_username = normalize_login_slug(st.session_state.get("npdb_username", ""))
    if current_username != normalize_login_slug(OWNER_EDITOR_USERNAME):
        return False

    role = str(st.session_state.get("npdb_role", "")).strip().lower()
    if role in OWNER_EDITOR_ROLES:
        return True

    auth_mode = str(st.session_state.get("npdb_auth_mode", "")).strip().lower()
    expected_username = get_secret_setting("NPDB_ACCESS_USERNAME", "access_username")
    return (
        auth_mode == "single_user"
        and bool(expected_username)
        and normalize_login_slug(expected_username) == normalize_login_slug(OWNER_EDITOR_USERNAME)
    )


def can_edit_database() -> bool:
    if not is_owner_editor():
        return False
    if use_supabase_backend() and not use_supabase_write_backend():
        return False
    return True


def cloud_write_is_blocked() -> bool:
    return is_owner_editor() and use_supabase_backend() and not use_supabase_write_backend()


def render_read_only_notice(feature_label: str):
    if cloud_write_is_blocked():
        st.warning(
            "Cloud write mode is not active in this deployment yet. "
            "To prevent data divergence, editing is temporarily disabled until "
            "`SUPABASE_SERVICE_ROLE_KEY` or `SUPABASE_SECRET_KEY` is configured in secure server-side secrets."
        )
        return
    st.info(
        f"Read-only access. Only `{OWNER_EDITOR_USERNAME}` can {feature_label}. "
        "Other approved users can still browse records, search structures, and review spectra."
    )


def clear_structure_search_state():
    st.session_state["structure_search_results"] = []
    st.session_state["structure_search_error"] = ""
    st.session_state["structure_search_mode_label"] = ""
    st.session_state["structure_search_attempted"] = False

# =========================
# Session state defaults
# =========================
if "nav_section" not in st.session_state:
    st.session_state["nav_section"] = "Dashboard"
elif st.session_state["nav_section"] in LEGACY_NAV_MAP:
    st.session_state["nav_section"] = LEGACY_NAV_MAP[st.session_state["nav_section"]]

if "selected_compound_id" not in st.session_state:
    st.session_state["selected_compound_id"] = None

if "compound_page" not in st.session_state:
    st.session_state["compound_page"] = "Browse Record"
elif st.session_state["compound_page"] in LEGACY_COMPOUND_PAGE_MAP:
    st.session_state["compound_page"] = LEGACY_COMPOUND_PAGE_MAP[st.session_state["compound_page"]]

if "compound_wizard_step" not in st.session_state:
    st.session_state["compound_wizard_step"] = 1

# Pending widget-state sync helpers.
# These avoid Streamlit errors when navigation is changed from buttons
# after a radio widget has already been instantiated in the same run.
if "_pending_main_section_radio" in st.session_state:
    st.session_state["main_section_radio"] = st.session_state.pop("_pending_main_section_radio")
elif "main_section_radio" not in st.session_state:
    st.session_state["main_section_radio"] = st.session_state["nav_section"]

if "_pending_compound_page_radio" in st.session_state:
    st.session_state["compound_page_radio"] = st.session_state.pop("_pending_compound_page_radio")
elif "compound_page_radio" not in st.session_state:
    st.session_state["compound_page_radio"] = st.session_state["compound_page"]

# =========================
# Navigation helpers
# =========================
def set_main_nav(section: str):
    st.session_state["nav_section"] = section
    st.session_state["_pending_main_section_radio"] = section

def set_compound_page(page_name: str):
    st.session_state["compound_page"] = page_name
    st.session_state["_pending_compound_page_radio"] = page_name


def build_internal_nav_href(section: str, compound_page: str | None = None) -> str:
    params = {"nav": section}
    if compound_page:
        params["compound_page"] = compound_page
    return f"?{urlencode(params)}"


def navigate_internal(section: str, compound_page: str | None = None):
    set_main_nav(section)
    if compound_page:
        set_compound_page(compound_page)
    try:
        st.query_params["nav"] = section
        if compound_page:
            st.query_params["compound_page"] = compound_page
        elif "compound_page" in st.query_params:
            del st.query_params["compound_page"]
    except Exception:
        pass


def apply_navigation_query_params():
    query_params = st.query_params
    target_section = query_params.get("nav")
    target_compound_page = query_params.get("compound_page")

    if isinstance(target_section, list):
        target_section = target_section[0] if target_section else None
    if isinstance(target_compound_page, list):
        target_compound_page = target_compound_page[0] if target_compound_page else None

    if target_section:
        set_main_nav(str(target_section))
        if target_compound_page:
            set_compound_page(str(target_compound_page))
        try:
            st.query_params.clear()
        except Exception:
            pass

def open_compound_detail(compound_id: int):
    st.session_state["selected_compound_id"] = int(compound_id)
    set_main_nav("Compound Workspace")
    set_compound_page("Browse Record")

def open_compound_editor(compound_id: int):
    st.session_state["selected_compound_id"] = int(compound_id)
    set_main_nav("Compound Workspace")
    set_compound_page("Update Metadata")

# =========================
# Custom styling
# =========================
st.markdown("""
<style>
:root {
    --bg-soft: rgba(255,255,255,0.028);
    --bg-soft-2: rgba(255,255,255,0.05);
    --bg-panel: rgba(12, 24, 40, 0.74);
    --bg-panel-strong: rgba(14, 27, 45, 0.9);
    --border-soft: rgba(255,255,255,0.10);
    --text-soft: #AEB8C6;
    --text-main: #F5F8FD;
    --text-strong: #FFFFFF;
    --accent-cyan: #61D8ED;
    --accent-blue: #4C8EFF;
    --accent-purple: #9C63F1;
    --accent-green: #7EF0C2;
    --accent-coral: #FF7F6D;
    --accent-gold: #F2C66D;
    --shadow-soft: 0 18px 44px rgba(0,0,0,0.22);
    --shadow-deep: 0 24px 60px rgba(0,0,0,0.34);
    --glow-soft: 0 0 0 1px rgba(255,255,255,0.04), 0 10px 30px rgba(97,216,237,0.05);
    --radius-card: 22px;
    --radius-pill: 999px;
}

[data-testid="stAppViewContainer"] {
    background:
        radial-gradient(circle at 14% 16%, rgba(97, 216, 237, 0.11), transparent 28%),
        radial-gradient(circle at 84% 12%, rgba(156, 99, 241, 0.14), transparent 30%),
        radial-gradient(circle at 62% 82%, rgba(255, 127, 109, 0.07), transparent 24%),
        linear-gradient(180deg, #06101c 0%, #081321 40%, #07111b 100%);
}

.block-container {
    padding-top: 1rem;
    padding-bottom: 4.25rem;
    max-width: 1360px;
}

[data-testid="stSidebar"] {
    border-right: 1px solid rgba(255,255,255,0.06);
    background:
        linear-gradient(180deg, rgba(8, 17, 30, 0.96), rgba(8, 14, 24, 0.98)) !important;
    min-width: 20rem !important;
    max-width: 20rem !important;
}

[data-testid="stSidebar"] .block-container {
    padding-top: 1.1rem;
    padding-left: 0.74rem;
    padding-right: 0.74rem;
}

.sidebar-note {
    border-radius: 18px;
    padding: 0.9rem 0.95rem;
    background: linear-gradient(180deg, rgba(17, 28, 46, 0.78), rgba(9, 18, 31, 0.82));
    border: 1px solid rgba(255,255,255,0.06);
    color: var(--text-soft);
    font-size: 0.88rem;
    line-height: 1.5;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
    margin-bottom: 0.95rem;
}

.sidebar-brand {
    padding: 0.2rem 0.1rem 0.45rem 0.1rem;
    margin-bottom: 0.95rem;
}

.sidebar-brand-head {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 0.72rem;
    text-align: center;
}

.sidebar-brand-logo-shell {
    width: 126px;
    height: 126px;
    border-radius: 28px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(0,0,0,0.18);
    border: 1px solid rgba(97,216,237,0.12);
    overflow: hidden;
    box-shadow: 0 18px 36px rgba(0,0,0,0.24), 0 0 0 1px rgba(156,99,241,0.05);
}

.sidebar-brand-logo {
    width: 118px;
    height: 118px;
    object-fit: contain;
    display: block;
}

.sidebar-brand-copy {
    min-width: 0;
}

.sidebar-brand-title {
    color: var(--text-main);
    font-size: 1.04rem;
    line-height: 1.23;
    font-weight: 760;
    letter-spacing: 0;
    text-align: center;
}

.sidebar-doc-cta {
    margin: 0.25rem 0 0.9rem 0;
}

.sidebar-nav-icon-shell {
    width: 2.95rem;
    height: 2.95rem;
    border-radius: 17px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: linear-gradient(180deg, rgba(19, 33, 55, 0.96), rgba(10, 18, 31, 0.98));
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.05), 0 10px 22px rgba(0,0,0,0.16);
    margin-top: 0.08rem;
}

.sidebar-nav-icon {
    width: 1.45rem;
    height: 1.45rem;
    object-fit: contain;
    display: block;
    filter: drop-shadow(0 0 12px rgba(97, 216, 237, 0.16));
}

.sidebar-nav-link {
    display: flex;
    align-items: center;
    gap: 0.78rem;
    min-height: 56px;
    padding: 0.55rem 0.82rem;
    margin-bottom: 0.28rem;
    border-radius: 18px;
    color: rgba(240, 246, 255, 0.92) !important;
    text-decoration: none !important;
    border: 1px solid transparent;
    background: transparent;
    transition: background 0.16s ease, border-color 0.16s ease, transform 0.16s ease;
}

.sidebar-nav-link:hover {
    background: linear-gradient(180deg, rgba(255,255,255,0.032), rgba(255,255,255,0.014));
    border-color: rgba(255,255,255,0.07);
    transform: translateY(-1px);
}

.sidebar-nav-link.is-active {
    background: linear-gradient(90deg, rgba(120, 66, 209, 0.94), rgba(48, 79, 167, 0.94));
    border-color: rgba(120, 176, 255, 0.2);
    box-shadow: 0 12px 24px rgba(0,0,0,0.16);
}

.sidebar-nav-link .sidebar-nav-icon-shell {
    width: 2.34rem;
    height: 2.34rem;
    border-radius: 13px;
    margin: 0;
    flex: 0 0 auto;
}

.sidebar-nav-link-label {
    display: block;
    min-width: 0;
    color: inherit;
    font-size: 1rem;
    line-height: 1.2;
    font-weight: 680;
    text-align: left;
}

.sidebar-stats {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 0.58rem;
    margin-bottom: 1rem;
}

.sidebar-stat {
    border-radius: 15px;
    padding: 0.58rem 0.62rem;
    background: linear-gradient(180deg, rgba(17, 28, 46, 0.78), rgba(10, 18, 31, 0.88));
    border: 1px solid rgba(255,255,255,0.06);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
    min-width: 0;
}

.sidebar-stat-head {
    display: flex;
    align-items: center;
    gap: 0.48rem;
    min-width: 0;
}

.sidebar-stat-icon {
    width: 20px;
    height: 20px;
    object-fit: contain;
    filter: drop-shadow(0 0 10px rgba(97, 216, 237, 0.08));
    flex: 0 0 auto;
}

.sidebar-stat-value {
    color: var(--text-main);
    font-size: 0.88rem;
    font-weight: 760;
    line-height: 1.1;
    min-width: 0;
    white-space: nowrap;
}

.sidebar-stat-label {
    color: var(--text-soft);
    font-size: 0.68rem;
    line-height: 1.2;
    margin-top: 0.25rem;
    overflow-wrap: normal;
}

.sidebar-meta-block {
    margin-top: 0.15rem;
    margin-bottom: 0.95rem;
}

.sidebar-meta-title {
    color: rgba(174,184,198,0.78);
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 0.28rem;
}

.sidebar-meta-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    color: var(--text-main);
    font-size: 0.96rem;
    font-weight: 640;
    margin-bottom: 0.55rem;
}

.sidebar-status-dot {
    width: 14px;
    height: 14px;
    border-radius: 999px;
    background: linear-gradient(180deg, #34e38b, #22c96d);
    box-shadow: 0 0 0 4px rgba(52, 227, 139, 0.08);
}

.sidebar-meta-divider {
    height: 1px;
    background: rgba(255,255,255,0.08);
    margin: 0.55rem 0 0.75rem 0;
}

.sidebar-quality-value {
    color: var(--text-main);
    font-size: 0.96rem;
    font-weight: 720;
    margin-bottom: 0.42rem;
}

.sidebar-quality-track {
    width: 100%;
    height: 12px;
    border-radius: 999px;
    background: rgba(93, 109, 141, 0.34);
    overflow: hidden;
}

.sidebar-quality-fill {
    height: 100%;
    border-radius: inherit;
    background: linear-gradient(90deg, #1ee4ef, #2ad59e);
}

.selector-card {
    border-radius: 20px;
    padding: 1rem 1.05rem;
    margin-bottom: 1rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.018));
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: var(--glow-soft);
}

.selector-title {
    color: var(--text-main);
    font-size: 0.98rem;
    font-weight: 720;
    margin-bottom: 0.18rem;
}

.selector-subtitle {
    color: var(--text-soft);
    font-size: 0.9rem;
    line-height: 1.5;
    margin-bottom: 0.75rem;
}

.inline-note {
    color: var(--text-soft);
    font-size: 0.92rem;
    line-height: 1.5;
}

.hero-shell {
    margin-top: 0.1rem;
    margin-bottom: 1.2rem;
}

.hero-banner-wrap {
    border-radius: 28px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.06);
    box-shadow: var(--shadow-deep);
}

.hero-image-fallback {
    border-radius: 20px;
    padding: 1.2rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.014));
    border: 1px solid rgba(255,255,255,0.08);
}

@media (max-width: 1100px) {
    .hero-shell {
        margin-bottom: 1rem;
    }
}
.section-title {
    margin-top: 0.2rem;
    margin-bottom: 0.28rem;
    font-size: 1.56rem;
    line-height: 1.16;
    font-weight: 800;
    letter-spacing: -0.03em;
    color: var(--text-strong);
    text-wrap: balance;
}

.app-credit-footer {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    margin: 0;
    padding: 0.34rem 0.82rem;
    border-radius: 999px;
    background: linear-gradient(180deg, rgba(11, 21, 34, 0.9), rgba(7, 15, 27, 0.88));
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: 0 8px 20px rgba(0,0,0,0.18);
    color: rgba(245, 248, 253, 0.92);
    font-size: 0.68rem;
    letter-spacing: 0.01em;
    backdrop-filter: blur(8px);
    white-space: nowrap;
}

.sidebar-credit-wrap {
    margin-top: 1rem;
    padding-left: 0.08rem;
    text-align: left;
}

.sidebar-credit-wrap .app-credit-footer {
    justify-content: flex-start;
}

.sidebar-session-summary {
    margin-top: 1rem;
    padding: 0.82rem 0.9rem 0.72rem 0.9rem;
    border-radius: 18px;
    background: linear-gradient(180deg, rgba(17, 28, 46, 0.72), rgba(9, 18, 31, 0.84));
    border: 1px solid rgba(255,255,255,0.07);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.035);
}

.sidebar-session-title {
    color: rgba(174,184,198,0.78);
    font-size: 0.7rem;
    font-weight: 760;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 0.28rem;
}

.sidebar-session-user {
    color: var(--text-main);
    font-size: 0.86rem;
    line-height: 1.35;
    font-weight: 650;
}

.section-subtitle {
    color: var(--text-soft);
    margin-bottom: 0.92rem;
    line-height: 1.55;
    max-width: 62rem;
    font-size: 0.93rem;
}

.metric-card {
    border-radius: var(--radius-card);
    padding: 0.92rem 0.98rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.042), rgba(255,255,255,0.02));
    border: 1px solid rgba(255,255,255,0.08);
    margin-bottom: 0.72rem;
    box-shadow: var(--glow-soft), var(--shadow-soft);
    min-height: 96px;
    transition: transform 0.18s ease, border-color 0.18s ease, box-shadow 0.18s ease;
}

.metric-card:hover,
.panel-card:hover,
.compound-card:hover,
.result-card:hover,
.helper-card:hover,
.kv-card:hover,
.structure-card:hover {
    transform: translateY(-1px);
    border-color: rgba(97,216,237,0.2);
    box-shadow: 0 18px 42px rgba(0,0,0,0.26), 0 0 0 1px rgba(97,216,237,0.05);
}

.metric-card-label {
    color: var(--text-soft);
    font-size: 0.8rem;
    font-weight: 580;
    margin-bottom: 0.38rem;
    line-height: 1.35;
}

.metric-card-value {
    font-size: 1.86rem;
    font-weight: 780;
    line-height: 1;
    letter-spacing: -0.03em;
    color: var(--text-strong);
}

.dashboard-section {
    margin-top: 0.36rem;
    margin-bottom: 0.92rem;
}

.dashboard-dataframe-note {
    margin-top: -0.25rem;
    margin-bottom: 0.9rem;
    color: var(--text-soft);
    font-size: 0.9rem;
}

.clean-stat {
    padding: 0.82rem 0.9rem;
    border-radius: 18px;
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.016));
    border: 1px solid rgba(255,255,255,0.07);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.035);
    min-height: 96px;
}

.clean-stat-label {
    color: var(--text-soft);
    font-size: 0.88rem;
    font-weight: 560;
    margin-bottom: 0.3rem;
    line-height: 1.5;
}

.clean-stat-value {
    color: var(--text-main);
    font-size: 1.78rem;
    font-weight: 780;
    letter-spacing: -0.03em;
    line-height: 1;
}

.panel-card {
    padding: 0.92rem 0.98rem;
    border-radius: 24px;
    background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.018));
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: var(--glow-soft), var(--shadow-soft);
    margin-bottom: 1rem;
    backdrop-filter: blur(10px);
}

.quick-card {
    padding: 1rem 1.05rem;
    border-radius: 18px;
    border: 1px solid rgba(255,255,255,0.08);
    background: rgba(255,255,255,0.022);
    margin-bottom: 0.8rem;
}

.compound-card {
    padding: 0.82rem 0.94rem;
    border-radius: 20px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.034), rgba(255,255,255,0.016));
    margin-bottom: 0.85rem;
    box-shadow: var(--glow-soft), var(--shadow-soft);
}

.compound-thumb-shell {
    width: 100%;
    height: 144px;
    display: flex;
    align-items: center;
    justify-content: center;
    overflow: hidden;
    border-radius: 18px;
    background: linear-gradient(180deg, rgba(255,255,255,0.065), rgba(255,255,255,0.018));
    border: 1px solid rgba(255,255,255,0.08);
}

.compound-thumb-shell img {
    width: 100%;
    height: 144px;
    object-fit: contain;
    display: block;
    background: rgba(255,255,255,0.96);
}

.compound-card:hover {
    border-color: rgba(115,231,255,0.22);
}

.result-card {
    padding: 1.05rem 1.1rem;
    border-radius: 22px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.034), rgba(255,255,255,0.018));
    margin-bottom: 0.8rem;
    box-shadow: var(--glow-soft), var(--shadow-soft);
}

.best-match-card {
    padding: 1.15rem 1.2rem;
    border-radius: 20px;
    border: 1px solid rgba(126, 240, 194, 0.30);
    background: linear-gradient(135deg, rgba(11, 103, 83, 0.18), rgba(53, 81, 152, 0.14));
    margin-bottom: 1rem;
}

.result-title {
    font-size: 1rem;
    font-weight: 780;
    margin-bottom: 0.18rem;
    color: var(--text-strong);
}

.result-subtitle {
    color: var(--text-soft);
    font-size: 0.85rem;
    margin-bottom: 0.42rem;
}

.badge-row {
    margin-top: 0.28rem;
    margin-bottom: 0.28rem;
    color: #D9DDE5;
    font-size: 0.95rem;
}

.info-chip-row {
    display: flex;
    flex-wrap: wrap;
    gap: 0.36rem;
    margin-top: 0.42rem;
}

.info-chip {
    display: inline-block;
    border-radius: 999px;
    padding: 0.3rem 0.62rem;
    font-size: 0.76rem;
    color: #E8EEF8;
    background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.025));
    border: 1px solid rgba(255,255,255,0.07);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
}

.kv-card {
    height: 100%;
    border-radius: 18px;
    padding: 1rem 1.05rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.028), rgba(255,255,255,0.014));
    border: 1px solid rgba(255,255,255,0.08);
    margin-bottom: 0.75rem;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
}

.kv-title {
    color: var(--text-soft);
    font-size: 0.87rem;
    margin-bottom: 0.18rem;
}

.kv-value {
    font-size: 1rem;
    font-weight: 660;
    color: var(--text-main);
    word-break: break-word;
    line-height: 1.55;
}

.structure-card {
    border-radius: 24px;
    padding: 1.05rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.015));
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: var(--glow-soft), var(--shadow-soft);
}

.record-shell {
    margin-top: 0.55rem;
}

.record-section-note {
    color: var(--text-soft);
    font-size: 0.94rem;
    line-height: 1.6;
    margin-top: -0.25rem;
    margin-bottom: 0.85rem;
}

.record-badge-strip {
    display: flex;
    flex-wrap: wrap;
    gap: 0.55rem;
    margin-bottom: 1rem;
}

.record-badge-strip.compact {
    margin-top: -0.35rem;
    margin-bottom: 0.85rem;
}

.record-badge {
    display: inline-flex;
    align-items: center;
    padding: 0.48rem 0.82rem;
    border-radius: 999px;
    background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.026));
    border: 1px solid rgba(255,255,255,0.08);
    color: #E8EEF8;
    font-size: 0.88rem;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
}

.structure-result-grid {
    display: grid;
    grid-template-columns: minmax(220px, 280px) 1fr;
    gap: 1rem;
    align-items: start;
}

.structure-result-meta {
    display: grid;
    gap: 0.45rem;
}

.structure-result-stat {
    color: var(--text-soft);
    font-size: 0.92rem;
    line-height: 1.5;
}

.query-summary-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 0.7rem;
}

.structure-search-shell {
    display: grid;
    grid-template-columns: minmax(0, 1.32fr) minmax(280px, 0.82fr);
    gap: 1rem;
    align-items: start;
}

.structure-search-shell .panel-card {
    border-radius: 24px;
    background: linear-gradient(180deg, rgba(12, 21, 36, 0.92), rgba(9, 16, 29, 0.94));
}

.structure-search-editor-title {
    color: var(--text-strong);
    font-size: 1.35rem;
    font-weight: 760;
    margin-bottom: 0.45rem;
}

.structure-search-editor-subtitle {
    color: var(--text-soft);
    font-size: 0.92rem;
    line-height: 1.55;
    margin-bottom: 0.9rem;
}

.structure-editor-note {
    margin-top: 0.55rem;
    color: var(--text-soft);
    font-size: 0.88rem;
    line-height: 1.5;
}

.search-preset-strip {
    display: flex;
    flex-wrap: wrap;
    gap: 0.5rem;
    margin-top: 0.7rem;
}

.search-preset-pill {
    display: inline-flex;
    align-items: center;
    gap: 0.42rem;
    padding: 0.5rem 0.72rem;
    border-radius: 999px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.015));
    color: var(--text-main);
    font-size: 0.82rem;
}

.search-preset-pill strong {
    color: var(--text-strong);
}

.query-summary-card {
    border-radius: 18px;
    padding: 0.95rem 1rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.036), rgba(255,255,255,0.018));
    border: 1px solid rgba(255,255,255,0.08);
}

.query-summary-label {
    color: var(--text-soft);
    font-size: 0.84rem;
    margin-bottom: 0.18rem;
}

.query-summary-value {
    color: var(--text-strong);
    font-size: 1.15rem;
    font-weight: 740;
    line-height: 1.2;
}

.detail-table-wrap {
    padding: 0.2rem 0 0.9rem 0;
}

@media (max-width: 900px) {
    .structure-result-grid {
        grid-template-columns: 1fr;
    }
}

.small-note {
    color: var(--text-soft);
    font-size: 0.92rem;
}

div[data-baseweb="select"] > div {
    border-radius: 16px !important;
    background: rgba(255,255,255,0.028) !important;
    border-color: rgba(255,255,255,0.09) !important;
    min-height: 44px !important;
}

div[data-testid="stTextInput"] input,
div[data-testid="stTextArea"] textarea,
div[data-testid="stNumberInput"] input {
    border-radius: 16px !important;
    background: rgba(255,255,255,0.026) !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    color: var(--text-main) !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.025);
}

div[data-testid="stTextInput"] input:focus,
div[data-testid="stTextArea"] textarea:focus,
div[data-testid="stNumberInput"] input:focus {
    border-color: rgba(97,216,237,0.34) !important;
    box-shadow: 0 0 0 1px rgba(97,216,237,0.18), 0 0 0 6px rgba(97,216,237,0.05) !important;
}

button[kind="primary"] {
    border-radius: 16px !important;
    min-height: 42px !important;
}

div[data-testid="stButton"] button,
div[data-testid="stDownloadButton"] button {
    border-radius: 16px !important;
    min-height: 46px !important;
    font-weight: 660 !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.055), rgba(255,255,255,0.028)) !important;
    border: 1px solid rgba(255,255,255,0.10) !important;
    color: #F5F8FD !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.04), 0 8px 22px rgba(0,0,0,0.16);
    transition: transform 0.16s ease, box-shadow 0.16s ease, border-color 0.16s ease !important;
}

div[data-testid="stButton"] button:hover,
div[data-testid="stDownloadButton"] button:hover {
    border-color: rgba(97, 216, 237, 0.36) !important;
    transform: translateY(-1px);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.05), 0 14px 26px rgba(0,0,0,0.2);
}

div[data-testid="stButton"] button[kind="primary"],
div[data-testid="stButton"] button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(90deg, #1EAFFF 0%, #2E86FF 38%, #8246FF 72%, #CB3EFF 100%) !important;
    border-color: rgba(255,255,255,0.18) !important;
    color: #FFFFFF !important;
    box-shadow: 0 18px 34px rgba(92, 85, 255, 0.24), inset 0 1px 0 rgba(255,255,255,0.18) !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button {
    min-height: 46px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important;
    text-align: left !important;
    padding: 0.48rem 0.78rem !important;
    border-radius: 14px !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    box-shadow: none !important;
    font-size: 0.92rem !important;
    line-height: 1.15 !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button > div {
    width: 100% !important;
    display: flex !important;
    align-items: center !important;
    justify-content: start !important;
    gap: 0.72rem !important;
    min-width: 0 !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button p {
    margin: 0 !important;
    text-align: left !important;
    white-space: normal !important;
    overflow-wrap: normal !important;
    word-break: normal !important;
    min-width: 0 !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button svg,
[data-testid="stSidebar"] div[data-testid="stButton"] button span[data-testid="stIconMaterial"] {
    width: 1.36rem !important;
    height: 1.36rem !important;
    min-width: 1.36rem !important;
    flex: 0 0 1.56rem !important;
    margin: 0 !important;
    font-size: 1.34rem !important;
    line-height: 1 !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="primary"] {
    background: linear-gradient(90deg, rgba(120, 66, 209, 0.92), rgba(48, 79, 167, 0.92)) !important;
    border-color: rgba(120, 176, 255, 0.18) !important;
    box-shadow: 0 12px 24px rgba(0,0,0,0.16) !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(90deg, rgba(120, 66, 209, 0.92), rgba(48, 79, 167, 0.92)) !important;
    border-color: rgba(120, 176, 255, 0.18) !important;
    box-shadow: 0 12px 24px rgba(0,0,0,0.16) !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="secondary"] {
    color: rgba(240, 246, 255, 0.92) !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button:hover {
    border-color: rgba(255,255,255,0.06) !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.015)) !important;
    transform: none !important;
}

[data-testid="stSidebar"] div[data-testid="stButton"] button[kind="primary"]:hover {
    background: linear-gradient(90deg, rgba(120, 66, 209, 0.96), rgba(48, 79, 167, 0.96)) !important;
}

main div[data-testid="stButton"] button[kind="primary"],
main div[data-testid="stButton"] button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(90deg, #1EAFFF 0%, #2E86FF 38%, #8246FF 72%, #CB3EFF 100%) !important;
    border-color: rgba(255,255,255,0.18) !important;
    color: #FFFFFF !important;
    box-shadow: 0 18px 34px rgba(92, 85, 255, 0.24), inset 0 1px 0 rgba(255,255,255,0.18) !important;
}

main div[data-testid="stButton"] button[kind="primary"] p,
main div[data-testid="stButton"] button[data-testid="stBaseButton-primary"] p {
    white-space: nowrap !important;
}

main div[data-testid="stButton"] button[kind="primary"]:hover,
main div[data-testid="stButton"] button[data-testid="stBaseButton-primary"]:hover {
    filter: brightness(1.05);
    transform: translateY(-1px);
}

div[data-testid="stRadio"] > div {
    display: flex !important;
    flex-direction: row !important;
    flex-wrap: wrap !important;
    align-items: center !important;
    gap: 0.6rem;
}

div[data-testid="stRadio"] label {
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.018));
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 999px;
    padding: 0.48rem 1rem;
    transition: all 0.18s ease;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
}

div[data-testid="stRadio"] label p {
    font-size: 0.88rem !important;
    font-weight: 600 !important;
    white-space: nowrap !important;
}

div[data-testid="stRadio"] label:has(input:checked) {
    background: linear-gradient(90deg, rgba(97,216,237,0.26), rgba(156,99,241,0.28));
    border-color: rgba(97,216,237,0.42);
    box-shadow: 0 0 0 1px rgba(97,216,237,0.06), 0 10px 24px rgba(76,142,255,0.12);
}

.action-strip {
    border-radius: 20px;
    padding: 1rem 1.05rem;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.018));
    margin-bottom: 1rem;
    box-shadow: var(--glow-soft);
}

.helper-card {
    border-radius: 20px;
    padding: 1.05rem 1.08rem;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.016));
    margin-bottom: 0.9rem;
    box-shadow: var(--glow-soft);
}

.helper-title {
    color: var(--text-main);
    font-size: 1rem;
    font-weight: 720;
    margin-bottom: 0.24rem;
}

.helper-text {
    color: var(--text-soft);
    font-size: 0.93rem;
    line-height: 1.5;
}

.section-banner {
    margin-bottom: 1rem;
    border-radius: 22px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
    background: rgba(255,255,255,0.02);
    box-shadow: var(--shadow-soft);
}

.accent-logo-wrap {
    margin-top: 1rem;
    padding-top: 0.2rem;
    text-align: center;
}

.sidebar-menu-caption {
    color: rgba(174,184,198,0.8);
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-top: 0.6rem;
    margin-bottom: 0.45rem;
}

.dashboard-hero-card {
    position: relative;
    min-height: 238px;
    border-radius: 24px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
    background:
        linear-gradient(115deg, rgba(5, 11, 26, 0.92), rgba(8, 18, 34, 0.72)),
        radial-gradient(circle at 82% 22%, rgba(197, 94, 255, 0.18), transparent 28%),
        radial-gradient(circle at 20% 18%, rgba(97, 216, 237, 0.12), transparent 24%);
    box-shadow: var(--shadow-deep);
}

.dashboard-hero-card.has-image {
    background-size: cover;
    background-position: center;
}

.dashboard-hero-card pre,
.dashboard-hero-card code,
.dashboard-hero-card [data-testid="stCodeBlock"],
.dashboard-hero-card [data-testid="stMarkdownContainer"] pre {
    display: none !important;
}

.dashboard-hero-overlay {
    position: absolute;
    inset: 0;
    background:
        linear-gradient(90deg, rgba(4, 10, 20, 0.9) 0%, rgba(4, 10, 20, 0.62) 48%, rgba(4, 10, 20, 0.18) 100%);
}

.dashboard-hero-content {
    position: relative;
    z-index: 1;
    max-width: 940px;
    padding: 2.1rem 1.65rem 1.55rem 1.65rem;
}

.dashboard-hero-kicker {
    color: rgba(174,184,198,0.88);
    font-size: 0.76rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
}

.dashboard-hero-title {
    margin: 0;
    color: #F7FBFF;
    font-size: clamp(2rem, 4.5vw, 3.35rem);
    line-height: 0.98;
    font-weight: 830;
    letter-spacing: -0.04em;
    text-wrap: balance;
}

.dashboard-hero-title .accent {
    background: linear-gradient(90deg, #60dfec, #4d8fff 46%, #ae5df4 78%, #f497d8);
    -webkit-background-clip: text;
    background-clip: text;
    color: transparent;
}

.workspace-headbar {
    position: relative;
    min-height: 132px;
    margin: 0 0 1.05rem 0;
    padding: 1.25rem 1.45rem;
    border-radius: 22px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
    background:
        linear-gradient(112deg, rgba(5, 11, 26, 0.92), rgba(8, 18, 34, 0.66)),
        radial-gradient(circle at 82% 24%, rgba(197, 94, 255, 0.16), transparent 30%),
        radial-gradient(circle at 22% 18%, rgba(97, 216, 237, 0.12), transparent 26%);
    box-shadow: var(--shadow-soft);
    display: flex;
    align-items: center;
}

.workspace-headbar.has-image {
    background-size: cover;
    background-position: center right;
}

.workspace-headbar::after {
    content: "";
    position: absolute;
    inset: 0;
    background: linear-gradient(90deg, rgba(4,10,20,0.88) 0%, rgba(4,10,20,0.58) 52%, rgba(4,10,20,0.22) 100%);
    pointer-events: none;
}

.workspace-headbar-copy {
    position: relative;
    z-index: 1;
}

.workspace-headbar-kicker {
    color: rgba(174,184,198,0.84);
    font-size: 0.72rem;
    font-weight: 760;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 0.34rem;
}

.workspace-headbar-title {
    color: #F7FBFF;
    font-size: clamp(1.55rem, 3.2vw, 2.55rem);
    line-height: 1.02;
    font-weight: 830;
    letter-spacing: 0;
}

.workspace-headbar-title span {
    background: linear-gradient(90deg, #60dfec, #4d8fff 46%, #ae5df4 78%, #f497d8);
    -webkit-background-clip: text;
    background-clip: text;
    color: transparent;
}

.dashboard-workspace-card,
.workflow-card,
.dashboard-workflow-shell {
    border-radius: 26px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.045), rgba(255,255,255,0.018));
    box-shadow: var(--shadow-soft);
}

.dashboard-cta-kicker {
    color: rgba(174,184,198,0.84);
    font-size: 0.78rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 0.45rem;
}

.dashboard-cta-title {
    color: var(--text-strong);
    font-size: 1.58rem;
    font-weight: 760;
    letter-spacing: -0.03em;
    margin-bottom: 0.3rem;
}

.dashboard-cta-copy {
    color: var(--text-soft);
    font-size: 0.96rem;
    line-height: 1.6;
    max-width: 34rem;
}

.dashboard-workspace-card {
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    min-height: 214px;
    padding: 0.82rem 0.92rem 0.72rem 0.92rem;
    overflow: hidden;
}

.dashboard-workspace-title {
    color: var(--text-strong);
    font-size: 0.94rem;
    font-weight: 750;
    margin-bottom: 0.22rem;
}

.dashboard-workspace-copy {
    color: var(--text-soft);
    font-size: 0.68rem;
    line-height: 1.3;
    margin-bottom: 0.12rem;
}

.dashboard-workspace-art {
    width: 100%;
    max-height: 102px;
    object-fit: contain;
    align-self: center;
}

.dashboard-workflow-shell {
    position: relative;
    padding: 1.02rem 1.05rem 1rem 1.05rem;
    overflow: hidden;
}

.dashboard-workflow-title {
    color: var(--text-strong);
    font-size: 1rem;
    font-weight: 760;
    margin-bottom: 1rem;
}

.dashboard-workflow-grid {
    position: relative;
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 0.84rem;
    align-items: stretch;
}

.dashboard-workflow-grid::before {
    content: "";
    position: absolute;
    top: 1.05rem;
    left: 8%;
    right: 8%;
    border-top: 1px dotted rgba(174, 184, 198, 0.34);
    z-index: 0;
}

.workflow-card {
    display: block;
    padding: 2.68rem 0.78rem 0.82rem 0.78rem;
    min-height: 150px;
    position: relative;
    background: linear-gradient(180deg, rgba(20, 31, 51, 0.82), rgba(10, 18, 31, 0.92));
    text-decoration: none;
    z-index: 1;
}

.dashboard-chart-card {
    border-radius: 24px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(255,255,255,0.036), rgba(255,255,255,0.016));
    box-shadow: var(--shadow-soft);
    padding: 0.8rem 0.86rem 0.68rem 0.86rem;
    min-height: 420px;
}

.chart-legend-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
    gap: 0.32rem 0.65rem;
    margin: 0.04rem 0 0.95rem 0;
}

.chart-legend-item {
    display: grid;
    grid-template-columns: 0.72rem minmax(0, 1fr) auto;
    align-items: center;
    gap: 0.45rem;
    color: #DDE6F3;
    font-size: 0.78rem;
    line-height: 1.25;
    min-width: 0;
}

.chart-legend-swatch {
    width: 0.72rem;
    height: 0.72rem;
    border-radius: 3px;
    border: 1px solid rgba(255,255,255,0.22);
}

.chart-legend-label {
    min-width: 0;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}

.chart-legend-percent {
    color: rgba(221,230,243,0.72);
    font-variant-numeric: tabular-nums;
}

.workflow-card.is-primary {
    border-color: rgba(22, 228, 240, 0.8);
    box-shadow: 0 0 0 1px rgba(22, 228, 240, 0.18), 0 24px 42px rgba(0,0,0,0.26);
}

.workflow-card.is-primary .workflow-step {
    background: linear-gradient(135deg, rgba(16, 210, 232, 0.92), rgba(88, 72, 214, 0.92));
    border-color: rgba(97, 216, 237, 0.72);
    box-shadow: 0 0 0 4px rgba(16, 210, 232, 0.12), 0 0 24px rgba(97, 216, 237, 0.2);
}

.workflow-card:hover {
    transform: translateY(-1px);
    border-color: rgba(97,216,237,0.28);
}

.workflow-step {
    position: absolute;
    top: -0.08rem;
    left: 50%;
    transform: translateX(-50%);
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 2.08rem;
    height: 2.08rem;
    border-radius: 999px;
    background: linear-gradient(90deg, rgba(97,216,237,0.26), rgba(156,99,241,0.3));
    border: 1px solid rgba(97,216,237,0.28);
    color: var(--text-strong);
    font-weight: 760;
    z-index: 2;
}

.workflow-title {
    color: var(--text-strong);
    font-size: 0.84rem;
    font-weight: 720;
    margin-bottom: 0.26rem;
    line-height: 1.22;
}

.workflow-card-icon-shell {
    width: 2.72rem;
    height: 2.72rem;
    margin-bottom: 0.62rem;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    border-radius: 18px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(180deg, rgba(22, 30, 56, 0.88), rgba(9, 17, 29, 0.92));
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
}

.workflow-card-icon {
    width: 1.62rem;
    height: 1.62rem;
    object-fit: contain;
    filter: drop-shadow(0 0 10px rgba(97, 216, 237, 0.18));
}

.workflow-copy {
    color: var(--text-soft);
    font-size: 0.68rem;
    line-height: 1.34;
}

.dashboard-hero-header {
    display: block;
}

.dashboard-hero-text {
    flex: 1 1 auto;
}

.dashboard-hero-title-shell {
    max-width: 42rem;
}

.dashboard-tagline {
    margin: 0.7rem 0 0 0;
    color: rgba(230, 237, 247, 0.78);
    font-size: 0.95rem;
    line-height: 1.55;
}

.dashboard-stat-board {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 0.2rem;
    margin-top: 1.05rem;
    max-width: 42rem;
}

.dashboard-stat-board-item {
    position: relative;
    padding: 0.25rem 1rem 0.15rem 1rem;
}

.dashboard-stat-board-item + .dashboard-stat-board-item::before {
    content: "";
    position: absolute;
    left: 0;
    top: 0.45rem;
    bottom: 0.45rem;
    width: 1px;
    background: linear-gradient(180deg, rgba(255,255,255,0), rgba(255,255,255,0.24), rgba(255,255,255,0));
}

.dashboard-stat-board-head {
    display: flex;
    align-items: center;
    gap: 0.72rem;
    margin-bottom: 0.25rem;
}

.dashboard-stat-board-head.is-updated-head {
    align-items: center;
}

.dashboard-stat-board-icon {
    width: 36px;
    height: 36px;
    border-radius: 14px;
    object-fit: contain;
    background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.02));
    border: 1px solid rgba(255,255,255,0.08);
    padding: 0.36rem;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
}

.dashboard-stat-board-copy {
    min-width: 0;
    display: flex;
    flex-direction: column;
    justify-content: flex-start;
}

.dashboard-stat-board-copy.is-updated {
    padding-top: 0;
}

.dashboard-stat-board-value {
    color: var(--text-strong);
    font-size: clamp(1.42rem, 2.4vw, 1.86rem);
    font-weight: 780;
    line-height: 1;
    letter-spacing: -0.04em;
    margin-bottom: 0.22rem;
}

.dashboard-stat-board-label {
    color: rgba(230, 237, 247, 0.88);
    font-size: 0.84rem;
    line-height: 1.35;
}

.dashboard-stat-board-label-top {
    margin-top: 0;
    margin-bottom: 0.18rem;
}

.dashboard-stat-board-date {
    color: rgba(230, 237, 247, 0.92);
    font-size: 1rem;
    line-height: 1.4;
}

.dashboard-search-strip {
    margin-top: 0.6rem;
    display: flex;
    align-items: center;
    gap: 1rem;
    padding: 0.88rem 1rem;
    border-radius: 22px;
    border: 1px solid rgba(255,255,255,0.08);
    background: linear-gradient(90deg, rgba(11, 22, 40, 0.88), rgba(21, 24, 52, 0.9));
    box-shadow: var(--shadow-soft);
}

.dashboard-search-strip-icon-shell {
    flex: 0 0 auto;
}

.dashboard-search-strip-icon {
    width: 72px;
    height: 72px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: transparent;
    color: #f7fbff;
    font-size: 2rem;
    box-shadow: none;
}

.dashboard-search-strip-icon-image {
    width: 72px;
    height: 72px;
    object-fit: contain;
    display: block;
    filter: drop-shadow(0 14px 22px rgba(30, 87, 196, 0.18));
}

.dashboard-search-strip-copy-shell {
    min-width: 0;
}

.dashboard-search-strip-title {
    color: var(--text-strong);
    font-size: 1.14rem;
    font-weight: 760;
    margin-bottom: 0.16rem;
}

.dashboard-search-strip-copy {
    color: var(--text-soft);
    font-size: 0.84rem;
    line-height: 1.42;
}

.dashboard-search-cta {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 100%;
    min-height: 56px;
    padding: 0.7rem 0.9rem;
    border-radius: 18px;
    border: 1px solid rgba(255,255,255,0.18);
    background: linear-gradient(90deg, #1EAFFF 0%, #2E86FF 38%, #8246FF 72%, #CB3EFF 100%);
    color: #FFFFFF !important;
    -webkit-text-fill-color: #FFFFFF;
    font-size: 0.98rem;
    font-weight: 800;
    text-decoration: none;
    letter-spacing: 0.01em;
    box-shadow: 0 18px 34px rgba(92, 85, 255, 0.24), inset 0 1px 0 rgba(255,255,255,0.18);
    text-shadow: none;
    white-space: nowrap;
    line-height: 1.05;
}

.dashboard-search-cta:visited {
    color: #FFFFFF !important;
    -webkit-text-fill-color: #FFFFFF;
}

.dashboard-search-cta:hover {
    filter: brightness(1.05);
}

[data-testid="stDataFrame"] {
    border-radius: 18px;
    overflow: hidden;
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow: var(--glow-soft);
}

[data-testid="stDataFrame"] [role="grid"] {
    background: rgba(8, 17, 30, 0.45) !important;
}

[data-testid="stTabs"] [data-baseweb="tab-list"] {
    gap: 0.55rem;
    padding: 0.18rem;
    background: linear-gradient(180deg, rgba(255,255,255,0.026), rgba(255,255,255,0.014));
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 18px;
    margin-bottom: 1rem;
}

[data-testid="stTabs"] button[role="tab"] {
    min-height: 38px;
    border-radius: 14px !important;
    color: var(--text-soft) !important;
    font-weight: 650 !important;
    transition: all 0.18s ease;
}

[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background: linear-gradient(90deg, rgba(97,216,237,0.22), rgba(156,99,241,0.24)) !important;
    color: var(--text-strong) !important;
    box-shadow: 0 8px 18px rgba(76,142,255,0.12);
}

[data-testid="stExpander"] {
    border-radius: 18px !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.014)) !important;
    box-shadow: var(--glow-soft);
    overflow: hidden;
}

[data-testid="stExpander"] details summary {
    padding-top: 0.25rem;
    padding-bottom: 0.25rem;
}

[data-testid="stFileUploader"] section {
    border-radius: 18px !important;
    border: 1px dashed rgba(97,216,237,0.24) !important;
    background: linear-gradient(180deg, rgba(255,255,255,0.028), rgba(255,255,255,0.012)) !important;
}

[data-testid="stAlert"] {
    border-radius: 18px !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    box-shadow: var(--glow-soft);
}

hr {
    border-color: rgba(255,255,255,0.07);
}

header[data-testid="stHeader"] {
    background: rgba(7, 17, 29, 0.32);
}

/* Typography harmony pass: keep the app polished and readable across menus. */
.section-title,
.structure-search-editor-title {
    font-size: clamp(1.34rem, 2vw, 1.68rem);
    line-height: 1.15;
    letter-spacing: 0;
}

.section-subtitle,
.structure-search-editor-subtitle,
.helper-text,
.selector-subtitle,
.inline-note,
.record-section-note,
.dashboard-search-strip-copy,
.dashboard-workspace-copy,
.workflow-copy,
.result-subtitle,
.small-note,
div[data-testid="stCaptionContainer"],
[data-testid="stMarkdownContainer"] p {
    font-size: 0.9rem;
    line-height: 1.48;
}

.dashboard-workspace-title,
.dashboard-workflow-title,
.helper-title,
.selector-title,
.result-title,
.query-summary-value {
    font-size: 1rem;
    line-height: 1.24;
    letter-spacing: 0;
}

.dashboard-hero-title {
    font-size: clamp(2rem, 4vw, 3rem);
    letter-spacing: 0;
}

.workspace-headbar-title {
    font-size: clamp(1.52rem, 2.8vw, 2.35rem);
}

.dashboard-tagline,
.dashboard-cta-copy,
.badge-row,
.kv-value {
    font-size: 0.94rem;
    line-height: 1.48;
}

.kv-title,
.metric-card-label,
.clean-stat-label,
.query-summary-label,
.structure-result-stat,
.chart-legend-item,
.info-chip,
.record-badge {
    font-size: 0.82rem;
    line-height: 1.36;
}

.dashboard-workspace-copy,
.workflow-copy {
    font-size: 0.76rem;
}

.workflow-title {
    font-size: 0.9rem;
    line-height: 1.25;
    letter-spacing: 0;
}

div[data-testid="stButton"] button,
div[data-testid="stDownloadButton"] button,
div[data-testid="stRadio"] label p,
[data-testid="stTabs"] button[role="tab"],
div[data-testid="stTextInput"] input,
div[data-testid="stTextArea"] textarea,
div[data-testid="stNumberInput"] input,
div[data-baseweb="select"] {
    font-size: 0.9rem !important;
    line-height: 1.28 !important;
}

[data-testid="stDataFrame"] {
    font-size: 0.84rem;
}

@media (max-width: 900px) {
    .section-title {
        font-size: 1.28rem;
    }

    .section-subtitle {
        font-size: 0.9rem;
    }

    .sidebar-stats {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }

    .dashboard-stat-strip {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }

    .dashboard-stat-board {
        grid-template-columns: repeat(3, minmax(0, 1fr));
    }

    .dashboard-hero-header {
        flex-direction: column;
        align-items: flex-start;
    }

    .app-credit-footer {
        display: flex;
        text-align: center;
        white-space: normal;
    }

    .dashboard-search-strip {
        flex-direction: column;
        align-items: flex-start;
    }
}

@media (max-width: 700px) {
    .dashboard-workflow-grid {
        grid-template-columns: 1fr;
    }

    .dashboard-stat-strip {
        grid-template-columns: 1fr;
    }

    .dashboard-stat-board {
        grid-template-columns: 1fr;
    }

    .dashboard-stat-board-item {
        padding-left: 0;
    }

    .dashboard-stat-board-item + .dashboard-stat-board-item::before {
        display: none;
    }

    .dashboard-hero-content {
        padding: 1.5rem 1.2rem 1.2rem 1.2rem;
    }

    .workspace-headbar {
        min-height: 116px;
        padding: 1.1rem 1rem;
        border-radius: 18px;
        background-position: 62% center;
    }

    .structure-search-shell {
        grid-template-columns: 1fr;
    }

}

@media (max-width: 1180px) {
    .block-container {
        max-width: 100% !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }

    [data-testid="stSidebar"] {
        min-width: 16rem !important;
        max-width: 16rem !important;
    }

    .dashboard-hero-card {
        min-height: 190px;
    }

    .dashboard-search-cta {
        white-space: normal;
        text-align: center;
    }
}

@media (max-width: 960px) {
    .dashboard-workspace-card {
        min-height: auto;
    }

    .dashboard-workspace-art {
        max-height: 92px;
    }

    .dashboard-workflow-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }

    .dashboard-workflow-grid::before {
        display: none;
    }

    .workflow-card {
        min-height: auto;
    }

    .dashboard-search-strip-icon,
    .dashboard-search-strip-icon-image {
        width: 54px;
        height: 54px;
    }
}

@media (max-width: 820px) {
    [data-testid="stSidebar"] {
        min-width: 14.5rem !important;
        max-width: 14.5rem !important;
    }

    .sidebar-brand-head {
        gap: 0.56rem;
    }

    .sidebar-brand-logo-shell {
        width: 104px;
        height: 104px;
        border-radius: 24px;
    }

    .sidebar-brand-logo {
        width: 98px;
        height: 98px;
    }

    .sidebar-brand-title {
        font-size: 0.94rem;
    }

    .sidebar-nav-link {
        min-height: 48px;
        padding: 0.46rem 0.62rem;
    }

    .sidebar-nav-link .sidebar-nav-icon-shell {
        width: 2rem;
        height: 2rem;
    }

    .sidebar-nav-link-label {
        font-size: 0.9rem;
    }
}

.panel-card,
.compound-card,
.result-card,
.helper-card,
.kv-card,
.workflow-card,
.dashboard-workspace-card,
.dashboard-workflow-shell,
.dashboard-search-strip,
.dashboard-hero-card,
.workspace-headbar {
    min-width: 0;
}

.section-title,
.dashboard-hero-title,
.workspace-headbar-title,
.dashboard-search-strip-title,
.dashboard-workspace-title,
.workflow-title,
.result-title,
.sidebar-nav-link-label,
.info-chip,
.kv-value {
    overflow-wrap: anywhere;
}

</style>
""", unsafe_allow_html=True)

# =========================
# Database connection
# =========================
def ensure_project_dirs():
    for directory in [
        DATABASE_DIR,
        BRANDING_DIR,
        BRANDING_OPTIMIZED_DIR,
        STRUCTURES_DIR,
        SPECTRA_DIR,
        TEMPLATES_DIR,
        SUBMISSIONS_DIR,
        SUBMISSIONS_INBOX_DIR,
        SUBMISSIONS_REVIEWED_DIR,
        SUBMISSIONS_APPROVED_DIR,
        EXPORTS_DIR,
        DOCS_DIR,
        BACKUPS_DIR,
    ]:
        directory.mkdir(parents=True, exist_ok=True)


def get_connection():
    ensure_project_dirs()
    connection = sqlite3.connect(DB_PATH)
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute("PRAGMA synchronous = NORMAL")
    connection.execute("PRAGMA temp_store = MEMORY")
    connection.execute("PRAGMA cache_size = -32000")
    return connection


def invalidate_cached_views():
    try:
        st.cache_data.clear()
    except Exception:
        pass


def table_exists(table_name: str) -> bool:
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        )
        return cursor.fetchone() is not None
    finally:
        conn.close()


def get_table_columns(table_name: str):
    if not table_exists(table_name):
        return set()

    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        rows = cursor.fetchall()
        return {row[1] for row in rows}
    finally:
        conn.close()


def ensure_database_schema():
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS compounds (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trivial_name TEXT NOT NULL,
                iupac_name TEXT,
                molecular_formula TEXT,
                smiles TEXT,
                inchi TEXT,
                inchikey TEXT,
                compound_class TEXT,
                compound_subclass TEXT,
                source_category TEXT,
                source_organism TEXT,
                source_material TEXT,
                sample_code TEXT,
                collection_location TEXT,
                gps_coordinates TEXT,
                depth_m REAL,
                uv_data TEXT,
                ftir_data TEXT,
                cd_data TEXT,
                optical_rotation TEXT,
                melting_point TEXT,
                crystallization_method TEXT,
                structure_image_path TEXT,
                journal_name TEXT,
                article_title TEXT,
                publication_year TEXT,
                volume TEXT,
                issue TEXT,
                pages TEXT,
                doi TEXT,
                ccdc_number TEXT,
                molecular_weight REAL,
                hrms_data TEXT,
                data_source TEXT,
                curation_status TEXT DEFAULT 'curated',
                note TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS proton_nmr (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compound_id INTEGER NOT NULL,
                delta_ppm REAL NOT NULL,
                multiplicity TEXT,
                j_value TEXT,
                proton_count TEXT,
                assignment TEXT,
                solvent TEXT,
                instrument_mhz REAL,
                note TEXT,
                FOREIGN KEY (compound_id) REFERENCES compounds(id) ON DELETE CASCADE
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS carbon_nmr (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compound_id INTEGER NOT NULL,
                delta_ppm REAL NOT NULL,
                carbon_type TEXT,
                assignment TEXT,
                solvent TEXT,
                instrument_mhz REAL,
                note TEXT,
                FOREIGN KEY (compound_id) REFERENCES compounds(id) ON DELETE CASCADE
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS spectra_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compound_id INTEGER NOT NULL,
                spectrum_type TEXT NOT NULL,
                file_path TEXT NOT NULL,
                note TEXT,
                FOREIGN KEY (compound_id) REFERENCES compounds(id) ON DELETE CASCADE
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS bioactivity_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compound_id INTEGER NOT NULL,
                activity_label TEXT NOT NULL,
                target_name TEXT,
                target_category TEXT,
                assay_type TEXT,
                potency_type TEXT,
                potency_relation TEXT,
                potency_value REAL,
                potency_unit TEXT,
                outcome TEXT,
                assay_medium TEXT,
                selectivity TEXT,
                assay_source TEXT,
                note TEXT,
                FOREIGN KEY (compound_id) REFERENCES compounds(id) ON DELETE CASCADE
            )
            """
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_trivial_name ON compounds(trivial_name)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_sample_code ON compounds(sample_code)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_doi ON compounds(doi)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_inchikey ON compounds(inchikey)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_smiles ON compounds(smiles)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_updated_at ON compounds(updated_at)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_curation_status ON compounds(curation_status)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_compound_class ON compounds(compound_class)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_source_category ON compounds(source_category)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_source_organism ON compounds(source_organism)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_compounds_data_source ON compounds(data_source)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_proton_compound ON proton_nmr(compound_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_proton_compound_delta ON proton_nmr(compound_id, delta_ppm)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_carbon_compound ON carbon_nmr(compound_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_carbon_compound_delta ON carbon_nmr(compound_id, delta_ppm)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_spectra_compound ON spectra_files(compound_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_spectra_type ON spectra_files(spectrum_type)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_bioactivity_compound ON bioactivity_records(compound_id)"
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_bioactivity_target_category ON bioactivity_records(target_category)"
        )
        conn.commit()
        invalidate_cached_views()
    finally:
        conn.close()


def ensure_compounds_schema():
    required_columns = {
        "issue": "TEXT",
        "ccdc_number": "TEXT",
        "molecular_weight": "REAL",
        "smiles": "TEXT",
        "inchi": "TEXT",
        "inchikey": "TEXT",
        "hrms_data": "TEXT",
        "source_category": "TEXT",
        "source_organism": "TEXT",
        "cd_data": "TEXT",
        "article_title": "TEXT",
        "curation_status": "TEXT DEFAULT 'curated'",
        "created_at": "TEXT",
        "updated_at": "TEXT",
    }

    existing = get_table_columns("compounds")
    missing = {name: dtype for name, dtype in required_columns.items() if name not in existing}
    conn = get_connection()
    try:
        cursor = conn.cursor()
        for column_name, data_type in missing.items():
            cursor.execute(f"ALTER TABLE compounds ADD COLUMN {column_name} {data_type}")
        cursor.execute(
            """
            UPDATE compounds
            SET curation_status = 'curated'
            WHERE curation_status IS NULL OR TRIM(curation_status) = ''
            """
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_compounds_inchikey ON compounds(inchikey)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_compounds_smiles ON compounds(smiles)")
        conn.commit()
        invalidate_cached_views()
    finally:
        conn.close()


def ensure_bioactivity_schema():
    required_columns = {
        "activity_label": "TEXT",
        "target_name": "TEXT",
        "target_category": "TEXT",
        "assay_type": "TEXT",
        "potency_type": "TEXT",
        "potency_relation": "TEXT",
        "potency_value": "REAL",
        "potency_unit": "TEXT",
        "outcome": "TEXT",
        "assay_medium": "TEXT",
        "selectivity": "TEXT",
        "assay_source": "TEXT",
        "note": "TEXT",
    }

    if not table_exists("bioactivity_records"):
        ensure_database_schema()
        return

    existing = get_table_columns("bioactivity_records")
    missing = {name: dtype for name, dtype in required_columns.items() if name not in existing}
    if not missing:
        return

    conn = get_connection()
    try:
        cursor = conn.cursor()
        for column_name, data_type in missing.items():
            cursor.execute(f"ALTER TABLE bioactivity_records ADD COLUMN {column_name} {data_type}")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_bioactivity_compound ON bioactivity_records(compound_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_bioactivity_target_category ON bioactivity_records(target_category)")
        conn.commit()
        invalidate_cached_views()
    finally:
        conn.close()


if should_initialize_sqlite_schema():
    ensure_database_schema()
    ensure_compounds_schema()
    ensure_bioactivity_schema()

# =========================
# Generic helpers
# =========================
def clean_text(value):
    if pd.isna(value) or value is None:
        return "-"
    text = str(value).strip()
    return text if text else "-"


def html_text(value):
    return html.escape(clean_text(value), quote=True)


def maybe_blank(value):
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


@st.cache_data(show_spinner=False)
def image_to_data_uri(path_value: str, max_px: int = 720) -> str:
    return optimized_image_data_uri(path_value, max_px=max_px)


def render_cloud_sync_notice():
    if use_local_read_backend() and use_supabase_write_backend():
        return
    elif not use_supabase_backend():
        st.warning(
            "Local-only mode is active. Changes in this session are still being written to the desktop database only. "
            "Before production use, configure Supabase secrets in local and deployed environments so every submission goes to the same cloud database."
        )
    elif cloud_write_is_blocked():
        st.warning(
            "Supabase is already the active read backend, but cloud write access is still locked. "
            "This app is keeping editing disabled so new submissions or metadata changes do not get split between local storage and Supabase."
        )
    elif use_supabase_write_backend():
        return

def safe_float_or_none(value):
    text = maybe_blank(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_source_category(value: str) -> str:
    text = maybe_blank(value)
    if not text:
        return ""
    for option in DEFAULT_SOURCE_OPTIONS:
        if text.casefold() == option.casefold():
            return option
    return text


def infer_source_fields(source_category="", source_organism="", source_material="") -> tuple[str, str, str]:
    raw_category = maybe_blank(source_category)
    category = normalize_source_category(raw_category)
    organism = maybe_blank(source_organism)
    legacy = maybe_blank(source_material)
    raw_category_is_source_option = any(raw_category.casefold() == option.casefold() for option in DEFAULT_SOURCE_OPTIONS)
    raw_category_has_marine_hint = any(token in raw_category.casefold() for token in MARINE_SOURCE_ORGANISM_HINTS)
    legacy_is_source_option = any(legacy.casefold() == option.casefold() for option in DEFAULT_SOURCE_OPTIONS)
    legacy_has_marine_hint = any(token in legacy.casefold() for token in MARINE_SOURCE_ORGANISM_HINTS)

    if raw_category and raw_category_has_marine_hint and not raw_category_is_source_option:
        category = "Marine"
        if not organism and "natural product" not in raw_category.casefold():
            organism = raw_category

    if not category and legacy:
        legacy_normalized = normalize_source_category(legacy)
        if legacy_is_source_option or legacy_normalized.casefold() != legacy.casefold():
            category = legacy_normalized
        elif legacy_has_marine_hint:
            category = "Marine"

    if not organism and legacy:
        normalized_legacy_category = normalize_source_category(legacy)
        if (
            not normalized_legacy_category
            or normalized_legacy_category.casefold() != legacy.casefold()
            or (legacy_has_marine_hint and not legacy_is_source_option)
        ):
            organism = legacy

    summary = legacy
    if category and organism:
        summary = f"{category} | {organism}"
    elif category:
        summary = category
    elif organism:
        summary = organism

    return category, organism, summary


def source_summary_from_record(record) -> str:
    category, organism, summary = infer_source_fields(
        record.get("source_category"),
        record.get("source_organism"),
        record.get("source_material"),
    )
    return summary


def normalize_curation_status(value: str, default: str = "curated") -> str:
    text = maybe_blank(value).lower()
    if text in CURATION_STATUS_OPTIONS:
        return text
    return default


def infer_curation_status(record, default: str = "curated") -> str:
    explicit = maybe_blank(record.get("curation_status"))
    if explicit:
        return normalize_curation_status(explicit, default=default)
    note_text = maybe_blank(record.get("note")).lower()
    data_source = maybe_blank(record.get("data_source")).lower()
    imported_sources = {"coconut", "cmnpd", "np-mrd", "npatlas", "nmrshiftdb", "jeol"}
    if "imported from" in note_text or data_source in imported_sources:
        return "imported"
    return default


COMPOUND_REQUIRED_COLUMNS = [
    "id",
    "trivial_name",
    "iupac_name",
    "molecular_formula",
    "smiles",
    "inchi",
    "inchikey",
    "compound_class",
    "compound_subclass",
    "source_category",
    "source_organism",
    "source_material",
    "sample_code",
    "collection_location",
    "gps_coordinates",
    "depth_m",
    "uv_data",
    "ftir_data",
    "cd_data",
    "optical_rotation",
    "melting_point",
    "crystallization_method",
    "structure_image_path",
    "journal_name",
    "article_title",
    "publication_year",
    "volume",
    "issue",
    "pages",
    "doi",
    "ccdc_number",
    "molecular_weight",
    "hrms_data",
    "data_source",
    "curation_status",
    "note",
    "created_at",
    "updated_at",
]


def enrich_compounds_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    enriched = df.copy()
    for column_name in COMPOUND_REQUIRED_COLUMNS:
        if column_name not in enriched.columns:
            enriched[column_name] = ""

    if enriched.empty:
        return enriched[COMPOUND_REQUIRED_COLUMNS]

    source_fields = enriched.apply(
        lambda row: infer_source_fields(
            row.get("source_category"),
            row.get("source_organism"),
            row.get("source_material"),
        ),
        axis=1,
        result_type="expand",
    )
    source_fields.columns = ["source_category", "source_organism", "source_material"]
    enriched["source_category"] = source_fields["source_category"]
    enriched["source_organism"] = source_fields["source_organism"]
    enriched["source_material"] = source_fields["source_material"]
    enriched["curation_status"] = enriched.apply(lambda row: infer_curation_status(row), axis=1)
    return enriched


def is_raw_spectrum_type(spectrum_type_value: str) -> bool:
    text = maybe_blank(spectrum_type_value).lower()
    raw_tokens = ["raw", "jcamp", "mnova", "fid"]
    return any(token in text for token in raw_tokens)


def classify_storage_type(file_path_value: str) -> str:
    text = maybe_blank(file_path_value)
    if not text:
        return "Unknown"
    if is_supabase_storage_reference(text):
        return "Supabase Storage"
    if is_google_drive_url(text):
        return "Google Drive"
    if is_external_url(text):
        return "External URL"
    return "Local file"


def validate_spectrum_entry(file_path_value: str, spectrum_type_value: str) -> tuple[list[str], list[str]]:
    errors = []
    warnings = []
    path_text = maybe_blank(file_path_value)
    spectrum_text = maybe_blank(spectrum_type_value)

    if not path_text:
        errors.append("File path or URL is required.")
        return errors, warnings

    if is_external_url(path_text):
        if is_google_drive_url(path_text):
            file_id = extract_google_drive_file_id(path_text)
            if not file_id:
                warnings.append("Google Drive link was detected, but the file ID could not be extracted. Preview/download may fail.")
            if is_raw_spectrum_type(spectrum_text):
                warnings.append("Raw-data link saved in Google Drive mode. Make sure sharing permission is set to viewer/download for approved users.")
        elif is_raw_spectrum_type(spectrum_text):
            warnings.append("Raw-data link uses a non-Google external URL. Confirm that users can access it from outside your laptop.")
        return errors, warnings

    full_path = get_full_file_path(path_text)
    if full_path is None or not full_path.exists():
        warnings.append("Local file path was saved, but the file does not currently exist at that location.")
    elif is_raw_spectrum_type(spectrum_text):
        warnings.append("Raw-data file is stored locally. Google Drive is safer for public deployment and laptop storage.")
    return errors, warnings

def slugify_value(value: str, fallback: str = "file") -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", maybe_blank(value))
    text = text.strip("._")
    return text or fallback

def relative_project_path(path: Path) -> str:
    return str(path.relative_to(PROJECT_DIR))

def build_existing_options(df: pd.DataFrame, column_name: str, defaults=None):
    values = set(defaults or [])
    if column_name in df.columns:
        for value in df[column_name].dropna().astype(str):
            cleaned = value.strip()
            if cleaned:
                values.add(cleaned)
    return sorted(values)

def select_or_custom(label: str, options: list[str], key: str, value: str = "", help_text: str | None = None):
    normalized_value = maybe_blank(value)
    clean_options = [item for item in options if maybe_blank(item)]
    known_values = sorted(set(clean_options))

    custom_default = normalized_value if normalized_value and normalized_value not in known_values else ""
    select_options = [""] + known_values + ["Custom..."]
    default_value = normalized_value if normalized_value in known_values else ("Custom..." if custom_default else "")
    select_key = f"{key}_select"
    selectbox_kwargs = {
        "label": label,
        "options": select_options,
        "key": select_key,
        "help": help_text,
    }
    if select_key not in st.session_state:
        selectbox_kwargs["index"] = select_options.index(default_value)
    selected = st.selectbox(**selectbox_kwargs)
    show_custom_input = selected == "Custom..." or bool(custom_default)
    if show_custom_input:
        custom_key = f"{key}_custom"
        custom_kwargs = {
            "label": f"{label} (Custom, optional)",
            "key": custom_key,
            "placeholder": f"Type a new {label.lower()} here if it is not in the list.",
        }
        if custom_key not in st.session_state:
            custom_kwargs["value"] = custom_default
        custom_value = st.text_input(**custom_kwargs)
        custom_text = maybe_blank(custom_value)
        if custom_text:
            return custom_text
        return ""
    return selected

def reset_compound_wizard():
    wizard_keys = [
        "compound_wizard_step",
        "wizard_trivial_name",
        "wizard_iupac_name",
        "wizard_formula",
        "wizard_molecular_weight",
        "wizard_smiles",
        "wizard_inchi",
        "wizard_inchikey",
        "wizard_compound_class_select",
        "wizard_compound_class_custom",
        "wizard_compound_subclass_select",
        "wizard_compound_subclass_custom",
        "wizard_data_source_select",
        "wizard_data_source_custom",
        "wizard_source_category_select",
        "wizard_source_category_custom",
        "wizard_source_organism",
        "wizard_sample_code",
        "wizard_collection_location",
        "wizard_gps_coordinates",
        "wizard_depth_m",
        "wizard_uv_data",
        "wizard_ftir_data",
        "wizard_cd_data",
        "wizard_optical_rotation",
        "wizard_melting_point",
        "wizard_crystallization_method",
        "wizard_ccdc_number",
        "wizard_hrms_data",
        "wizard_structure_path",
        "wizard_structure_upload",
        "wizard_submission_spectrum_type_select",
        "wizard_submission_spectrum_type_custom",
        "wizard_submission_spectra_note",
        "wizard_submission_spectra_uploads",
        "wizard_journal_name",
        "wizard_article_title",
        "wizard_publication_year",
        "wizard_volume",
        "wizard_issue",
        "wizard_pages",
        "wizard_doi",
        "wizard_curation_status",
        "wizard_note",
    ]
    for key in wizard_keys:
        if key in st.session_state:
            del st.session_state[key]
        draft_key = f"_draft_{key}"
        if draft_key in st.session_state:
            del st.session_state[draft_key]
    st.session_state["compound_wizard_step"] = 1


def persist_wizard_inputs():
    wizard_keys = [
        "wizard_trivial_name",
        "wizard_iupac_name",
        "wizard_formula",
        "wizard_molecular_weight",
        "wizard_smiles",
        "wizard_inchi",
        "wizard_inchikey",
        "wizard_compound_class_select",
        "wizard_compound_class_custom",
        "wizard_compound_subclass_select",
        "wizard_compound_subclass_custom",
        "wizard_data_source_select",
        "wizard_data_source_custom",
        "wizard_source_category_select",
        "wizard_source_category_custom",
        "wizard_source_organism",
        "wizard_sample_code",
        "wizard_collection_location",
        "wizard_gps_coordinates",
        "wizard_depth_m",
        "wizard_uv_data",
        "wizard_ftir_data",
        "wizard_cd_data",
        "wizard_optical_rotation",
        "wizard_melting_point",
        "wizard_crystallization_method",
        "wizard_ccdc_number",
        "wizard_hrms_data",
        "wizard_structure_path",
        "wizard_structure_upload",
        "wizard_submission_spectrum_type_select",
        "wizard_submission_spectrum_type_custom",
        "wizard_submission_spectra_note",
        "wizard_submission_spectra_uploads",
        "wizard_journal_name",
        "wizard_article_title",
        "wizard_publication_year",
        "wizard_volume",
        "wizard_issue",
        "wizard_pages",
        "wizard_doi",
        "wizard_curation_status",
        "wizard_note",
    ]
    for key in wizard_keys:
        if key in st.session_state:
            st.session_state[f"_draft_{key}"] = st.session_state[key]


def get_wizard_value(key: str, default=""):
    draft_key = f"_draft_{key}"
    if draft_key in st.session_state:
        return st.session_state[draft_key]
    return st.session_state.get(key, default)


def hydrate_wizard_widget(key: str, default=""):
    draft_key = f"_draft_{key}"
    if key not in st.session_state:
        if draft_key in st.session_state:
            st.session_state[key] = st.session_state[draft_key]
        else:
            st.session_state[key] = default

def keyword_search_mask(df: pd.DataFrame, keyword: str) -> pd.Series:
    searchable_columns = [
        "trivial_name",
        "iupac_name",
        "smiles",
        "inchi",
        "inchikey",
        "sample_code",
        "source_category",
        "source_organism",
        "source_material",
        "collection_location",
        "compound_class",
        "compound_subclass",
        "journal_name",
        "doi",
        "note",
    ]
    combined = df[searchable_columns].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
    tokens = [token.lower() for token in re.split(r"\s+", keyword.strip()) if token]

    if not tokens:
        return pd.Series([True] * len(df), index=df.index)

    mask = pd.Series([True] * len(df), index=df.index)
    for token in tokens:
        mask &= combined.str.contains(re.escape(token), regex=True)
    return mask


def field_search_mask(df: pd.DataFrame, keyword: str, field_label: str, match_mode: str) -> pd.Series:
    columns = SEARCH_FIELD_COLUMN_MAP.get(field_label, SEARCH_FIELD_COLUMN_MAP["All searchable fields"])
    combined = df[columns].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
    query = maybe_blank(keyword).lower()
    if not query:
        return pd.Series([True] * len(df), index=df.index)

    if match_mode == "Exact phrase":
        return combined.str.contains(re.escape(query), regex=True)

    if match_mode == "Starts with":
        return combined.str.startswith(query)

    tokens = [token.lower() for token in re.split(r"\s+", query.strip()) if token]
    mask = pd.Series([True] * len(df), index=df.index)
    for token in tokens:
        mask &= combined.str.contains(re.escape(token), regex=True)
    return mask


def decode_uploaded_text(uploaded_file) -> str:
    if uploaded_file is None:
        return ""
    raw = uploaded_file.getvalue()
    for encoding in ["utf-8", "latin-1", "utf-16"]:
        try:
            return raw.decode(encoding)
        except Exception:
            continue
    return raw.decode("utf-8", errors="ignore")


def parse_peak_upload(uploaded_file) -> tuple[list[float], str]:
    if uploaded_file is None:
        return [], ""

    text = decode_uploaded_text(uploaded_file)
    suffix = Path(uploaded_file.name).suffix.lower()
    if not text.strip():
        return [], "Uploaded peak file is empty."

    if suffix in {".dx", ".jdx", ".jcamp"}:
        in_data_block = False
        peaks = []
        for line in text.splitlines():
            stripped = line.strip()
            upper = stripped.upper()
            if upper.startswith("##PEAK TABLE") or upper.startswith("##XYDATA") or upper.startswith("##XYPOINTS"):
                in_data_block = True
                continue
            if upper.startswith("##END"):
                in_data_block = False
                continue
            if upper.startswith("##") or not in_data_block or not stripped:
                continue
            values = re.findall(r"[-+]?\d*\.?\d+(?:[Ee][-+]?\d+)?", stripped)
            if not values:
                continue
            numeric = [float(value) for value in values]
            peaks.extend(numeric[::2] if len(numeric) > 1 else numeric)
        if peaks:
            return peaks, f"Read {len(peaks)} peak position(s) from `{uploaded_file.name}`."

    peaks = parse_peak_input(text)
    if peaks:
        return peaks, f"Read {len(peaks)} peak value(s) from `{uploaded_file.name}`."
    return [], f"No usable numeric peak values were detected in `{uploaded_file.name}`."


def is_structure_backend_available() -> bool:
    return Chem is not None and DataStructs is not None and AllChem is not None


def smiles_to_mol(smiles_value: str):
    if not is_structure_backend_available():
        return None
    smiles_text = maybe_blank(smiles_value)
    if not smiles_text:
        return None
    try:
        return Chem.MolFromSmiles(smiles_text)
    except Exception:
        return None


def structure_text_to_mol(structure_value: str):
    if not is_structure_backend_available():
        return None
    structure_text = maybe_blank(structure_value)
    if not structure_text:
        return None

    mol = smiles_to_mol(structure_text)
    if mol is not None:
        return mol

    try:
        return Chem.MolFromMolBlock(structure_text, sanitize=True, removeHs=True)
    except Exception:
        return None


def canonicalize_smiles(smiles_value: str) -> str:
    mol = structure_text_to_mol(smiles_value)
    if mol is None:
        return ""
    try:
        return Chem.MolToSmiles(mol, canonical=True)
    except Exception:
        return ""


def molecule_similarity_score(query_mol, candidate_mol) -> float:
    if not is_structure_backend_available() or query_mol is None or candidate_mol is None:
        return 0.0
    query_fp = AllChem.GetMorganFingerprintAsBitVect(query_mol, radius=2, nBits=2048)
    candidate_fp = AllChem.GetMorganFingerprintAsBitVect(candidate_mol, radius=2, nBits=2048)
    return float(DataStructs.TanimotoSimilarity(query_fp, candidate_fp))


def molecule_fingerprint(mol):
    if not is_structure_backend_available() or mol is None:
        return None
    try:
        return AllChem.GetMorganFingerprintAsBitVect(mol, radius=2, nBits=2048)
    except Exception:
        return None


@st.cache_resource(show_spinner=False)
def build_structure_candidate_index(_db_signature: float):
    if not is_structure_backend_available():
        return []

    candidates = []
    compounds_df = load_all_compounds()
    for _, row in compounds_df.iterrows():
        candidate_smiles = maybe_blank(row.get("smiles"))
        if not candidate_smiles:
            continue
        candidate_mol = smiles_to_mol(candidate_smiles)
        if candidate_mol is None:
            continue
        candidate_fp = molecule_fingerprint(candidate_mol)
        candidates.append(
            {
                "id": int(row.get("id")),
                "row": row.to_dict(),
                "smiles": candidate_smiles,
                "mol": candidate_mol,
                "fp": candidate_fp,
                "canonical": canonicalize_smiles(candidate_smiles),
            }
        )
    return candidates


def search_by_structure(
    compounds_df: pd.DataFrame,
    query_smiles: str,
    search_type: str,
    similarity_threshold: float = 0.35,
):
    query_text = maybe_blank(query_smiles)
    if not query_text:
        return [], "Please draw a structure and click Apply in the editor first, or paste a valid SMILES / Molfile query."

    if not is_structure_backend_available():
        return [], "Structure search requires RDKit. Install `rdkit>=2026.3` in both requirements.txt files before using this feature."

    query_mol = structure_text_to_mol(query_text)
    if query_mol is None:
        return [], "The structure could not be parsed. Please redraw the query or paste a valid SMILES / Molfile structure."

    query_canonical = canonicalize_smiles(query_text)
    query_fp = molecule_fingerprint(query_mol)
    allowed_ids = set(compounds_df["id"].astype(int).tolist()) if "id" in compounds_df.columns else set()
    indexed_candidates = build_structure_candidate_index(get_db_signature())
    results = []
    searchable_candidates = 0

    for candidate in indexed_candidates:
        if allowed_ids and int(candidate["id"]) not in allowed_ids:
            continue
        candidate_smiles = candidate["smiles"]
        candidate_mol = candidate["mol"]
        if candidate_mol is None:
            continue
        searchable_candidates += 1

        matched = False
        score = 0.0
        match_label = ""

        if search_type == "Identity Search":
            candidate_canonical = candidate.get("canonical", "")
            matched = bool(query_canonical and candidate_canonical and query_canonical == candidate_canonical)
            score = 1.0 if matched else 0.0
            match_label = "Identity"
        elif search_type == "Substructure Search":
            try:
                matched = candidate_mol.HasSubstructMatch(query_mol)
            except Exception:
                matched = False
            score = 1.0 if matched else 0.0
            match_label = "Substructure"
        else:
            candidate_fp = candidate.get("fp")
            if query_fp is None or candidate_fp is None:
                score = 0.0
            else:
                score = float(DataStructs.TanimotoSimilarity(query_fp, candidate_fp))
            matched = score >= similarity_threshold
            match_label = "Similarity"

        if matched:
            item = dict(candidate["row"])
            item["structure_score"] = score * 100
            item["structure_match_type"] = match_label
            item["query_smiles"] = query_text
            item["matched_smiles"] = candidate_smiles
            results.append(item)

    if search_type == "Similarity Search":
        results.sort(
            key=lambda item: (
                item.get("structure_score", 0.0),
                maybe_blank(item.get("trivial_name")).lower(),
            ),
            reverse=True,
        )
    else:
        results.sort(
            key=lambda item: (
                maybe_blank(item.get("trivial_name")).lower(),
                int(item.get("id", 0)),
            )
        )

    if searchable_candidates == 0:
        return [], (
            f"No searchable structures are available yet in the current filtered dataset. Searchable compounds right now: {searchable_candidates}. "
            "Structure search compares your drawn query against compounds that already have SMILES filled in, "
            "so please add SMILES to your records first or use the admin shortcut to save the current drawn structure into a compound record."
        )

    if not results:
        return [], (
            f"No compounds matched this {search_type.lower()} query in the current filtered dataset. Searchable compounds checked: {searchable_candidates}. "
            "Try lowering the similarity threshold, changing filters, or saving structure identifiers for more compounds first."
        )

    return results, ""


def export_structure_search_results(results: list[dict]) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(results, start=1):
        rows.append(
            {
                "Rank": i,
                "Compound ID": item.get("id"),
                "Trivial Name": clean_text(item.get("trivial_name")),
                "Molecular Formula": clean_text(item.get("molecular_formula")),
                "Compound Class": clean_text(item.get("compound_class")),
                "Source Category": clean_text(item.get("source_category")),
                "Source Organism": clean_text(item.get("source_organism")),
                "Source Summary": clean_text(source_summary_from_record(item)),
                "Match Type": clean_text(item.get("structure_match_type")),
                "Score (%)": round(float(item.get("structure_score", 0.0)), 2),
                "SMILES": clean_text(item.get("matched_smiles")),
            }
        )
    return pd.DataFrame(rows)


def render_structure_search_results(results: list[dict], search_type: str, limit: int = 10):
    if not results:
        st.info("No compounds matched the current structure query.")
        return

    section_header("Structure Search Results", f"Showing the top {min(limit, len(results))} candidate(s) for {search_type.lower()}.")
    st.caption(f"Results: {len(results)}")

    for i, item in enumerate(results[:limit], start=1):
        title_text = clean_text(item.get("trivial_name"))
        title = html_text(title_text)
        formula = html_text(item.get("molecular_formula"))
        compound_class = html_text(item.get("compound_class"))
        source_summary = html_text(source_summary_from_record(item))
        score = float(item.get("structure_score", 0.0))
        subtitle = html_text(f"{item.get('structure_match_type', search_type)} match | Score: {score:.1f}%")
        compound_id = html_text(item.get("id"))
        matched_smiles = html_text(item.get("matched_smiles"))

        with st.expander(f"#{i} · {title_text}", expanded=(i == 1)):
            st.markdown('<div class="result-card">', unsafe_allow_html=True)
            st.markdown(
                f"""
                <div class="result-title">{title}</div>
                <div class="result-subtitle">{subtitle}</div>
                """,
                unsafe_allow_html=True,
            )
            preview_col, meta_col = st.columns([1, 1.3])
            with preview_col:
                render_structure_preview(item.get("matched_smiles"), caption=f"Query candidate #{i}", size=(380, 260))
            with meta_col:
                st.markdown('<div class="structure-result-meta">', unsafe_allow_html=True)
                st.markdown(f'<div class="structure-result-stat"><strong>Compound ID:</strong> {compound_id}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="structure-result-stat"><strong>Molecular Formula:</strong> {formula}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="structure-result-stat"><strong>Compound Class:</strong> {compound_class}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="structure-result-stat"><strong>Source:</strong> {source_summary}</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="structure-result-stat"><strong>SMILES:</strong> {matched_smiles}</div>', unsafe_allow_html=True)
                st.progress(min(max(score / 100.0, 0.0), 1.0))
                st.markdown('</div>', unsafe_allow_html=True)

            action_left, action_right = st.columns([1, 1])
            with action_left:
                if st.button(f"Open Record #{i}", key=f"open_structure_result_{item.get('id')}_{i}"):
                    open_compound_detail(int(item["id"]))
                    st.rerun()
            with action_right:
                if can_edit_database():
                    if st.button(f"Update Metadata #{i}", key=f"edit_structure_result_{item.get('id')}_{i}"):
                        open_compound_editor(int(item["id"]))
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

def calculate_completeness_score(compound_row, proton_df, carbon_df, spectra_df):
    row = compound_row.iloc[0] if isinstance(compound_row, pd.DataFrame) else compound_row
    checks = [
        bool(maybe_blank(row.get("trivial_name"))),
        bool(maybe_blank(row.get("molecular_formula"))),
        bool(maybe_blank(row.get("smiles")) or maybe_blank(row.get("inchi")) or maybe_blank(row.get("inchikey"))),
        bool(maybe_blank(row.get("compound_class"))),
        bool(
            maybe_blank(row.get("source_category"))
            or maybe_blank(row.get("source_organism"))
            or maybe_blank(row.get("source_material"))
        ),
        bool(maybe_blank(row.get("data_source"))),
        bool(maybe_blank(row.get("hrms_data"))),
        bool(maybe_blank(row.get("doi")) or maybe_blank(row.get("journal_name"))),
        bool(maybe_blank(row.get("structure_image_path"))),
        not proton_df.empty,
        not carbon_df.empty,
        not spectra_df.empty,
    ]
    completed = sum(1 for item in checks if item)
    return round((completed / len(checks)) * 100)

def parse_peak_input(text: str):
    peaks = []
    for item in re.split(r"[\s,;]+", maybe_blank(text)):
        if not item:
            continue
        try:
            peaks.append(float(item))
        except ValueError:
            pass
    return peaks

def find_best_matches(query_peaks, db_peaks, tolerance):
    matched_query_peaks = []
    matched_db_indexes = set()

    for q in query_peaks:
        best_match = None
        best_diff = None
        best_index = None

        for i, db_peak in enumerate(db_peaks):
            if i in matched_db_indexes:
                continue

            diff = abs(q - db_peak)
            if diff <= tolerance:
                if best_diff is None or diff < best_diff:
                    best_diff = diff
                    best_match = db_peak
                    best_index = i

        if best_match is not None:
            matched_query_peaks.append((q, best_match, best_diff))
            matched_db_indexes.add(best_index)

    return matched_query_peaks

def is_external_url(value) -> bool:
    text = maybe_blank(value)
    if not text:
        return False
    parsed = urlparse(text)
    return parsed.scheme in {"http", "https"}


def is_google_drive_url(value) -> bool:
    text = maybe_blank(value).lower()
    return "drive.google.com" in text or "docs.google.com" in text


def extract_google_drive_file_id(value) -> str:
    text = maybe_blank(value)
    if not text:
        return ""

    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"[?&]id=([a-zA-Z0-9_-]+)",
        r"/d/([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def google_drive_preview_url(value) -> str:
    file_id = extract_google_drive_file_id(value)
    if not file_id:
        return ""
    return f"https://drive.google.com/thumbnail?id={file_id}&sz=w2000"


def google_drive_download_url(value) -> str:
    file_id = extract_google_drive_file_id(value)
    if not file_id:
        return value
    return f"https://drive.google.com/uc?export=download&id={file_id}"


def parse_supabase_storage_reference(value) -> tuple[str, str] | None:
    text = maybe_blank(value)
    if not text:
        return None
    if text.startswith("storage://"):
        bucket, _, object_path = text.removeprefix("storage://").partition("/")
        return (bucket, object_path) if bucket and object_path else None
    if not is_external_url(text):
        return None
    supabase_url = get_supabase_url()
    if not supabase_url:
        return None
    parsed = urlparse(text)
    supabase_host = urlparse(supabase_url).netloc
    if parsed.netloc != supabase_host:
        return None
    marker = "/storage/v1/object/public/"
    if marker not in parsed.path:
        return None
    bucket, _, object_path = parsed.path.split(marker, 1)[1].partition("/")
    if not bucket or not object_path:
        return None
    return unquote(bucket), unquote(object_path)


def is_supabase_storage_reference(value) -> bool:
    return parse_supabase_storage_reference(value) is not None


@st.cache_data(show_spinner=False, ttl=900)
def signed_supabase_storage_url(bucket: str, object_path: str, expires_in: int = 3600) -> str:
    if not use_supabase_write_backend():
        return ""
    response = _supabase_request(
        "POST",
        f"/storage/v1/object/sign/{bucket}/{quote(object_path, safe='/')}",
        body={"expiresIn": expires_in},
        write=True,
    )
    signed_url = response.get("signedURL", "") if isinstance(response, dict) else ""
    if not signed_url:
        return ""
    if signed_url.startswith("http"):
        return signed_url
    return f"{get_supabase_url().rstrip()}{signed_url}"


def display_asset_url(value) -> str:
    text = maybe_blank(value)
    reference = parse_supabase_storage_reference(text)
    if not reference:
        return text
    bucket, object_path = reference
    signed_url = signed_supabase_storage_url(bucket, object_path)
    return signed_url or text


def can_preview_external_image(file_path_value, spectrum_type_value="") -> bool:
    path_text = maybe_blank(file_path_value).lower()
    spectrum_text = maybe_blank(spectrum_type_value).lower()

    if any(raw_token in spectrum_text for raw_token in ["raw", "jcamp", "mnova", "fid"]):
        return False

    if is_google_drive_url(path_text):
        return True

    return path_text.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif"))


def get_full_file_path(relative_path):
    if relative_path is None:
        return None
    relative_path = str(relative_path).strip()
    if not relative_path:
        return None
    if is_external_url(relative_path):
        return None
    candidate = Path(relative_path)
    if candidate.is_absolute():
        return candidate
    return PROJECT_DIR / candidate

def is_image_file(path: Path):
    return path.suffix.lower() in [".png", ".jpg", ".jpeg", ".webp"]

def is_pdf_file(path: Path):
    return path.suffix.lower() == ".pdf"

def normalize_filter_value(value):
    text = clean_text(value)
    return text if text != "-" else None

def build_filter_options(df, column_name):
    values = []
    for value in df[column_name].tolist():
        normalized = normalize_filter_value(value)
        if normalized is not None:
            values.append(normalized)
    return ["All"] + sorted(set(values))

def apply_dataframe_filters(
    df,
    class_filter="All",
    subclass_filter="All",
    source_filter="All",
    data_source_filter="All"
):
    result = df.copy()

    if class_filter != "All":
        result = result[result["compound_class"].fillna("").astype(str).str.strip() == class_filter]

    if subclass_filter != "All":
        result = result[result["compound_subclass"].fillna("").astype(str).str.strip() == subclass_filter]

    if source_filter != "All":
        result = result[result["source_category"].fillna("").astype(str).str.strip() == source_filter]

    if data_source_filter != "All":
        result = result[result["data_source"].fillna("").astype(str).str.strip() == data_source_filter]

    return result

def filter_similarity_results(results, class_filter="All", source_filter="All", data_source_filter="All"):
    filtered = []

    for item in results:
        ok = True

        if class_filter != "All" and clean_text(item.get("compound_class")) != class_filter:
            ok = False

        item_source_category = normalize_source_category(item.get("source_category"))
        if not item_source_category:
            item_source_category = normalize_source_category(item.get("source_material"))
        if source_filter != "All" and clean_text(item_source_category) != source_filter:
            ok = False

        if data_source_filter != "All" and clean_text(item.get("data_source")) != data_source_filter:
            ok = False

        if ok:
            filtered.append(item)

    return filtered

def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    if Alignment is None and Font is None and PatternFill is None:
        raise ModuleNotFoundError("openpyxl is not available")

    output = io.BytesIO()
    export_df = df.copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.book[sheet_name]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 42)

        footer_row = worksheet.max_row + 2
        footer_col_end = max(1, worksheet.max_column)
        worksheet.cell(row=footer_row, column=1, value=OWNER_CREDIT)
        if footer_col_end > 1:
            worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=footer_col_end)

        footer_cell = worksheet.cell(row=footer_row, column=1)
        if Font is not None:
            footer_cell.font = Font(italic=True, size=10, color="4F5B6B")
        if Alignment is not None:
            footer_cell.alignment = Alignment(horizontal="right")
        if PatternFill is not None:
            footer_cell.fill = PatternFill(fill_type="solid", fgColor="F5F8FD")

        try:
            worksheet.oddFooter.right.text = OWNER_CREDIT
        except Exception:
            pass

    output.seek(0)
    return output.getvalue()


def download_dataframe_button(label: str, df: pd.DataFrame, file_name: str, key: str, sheet_name: str = "Data"):
    try:
        payload = dataframe_to_excel_bytes(df, sheet_name=sheet_name)
        resolved_name = file_name if file_name.lower().endswith(".xlsx") else f"{Path(file_name).stem}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    except Exception:
        payload = add_credit_to_text_bytes(dataframe_to_csv_bytes(df))
        resolved_name = f"{Path(file_name).stem}.csv"
        mime = "text/csv"

    st.download_button(
        label=label,
        data=payload,
        file_name=resolved_name,
        mime=mime,
        key=key,
    )


def add_credit_to_text_bytes(payload: bytes) -> bytes:
    text = payload.decode("utf-8")
    text = text.rstrip() + f"\n\n{OWNER_CREDIT}\n"
    return text.encode("utf-8")


def render_app_credit_footer():
    st.markdown(
        f'<div style="text-align:center;"><div class="app-credit-footer">{OWNER_CREDIT}</div></div>',
        unsafe_allow_html=True,
    )


def render_sidebar_credit():
    st.markdown(
        f'<div class="sidebar-credit-wrap"><div class="app-credit-footer">{OWNER_CREDIT}</div></div>',
        unsafe_allow_html=True,
    )


def normalize_structure_image(image_obj, size=(520, 360)):
    if Image is None or ImageOps is None or image_obj is None:
        return image_obj
    try:
        image = image_obj.convert("RGBA")
        contained = ImageOps.contain(image, size, method=Image.Resampling.LANCZOS)
        canvas = Image.new("RGBA", size, (255, 255, 255, 255))
        x = (size[0] - contained.width) // 2
        y = (size[1] - contained.height) // 2
        canvas.paste(contained, (x, y), contained)
        return canvas.convert("RGB")
    except Exception:
        return image_obj


def pil_image_to_data_uri(image_obj) -> str:
    data = pil_image_to_png_bytes(image_obj)
    if not data:
        return ""
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def pil_image_to_png_bytes(image_obj) -> bytes:
    if image_obj is None:
        return b""
    try:
        output = io.BytesIO()
        image = image_obj.convert("RGBA") if hasattr(image_obj, "convert") else image_obj
        image.save(output, format="PNG", optimize=True)
        return output.getvalue()
    except Exception:
        return b""


@st.cache_data(show_spinner=False)
def structure_smiles_png_bytes(smiles_text: str, size=(300, 220)) -> bytes:
    if Chem is None or Draw is None:
        return b""
    smiles_value = maybe_blank(smiles_text)
    if not smiles_value:
        return b""
    try:
        mol = structure_text_to_mol(smiles_value)
        if mol is None:
            return b""
        image = normalize_structure_image(Draw.MolToImage(mol, size=size), size=size)
        return pil_image_to_png_bytes(image)
    except Exception:
        return b""


@st.cache_data(show_spinner=False)
def structure_smiles_data_uri(smiles_text: str, size=(300, 220)) -> str:
    data = structure_smiles_png_bytes(smiles_text, size=size)
    if not data:
        return ""
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def structure_smiles_image_url(smiles_text: str) -> str:
    smiles_value = maybe_blank(smiles_text)
    if not smiles_value:
        return ""
    return f"https://cactus.nci.nih.gov/chemical/structure/{quote(smiles_value, safe='')}/image"


def load_standardized_structure_image(image_path: Path, size=(520, 360)):
    if Image is None or image_path is None or not image_path.exists():
        return None
    try:
        with Image.open(image_path) as image:
            return normalize_structure_image(image, size=size)
    except Exception:
        return None


def load_standardized_structure_source(source_value, size=(520, 360)):
    if Image is None or source_value is None:
        return None
    source_text = str(source_value).strip()
    if not source_text:
        return None

    if is_external_url(source_text):
        try:
            with urllib.request.urlopen(display_asset_url(source_text), timeout=30, context=_supabase_ssl_context()) as response:
                raw = response.read()
            with Image.open(io.BytesIO(raw)) as image:
                return normalize_structure_image(image, size=size)
        except Exception:
            return None

    full_path = get_full_file_path(source_text)
    if full_path is None or not full_path.exists():
        return None
    return load_standardized_structure_image(full_path, size=size)


def render_structure_preview(smiles_text: str, caption: str | None = None, empty_message: bool = True, size=(520, 360)):
    smiles_value = maybe_blank(smiles_text)
    if not smiles_value:
        if empty_message:
            st.info("No structure preview available for this record.")
        return
    if Chem is None or Draw is None:
        fallback_url = structure_smiles_image_url(smiles_value)
        if fallback_url:
            st.image(fallback_url, caption=caption or "Rendered from SMILES", width="stretch")
        elif empty_message:
            st.info("Structure preview becomes available when RDKit drawing support is active.")
        return
    try:
        mol = structure_text_to_mol(smiles_value)
        if mol is None:
            fallback_url = structure_smiles_image_url(smiles_value)
            if fallback_url:
                st.image(fallback_url, caption=caption or "Rendered from SMILES", width="stretch")
            elif empty_message:
                st.info("Stored structure could not be rendered from the available structure string.")
            return
        image = normalize_structure_image(Draw.MolToImage(mol, size=size), size=size)
        st.image(image, caption=caption, width="stretch")
    except Exception:
        fallback_url = structure_smiles_image_url(smiles_value)
        if fallback_url:
            st.image(fallback_url, caption=caption or "Rendered from SMILES", width="stretch")
        elif empty_message:
            st.info("Structure preview could not be rendered for this record.")

def get_backup_bytes():
    with open(DB_PATH, "rb") as f:
        return f.read()

def count_related_records(filtered_ids):
    filtered_ids = [int(item) for item in filtered_ids if str(item).strip()]
    if not filtered_ids:
        return 0, 0, 0
    conn = get_connection()

    try:
        placeholders = ",".join("?" * len(filtered_ids))
        proton_query = f"SELECT COUNT(*) AS n FROM proton_nmr WHERE compound_id IN ({placeholders})"
        carbon_query = f"SELECT COUNT(*) AS n FROM carbon_nmr WHERE compound_id IN ({placeholders})"
        spectra_query = f"SELECT COUNT(*) AS n FROM spectra_files WHERE compound_id IN ({placeholders})"

        proton_count = int(pd.read_sql_query(proton_query, conn, params=filtered_ids)["n"][0])
        carbon_count = int(pd.read_sql_query(carbon_query, conn, params=filtered_ids)["n"][0])
        spectra_count = int(pd.read_sql_query(spectra_query, conn, params=filtered_ids)["n"][0])
        return proton_count, carbon_count, spectra_count
    finally:
        conn.close()


def count_bioactivity_records(filtered_ids):
    filtered_ids = [int(item) for item in filtered_ids if str(item).strip()]
    if not filtered_ids:
        return 0
    conn = get_connection()
    try:
        placeholders = ",".join("?" * len(filtered_ids))
        query = f"SELECT COUNT(*) AS n FROM bioactivity_records WHERE compound_id IN ({placeholders})"
        return int(pd.read_sql_query(query, conn, params=filtered_ids)["n"][0])
    finally:
        conn.close()


@st.cache_data(show_spinner=False)
def count_database_totals(_db_signature: float):
    if use_supabase_backend():
        compounds_df = load_all_compounds()
        proton_df = load_all_proton_data()
        carbon_df = load_all_carbon_data()
        spectra_df = load_all_spectra_files()
        bioactivity_df = load_all_bioactivity_data()
        structure_columns = ["smiles", "inchi", "inchikey", "structure_image_path"]
        structures_count = 0
        if not compounds_df.empty:
            available_columns = [column for column in structure_columns if column in compounds_df.columns]
            if available_columns:
                structures_count = int(
                    compounds_df[available_columns]
                    .fillna("")
                    .astype(str)
                    .apply(lambda row: any(value.strip() for value in row), axis=1)
                    .sum()
                )
        return {
            "compounds": int(len(compounds_df)),
            "structures": structures_count,
            "proton": int(len(proton_df)),
            "carbon": int(len(carbon_df)),
            "spectra": int(len(spectra_df)),
            "bioactivity": int(len(bioactivity_df)),
        }

    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM compounds")
        compounds_count = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT COUNT(*) FROM proton_nmr")
        proton_count = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT COUNT(*) FROM carbon_nmr")
        carbon_count = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT COUNT(*) FROM spectra_files")
        spectra_count = int(cursor.fetchone()[0] or 0)
        cursor.execute(
            """
            SELECT COUNT(*) FROM compounds
            WHERE TRIM(COALESCE(smiles, '')) != ''
               OR TRIM(COALESCE(inchi, '')) != ''
               OR TRIM(COALESCE(inchikey, '')) != ''
               OR TRIM(COALESCE(structure_image_path, '')) != ''
            """
        )
        structures_count = int(cursor.fetchone()[0] or 0)
        cursor.execute("SELECT COUNT(*) FROM bioactivity_records")
        bioactivity_count = int(cursor.fetchone()[0] or 0)
        return {
            "compounds": compounds_count,
            "structures": structures_count,
            "proton": proton_count,
            "carbon": carbon_count,
            "spectra": spectra_count,
            "bioactivity": bioactivity_count,
        }
    finally:
        conn.close()


def calculate_workspace_health(compounds_df: pd.DataFrame):
    compounds_df = enrich_compounds_dataframe(compounds_df)
    if compounds_df.empty:
        return {
            "structure_ready": 0,
            "reference_ready": 0,
            "external_ready": 0,
            "submission_ready": 0,
            "bioactivity_ready": 0,
        }

    structure_ready = compounds_df[
        compounds_df["smiles"].fillna("").astype(str).str.strip().ne("")
        | compounds_df["inchi"].fillna("").astype(str).str.strip().ne("")
        | compounds_df["inchikey"].fillna("").astype(str).str.strip().ne("")
    ]
    reference_ready = compounds_df[
        compounds_df["doi"].fillna("").astype(str).str.strip().ne("")
        | compounds_df["journal_name"].fillna("").astype(str).str.strip().ne("")
    ]

    spectra_df = load_all_spectra_files()
    external_ready_ids = set(
        spectra_df[spectra_df["file_path"].fillna("").astype(str).apply(is_external_url)]["compound_id"].tolist()
    )
    bioactivity_df = load_all_bioactivity_data()
    bioactivity_ready_ids = set(bioactivity_df["compound_id"].tolist()) if not bioactivity_df.empty else set()
    submission_ready = compounds_df[
        compounds_df["trivial_name"].fillna("").astype(str).str.strip().ne("")
        & compounds_df["compound_class"].fillna("").astype(str).str.strip().ne("")
        & (
            compounds_df["source_category"].fillna("").astype(str).str.strip().ne("")
            | compounds_df["source_organism"].fillna("").astype(str).str.strip().ne("")
            | compounds_df["source_material"].fillna("").astype(str).str.strip().ne("")
        )
    ]

    return {
        "structure_ready": int(len(structure_ready)),
        "reference_ready": int(len(reference_ready)),
        "external_ready": int(len(external_ready_ids)),
        "submission_ready": int(len(submission_ready)),
        "bioactivity_ready": int(len(bioactivity_ready_ids)),
    }

# =========================
# UI helpers
# =========================
def section_header(title, subtitle=""):
    st.markdown(f'<div class="section-title">{html_text(title)}</div>', unsafe_allow_html=True)
    if subtitle:
        st.markdown(f'<div class="section-subtitle">{html_text(subtitle)}</div>', unsafe_allow_html=True)

def render_metric_card(label, value, col):
    with col:
        st.markdown(
            f"""
            <div class="metric-card">
                <div class="metric-card-label">{html_text(label)}</div>
                <div class="metric-card-value">{html_text(value)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_workspace_headbar(active_section: str):
    if active_section == "Dashboard":
        return
    hero_uri = image_to_data_uri(str(HERO_BANNER_PATH), max_px=1300) if HERO_BANNER_PATH.exists() else ""
    headbar_class = "workspace-headbar has-image" if hero_uri else "workspace-headbar"
    headbar_style = (
        "background-image: "
        "linear-gradient(112deg, rgba(5, 11, 26, 0.92), rgba(8, 18, 34, 0.48)), "
        f"url('{hero_uri}');"
    ) if hero_uri else ""
    section_label = html.escape(clean_text(active_section))
    render_raw_html(
        f"""
        <div class="{headbar_class}" style="{headbar_style}">
            <div class="workspace-headbar-copy">
                <div class="workspace-headbar-kicker">{section_label}</div>
                <div class="workspace-headbar-title">Natural Products<br><span>Spectral Database</span></div>
            </div>
        </div>
        """
    )


def render_clean_stat(label, value, col):
    with col:
        st.markdown(
            f"""
            <div class="clean-stat">
                <div class="clean-stat-label">{html_text(label)}</div>
                <div class="clean-stat-value">{html_text(value)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_helper_card(title, text):
    st.markdown(
        f"""
        <div class="helper-card">
            <div class="helper-title">{html_text(title)}</div>
            <div class="helper-text">{html_text(text)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_selector_card(title, subtitle):
    st.markdown(
        f"""
        <div class="selector-card">
            <div class="selector-title">{html_text(title)}</div>
            <div class="selector-subtitle">{html_text(subtitle)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_external_link_card(label: str, url: str, note: str | None = None):
    st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    st.markdown(f"**{label}**")
    if is_google_drive_url(url):
        st.markdown(f"[Open file]({url})")
        st.markdown(f"[Download file]({google_drive_download_url(url)})")
    else:
        st.markdown(f"[Open external file]({url})")
    if note:
        st.caption(note)
    st.markdown('</div>', unsafe_allow_html=True)


def render_dashboard_bar_chart(df: pd.DataFrame, x_col: str, y_col: str, color_hex: str = "#61D8ED"):
    if df.empty:
        st.info("No data available.")
        return

    chart_df = df[[x_col, y_col]].copy()
    chart_df[x_col] = chart_df[x_col].fillna("Uncategorized").astype(str).str.strip()
    chart_df[x_col] = chart_df[x_col].replace("", "Uncategorized")
    chart_df[y_col] = pd.to_numeric(chart_df[y_col], errors="coerce").fillna(0)
    chart_df = chart_df.sort_values(by=y_col, ascending=False).set_index(x_col)

    st.bar_chart(chart_df[[y_col]], color=color_hex)


def limit_chart_categories(df: pd.DataFrame, names_col: str, values_col: str, top_n: int = 8) -> pd.DataFrame:
    chart_df = df[[names_col, values_col]].copy()
    chart_df[names_col] = chart_df[names_col].fillna("Uncategorized").astype(str).str.strip().replace("", "Uncategorized")
    chart_df[values_col] = pd.to_numeric(chart_df[values_col], errors="coerce").fillna(0)
    chart_df = chart_df[chart_df[values_col] > 0].sort_values(values_col, ascending=False)
    if len(chart_df) <= top_n:
        return chart_df
    top_df = chart_df.head(top_n).copy()
    other_total = chart_df.iloc[top_n:][values_col].sum()
    if other_total > 0:
        other_label = "Other categories" if "source" in names_col.lower() else "Other classes"
        top_df = pd.concat(
            [top_df, pd.DataFrame([{names_col: other_label, values_col: other_total}])],
            ignore_index=True,
        )
    return top_df


def render_dashboard_pie_chart(df: pd.DataFrame, names_col: str, values_col: str, color_sequence: list[str] | None = None, top_n: int = 8):
    if df.empty:
        st.info("No data available.")
        return
    if px is None:
        st.dataframe(limit_chart_categories(df, names_col, values_col, top_n=top_n), width="stretch", hide_index=True)
        return
    chart_df = limit_chart_categories(df, names_col, values_col, top_n=top_n)
    if chart_df.empty:
        st.info("No data available.")
        return

    palette = color_sequence or ["#61D8ED", "#4C8EFF", "#9C63F1", "#FF7F6D", "#F2C66D", "#7EF0C2", "#BFA5FF", "#92F2D7", "#F7A68E"]
    chart_df = chart_df.copy()
    total_value = int(chart_df[values_col].sum())
    chart_df["_pct"] = chart_df[values_col].div(total_value).mul(100) if total_value else 0
    chart_df["_pct_label"] = chart_df["_pct"].apply(lambda value: f"{value:.1f}%" if value >= 5 else "")

    figure = px.pie(
        chart_df,
        names=names_col,
        values=values_col,
        hole=0.58,
        color_discrete_sequence=palette,
        custom_data=["_pct_label"],
    )
    figure.update_traces(
        textposition="inside",
        textinfo="text",
        texttemplate="%{customdata[0]}",
        textfont=dict(size=11, color="#F7FBFF"),
        insidetextorientation="auto",
        sort=False,
        pull=[0.025] + [0] * max(len(chart_df) - 1, 0),
        marker=dict(line=dict(color="rgba(7,17,29,0.95)", width=2)),
        hovertemplate=f"%{{label}}<br>%{{value}} {values_col.lower()}<br>%{{percent}}<extra></extra>",
    )
    figure.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        height=304,
        margin=dict(l=10, r=10, t=4, b=4),
        font=dict(color="#F5F8FD", family="Inter, system-ui, -apple-system, BlinkMacSystemFont, sans-serif"),
        showlegend=False,
        annotations=[
            dict(
                text=f"<b>{total_value:,}</b><br><span style='font-size:12px;color:#AFC0DA'>records</span>",
                x=0.5,
                y=0.5,
                showarrow=False,
                font=dict(size=15, color="#F5F8FD"),
            )
        ],
    )
    st.plotly_chart(figure, width="stretch", config={"displayModeBar": False})

    legend_items = []
    for idx, row in chart_df.iterrows():
        color = palette[idx % len(palette)]
        label = html_text(row[names_col])
        percent = float(row["_pct"])
        legend_items.append(
            f'<div class="chart-legend-item"><span class="chart-legend-swatch" style="background:{color}"></span>'
            f'<span class="chart-legend-label">{label}</span><span class="chart-legend-percent">{percent:.1f}%</span></div>'
        )
    st.markdown(f'<div class="chart-legend-grid">{"".join(legend_items)}</div>', unsafe_allow_html=True)


def format_latest_update_label(df: pd.DataFrame) -> str:
    if df.empty or "updated_at" not in df.columns:
        return "Not available"
    timestamps = pd.to_datetime(df["updated_at"], errors="coerce", utc=True).dropna()
    if timestamps.empty:
        return "Not available"
    latest = timestamps.max().tz_convert(None)
    return f"{latest.month}/{latest.day}/{latest.year}"


def format_metric_value(value) -> str:
    if value is None:
        return "0"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and not float(value).is_integer():
            return f"{value:,.1f}"
        return f"{int(value):,}"
    return str(value)


def build_dashboard_stat_markup(value, label: str, icon_path: Path | None = None) -> str:
    icon_uri = image_to_data_uri(str(icon_path), max_px=96) if icon_path and icon_path.exists() else ""
    icon_markup = f'<img class="dashboard-stat-board-icon" src="{icon_uri}" alt="{label} icon" />' if icon_uri else ""
    if label == "Updated":
        return f"""
            <div class="dashboard-stat-board-item is-updated">
                <div class="dashboard-stat-board-head is-updated-head">
                    {icon_markup}
                    <div class="dashboard-stat-board-copy is-updated">
                        <div class="dashboard-stat-board-label dashboard-stat-board-label-top">{label}</div>
                        <div class="dashboard-stat-board-date">{format_metric_value(value)}</div>
                    </div>
                </div>
            </div>
        """
    return f"""
        <div class="dashboard-stat-board-item">
            <div class="dashboard-stat-board-head">
                {icon_markup}
                <div class="dashboard-stat-board-copy">
                    <div class="dashboard-stat-board-value">{format_metric_value(value)}</div>
                    <div class="dashboard-stat-board-label">{label}</div>
                </div>
            </div>
        </div>
    """


def build_sidebar_stat_markup(value, label: str, icon_path: Path | None = None) -> str:
    icon_uri = image_to_data_uri(str(icon_path), max_px=88) if icon_path and icon_path.exists() else ""
    safe_label = html_text(label)
    icon_markup = f'<img class="sidebar-stat-icon" src="{icon_uri}" alt="{safe_label} icon" />' if icon_uri else ""
    return f"""
        <div class="sidebar-stat">
            <div class="sidebar-stat-head">
                {icon_markup}
                <div class="sidebar-stat-value">{html_text(format_metric_value(value))}</div>
            </div>
            <div class="sidebar-stat-label">{safe_label}</div>
        </div>
    """


def build_workflow_card_icon_markup(title: str) -> str:
    icon_path = WORKFLOW_CARD_ART_PATHS.get(title)
    if not icon_path or not icon_path.exists():
        return ""
    icon_uri = image_to_data_uri(str(icon_path), max_px=128)
    return (
        '<div class="workflow-card-icon-shell">'
        f'<img class="workflow-card-icon" src="{icon_uri}" alt="{title} icon" />'
        '</div>'
    )


def format_sidebar_nav_label(label: str) -> str:
    icon = SIDEBAR_NAV_LABEL_ICONS.get(label, "")
    return f"{icon}  {label}" if icon else label


def build_sidebar_nav_icon_markup(label: str) -> str:
    icon_path = SIDEBAR_NAV_ICON_PATHS.get(label)
    if not icon_path or not icon_path.exists():
        return ""
    icon_uri = image_to_data_uri(str(icon_path), max_px=72)
    return (
        '<div class="sidebar-nav-icon-shell">'
        f'<img class="sidebar-nav-icon" src="{icon_uri}" alt="{label} icon" />'
        '</div>'
    )


def render_sidebar_nav_link(group_title: str, item: dict, is_active: bool):
    target_section = item["section"]
    target_compound_page = item.get("compound_page")
    button_type = "primary" if is_active else "secondary"
    if st.button(
        item["label"],
        key=f"sidebar_nav_{group_title}_{item['label']}_{target_section}_{target_compound_page or 'none'}",
        width="stretch",
        type=button_type,
        icon=SIDEBAR_NAV_LABEL_ICONS.get(item["label"]),
    ):
        navigate_internal(target_section, target_compound_page)
        st.rerun()


def clear_npdb_login_session():
    for key in list(st.session_state.keys()):
        if str(key).startswith("npdb_"):
            st.session_state.pop(key, None)


def render_sidebar_session_controls():
    if not st.session_state.get("npdb_authenticated"):
        return

    username = clean_text(st.session_state.get("npdb_username") or "approved user")
    role = clean_text(st.session_state.get("npdb_role") or "viewer").replace("_", " ").title()
    st.markdown(
        f"""
        <div class="sidebar-session-summary">
            <div class="sidebar-session-title">Account</div>
            <div class="sidebar-session-user">{html.escape(username)} · {html.escape(role)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button("Log out", key="npdb_logout_button", width="stretch", icon=":material/logout:"):
        clear_npdb_login_session()
        st.rerun()


def build_snapshot_manifest() -> dict:
    compounds_df = load_all_compounds()
    proton_df = load_all_proton_data()
    carbon_df = load_all_carbon_data()
    spectra_df = load_all_spectra_files()
    bioactivity_df = load_all_bioactivity_data()
    return {
        "project": "npdb",
        "created_at_utc": datetime.now(UTC).isoformat(),
        "cloud_backend_active": use_supabase_backend(),
        "cloud_write_active": use_supabase_write_backend(),
        "source_of_truth": "Supabase" if use_supabase_write_backend() else "Local SQLite",
        "counts": {
            "compounds": int(len(compounds_df)),
            "proton_nmr": int(len(proton_df)),
            "carbon_nmr": int(len(carbon_df)),
            "spectra_files": int(len(spectra_df)),
            "bioactivity_records": int(len(bioactivity_df)),
        },
    }


def build_backup_bundle_bytes() -> tuple[bytes, str]:
    manifest = build_snapshot_manifest()
    bundle_timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    bundle_name = f"npdb_snapshot_{bundle_timestamp}.zip"
    tables = {
        "compounds": load_all_compounds(),
        "proton_nmr": load_all_proton_data(),
        "carbon_nmr": load_all_carbon_data(),
        "spectra_files": load_all_spectra_files(),
        "bioactivity_records": load_all_bioactivity_data(),
    }
    output = io.BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, indent=2))
        for table_name, df in tables.items():
            archive.writestr(f"tables/{table_name}.csv", df.to_csv(index=False))
        if use_local_read_backend() and DB_PATH.exists():
            archive.writestr("database/nmr.db", get_backup_bytes())
    output.seek(0)
    return output.getvalue(), bundle_name


def save_backup_bundle_locally(bundle_bytes: bytes, file_name: str) -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    bundle_path = BACKUPS_DIR / file_name
    bundle_path.write_bytes(bundle_bytes)
    return bundle_path


def upload_backup_bundle_to_cloud(bundle_bytes: bytes, file_name: str) -> str:
    ensure_write_target_ready()
    if use_local_read_backend():
        save_backup_bundle_locally(bundle_bytes, file_name)
    return supabase_upload_bytes(
        "backups",
        f"snapshots/{datetime.now(UTC).strftime('%Y/%m/%d')}/{file_name}",
        bundle_bytes,
        content_type="application/zip",
        public_bucket=False,
    )


def render_sidebar_workspace_summary(active_section: str, all_compounds_df: pd.DataFrame):
    all_compounds_df = enrich_compounds_dataframe(all_compounds_df)
    totals_snapshot = count_database_totals(get_db_signature())
    total_compounds = int(totals_snapshot["compounds"])
    structures_count = int(totals_snapshot.get("structures", 0))
    proton_count = int(totals_snapshot["proton"])
    carbon_count = int(totals_snapshot["carbon"])
    spectra_count = int(totals_snapshot["spectra"])
    bioactivity_count = int(totals_snapshot["bioactivity"])
    health = calculate_workspace_health(all_compounds_df)
    latest_update = format_latest_update_label(all_compounds_df)
    logo_uri = image_to_data_uri(str(SIDEBAR_LOGO_PATH), max_px=144) if SIDEBAR_LOGO_PATH.exists() else ""
    quality_pct = 0
    if total_compounds:
        quality_pct = int(round((health["submission_ready"] / total_compounds) * 100))
    quality_pct = max(0, min(100, quality_pct))
    stats_markup = "".join(
        [
            build_sidebar_stat_markup(total_compounds, "Compounds", COMPOUNDS_ART_PATH),
            build_sidebar_stat_markup(structures_count, "Structures", STRUCTURES_ART_PATH),
            build_sidebar_stat_markup(proton_count, "1H Peaks", SPECTRA_ART_PATH),
            build_sidebar_stat_markup(carbon_count, "13C Peaks", UPDATED_ART_PATH),
            build_sidebar_stat_markup(spectra_count, "Spectra", SPECTRA_ART_PATH),
            build_sidebar_stat_markup(bioactivity_count, "Bioactivity", BIOACTIVITY_ART_PATH),
        ]
    )

    st.markdown(
        f"""
        <div class="sidebar-brand">
            <div class="sidebar-brand-head">
                <div class="sidebar-brand-logo-shell">
                    <img class="sidebar-brand-logo" src="{logo_uri}" alt="NPDB logo" />
                </div>
                <div class="sidebar-brand-copy">
                    <div class="sidebar-brand-title">Natural Products<br>Spectral Database</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="sidebar-stats">
            {stats_markup}
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"""
        <div class="sidebar-meta-block">
            <div class="sidebar-meta-title">Last Update</div>
            <div class="sidebar-meta-row">
                <span>{latest_update}</span>
                <span class="sidebar-status-dot"></span>
            </div>
            <div class="sidebar-meta-divider"></div>
            <div class="sidebar-meta-title">Dataset Quality</div>
            <div class="sidebar-quality-value">{quality_pct}% Verified</div>
            <div class="sidebar-quality-track">
                <div class="sidebar-quality-fill" style="width:{quality_pct}%"></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_sidebar_navigation():
    current_section = st.session_state.get("nav_section", "Dashboard")
    current_compound_page = st.session_state.get("compound_page", "Browse Record")

    for group_title, items in SIDEBAR_NAV_GROUPS:
        if group_title == "Data Library":
            with st.expander("Additional Views", expanded=False):
                for item in items:
                    target_section = item["section"]
                    target_compound_page = item.get("compound_page")
                    if target_compound_page and not can_edit_database() and target_compound_page != "Browse Record":
                        continue
                    is_active = current_section == target_section
                    if target_compound_page:
                        is_active = is_active and current_compound_page == target_compound_page
                    render_sidebar_nav_link(group_title, item, is_active)
            continue

        st.markdown(f'<div class="sidebar-menu-caption">{group_title}</div>', unsafe_allow_html=True)
        for item in items:
            target_section = item["section"]
            target_compound_page = item.get("compound_page")
            if target_compound_page and not can_edit_database() and target_compound_page != "Browse Record":
                continue
            is_active = current_section == target_section
            if target_compound_page:
                is_active = is_active and current_compound_page == target_compound_page
            render_sidebar_nav_link(group_title, item, is_active)


def render_dashboard_showcase(
    filtered_df: pd.DataFrame,
    proton_count: int,
    carbon_count: int,
    spectra_count: int,
    bioactivity_count: int,
):
    hero_uri = image_to_data_uri(str(HERO_BANNER_PATH), max_px=1600) if HERO_BANNER_PATH.exists() else ""
    workspace_uri = image_to_data_uri(str(WORKSPACE_ART_PATH), max_px=540) if WORKSPACE_ART_PATH.exists() else ""
    search_icon_uri = image_to_data_uri(str(SEARCH_BIG_ART_PATH), max_px=160) if SEARCH_BIG_ART_PATH.exists() else ""
    hero_class = "dashboard-hero-card has-image" if hero_uri else "dashboard-hero-card"
    hero_style = f"background-image: linear-gradient(115deg, rgba(5, 11, 26, 0.9), rgba(8, 18, 34, 0.48)), url('{hero_uri}'); background-position: center right;" if hero_uri else ""
    render_raw_html(
        f"""
        <div class="{hero_class}" style="{hero_style}">
            <div class="dashboard-hero-overlay"></div>
            <div class="dashboard-hero-content">
                <div class="dashboard-hero-header">
                    <div class="dashboard-hero-text">
                        <div class="dashboard-hero-title-shell">
                            <h1 class="dashboard-hero-title">Natural Products<br><span class="accent">Spectral Database</span></h1>
                            <div class="dashboard-tagline">Explore. Analyze. Discover.</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
    )

    search_col, search_button_col = st.columns([3.25, 1.15], gap="medium")
    with search_col:
        st.markdown(
            f"""
            <div class="dashboard-search-strip">
                <div class="dashboard-search-strip-icon-shell">
                    <div class="dashboard-search-strip-icon">
                        {"<img class='dashboard-search-strip-icon-image' src='" + search_icon_uri + "' alt='Search icon' />" if search_icon_uri else "⌕"}
                    </div>
                </div>
                <div class="dashboard-search-strip-copy-shell">
                    <div class="dashboard-search-strip-title">Search Spectra</div>
                    <div class="dashboard-search-strip-copy">
                        Search by name, shift, formula, structure, or source organism.
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with search_button_col:
        st.write("")
        st.write("")
        if st.button("Start Searching ->", key="dashboard_start_search", width="stretch", type="primary", icon=":material/search:"):
            navigate_internal("Search & Match")
            st.rerun()

    workflow_col, workspace_col = st.columns([2.05, 0.82], gap="medium")
    with workflow_col:
        user_can_edit = can_edit_database()
        workflow_title = "Compound Workflow" if user_can_edit else "Research Workflow"
        workflow_steps = (
            DASHBOARD_WORKFLOW_STEPS
            if user_can_edit
            else [
                ("Search Spectra", "Find compounds by name, shift, formula, structure, or source."),
                ("Browse Record", "Open curated compound dossiers and linked spectra."),
                ("Download Data", "Export important metadata, NMR tables, and bioactivity summaries."),
            ]
        )
        workflow_cards = []
        for idx, (title, copy) in enumerate(workflow_steps, start=1):
            is_primary = idx == 1
            icon_markup = build_workflow_card_icon_markup(title)
            workflow_cards.append(
                f"""
                <div class="workflow-card {'is-primary' if is_primary else ''}">
                    <div class="workflow-step">{idx}</div>
                    {icon_markup}
                    <div class="workflow-title">{title}</div>
                    <div class="workflow-copy">{copy}</div>
                </div>
                """
            )
        render_raw_html(
            f"""
            <div class="dashboard-workflow-shell">
                <div class="dashboard-workflow-title">{workflow_title}</div>
                <div class="dashboard-workflow-grid">
                    {''.join(workflow_cards)}
                </div>
            </div>
            """
        )

    with workspace_col:
        art_markup = f'<img class="dashboard-workspace-art" src="{workspace_uri}" alt="Compound workspace visual" />' if workspace_uri else ""
        workspace_copy = (
            ("A unified workspace to manage curated data.", "Browse, create, import, and maintain high-quality spectral records in one place.")
            if can_edit_database()
            else ("A read-only workspace for curated records.", "Browse compound dossiers, linked spectra, and downloadable scientific tables.")
        )
        workspace_button_label = "Open Workspace" if can_edit_database() else "Browse Records"
        st.markdown(
            f"""
            <div class="dashboard-workspace-card">
                <div>
                    <div class="dashboard-workspace-title">Compound Workspace</div>
                    <div class="dashboard-workspace-copy">
                        {workspace_copy[0]}
                    </div>
                    <div class="dashboard-workspace-copy">
                        {workspace_copy[1]}
                    </div>
                </div>
                {art_markup}
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(workspace_button_label, key="dashboard_workspace_cta", width="stretch", icon=":material/folder_open:"):
            navigate_internal("Compound Workspace", "Browse Record")
            st.rerun()


def show_section_banner(image_path: Path, caption: str | None = None):
    if not image_path.exists():
        return
    st.markdown('<div class="section-banner">', unsafe_allow_html=True)
    st.image(str(image_path), width="stretch")
    st.markdown('</div>', unsafe_allow_html=True)
    if caption:
        st.caption(caption)


def render_batch_import_workspace():
    section_header(
        "Batch Import",
        "Use these ready-to-fill CSV templates to add compounds, 1H peaks, 13C peaks, and spectra records without guessing column names.",
    )

    write_batch_import_templates()
    template_map = build_batch_import_template_map()

    tabs = st.tabs(["Compounds", "1H Peaks", "13C Peaks", "Spectra Files"])
    import_specs = [
        (
            "compounds_batch_import_template.csv",
            COMPOUND_IMPORT_COLUMNS,
            ["trivial_name"],
            "Create new compound records in one pass. This is the best place to add metadata-heavy submissions from papers or lab notebooks.",
            import_compounds_from_dataframe,
        ),
        (
            "proton_nmr_batch_import_template.csv",
            PROTON_IMPORT_COLUMNS,
            ["delta_ppm", "assignment"],
            "Add many 1H NMR peaks at once. Use either compound_id or an exact compound_name already present in the database.",
            import_proton_from_dataframe,
        ),
        (
            "carbon_nmr_batch_import_template.csv",
            CARBON_IMPORT_COLUMNS,
            ["delta_ppm", "assignment"],
            "Add many 13C NMR peaks at once. Use either compound_id or an exact compound_name already present in the database.",
            import_carbon_from_dataframe,
        ),
        (
            "spectra_files_batch_import_template.csv",
            SPECTRA_IMPORT_COLUMNS,
            ["spectrum_type", "file_path"],
            "Register many spectra file links quickly. The file_path can point to an existing relative path inside data/spectra or to an external URL such as a Google Drive sharing link.",
            import_spectra_from_dataframe,
        ),
    ]

    for tab, (filename, expected_columns, required_columns, helper_text, import_function) in zip(tabs, import_specs):
        with tab:
            template_df = template_map[filename]
            template_path = TEMPLATES_DIR / filename

            render_helper_card("Template guide", helper_text)
            st.download_button(
                label=f"Download {filename}",
                data=dataframe_to_csv_bytes(template_df),
                file_name=filename,
                mime="text/csv",
                key=f"download_{filename}",
            )
            st.caption(f"Local template file: {template_path}")
            st.dataframe(template_df, width="stretch", hide_index=True)

            uploaded_file = st.file_uploader(
                f"Upload completed {filename}",
                type=["csv"],
                key=f"upload_{filename}",
            )

            if uploaded_file is None:
                continue

            try:
                uploaded_df = pd.read_csv(uploaded_file).fillna("")
            except Exception as exc:
                st.error(f"Could not read the CSV file: {exc}")
                continue

            missing_required_columns = validate_import_columns(uploaded_df, required_columns)
            if missing_required_columns:
                st.error(
                    "Missing required column(s): "
                    + ", ".join(missing_required_columns)
                    + ". Keep the original template headers unchanged."
                )
                continue

            preview_df = align_import_columns(uploaded_df, expected_columns)
            st.markdown("**Preview before import**")
            st.dataframe(preview_df, width="stretch", hide_index=True)

            if st.button(f"Import {filename}", key=f"import_{filename}", width="stretch"):
                inserted, skipped, errors = import_function(uploaded_df)
                status, headline = summarize_import_result(inserted, skipped, errors)
                getattr(st, status)(headline)

                if errors:
                    note_df = pd.DataFrame({"Import notes": errors[:30]})
                    st.dataframe(note_df, width="stretch", hide_index=True)

def render_kv(title, value):
    st.markdown(
        f"""
        <div class="kv-card">
            <div class="kv-title">{html_text(title)}</div>
            <div class="kv-value">{html_text(value)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

def render_compound_card(row, show_preview: bool = True):
    title = html_text(row["trivial_name"])
    formula = html_text(row["molecular_formula"])
    compound_class = html_text(row["compound_class"])
    subclass = html_text(row["compound_subclass"])
    source_summary = html_text(source_summary_from_record(row))
    sample_code = html_text(row["sample_code"])
    curation_status = html_text(clean_text(normalize_curation_status(row.get("curation_status"))).title())
    st.markdown('<div class="compound-card">', unsafe_allow_html=True)
    if show_preview:
        preview_col, info_col = st.columns([0.92, 4.08])
    else:
        preview_col = None
        info_col = st.container()
    if show_preview and preview_col is not None:
        with preview_col:
            source_value = row.get("structure_image_path")
            standardized_image = load_standardized_structure_source(source_value, size=(300, 220))
            if standardized_image is not None:
                st.image(standardized_image, width="stretch")
            elif source_value and is_external_url(str(source_value).strip()):
                safe_url = display_asset_url(source_value).replace('"', "&quot;")
                st.image(safe_url, width="stretch")
            else:
                structure_png = structure_smiles_png_bytes(row.get("smiles"), size=(300, 220))
                if structure_png:
                    st.image(structure_png, width="stretch")
                else:
                    structure_url = structure_smiles_image_url(row.get("smiles"))
                    if structure_url:
                        st.image(structure_url, width="stretch")
    with info_col:
        st.markdown(
            f"""
            <div class="result-title">{title}</div>
            <div class="result-subtitle">{formula}</div>
            <div class="info-chip-row">
                <span class="info-chip">Class: {compound_class}</span>
                <span class="info-chip">Subclass: {subclass}</span>
                <span class="info-chip">Source: {source_summary}</span>
                <span class="info-chip">Sample: {sample_code}</span>
                <span class="info-chip">Status: {curation_status}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.markdown('</div>', unsafe_allow_html=True)

# =========================
# Data loading
# =========================
@st.cache_data(show_spinner=False)
def load_all_compounds():
    conn = get_connection()
    query = """
        SELECT id, trivial_name, iupac_name, molecular_formula,
               smiles, inchi, inchikey,
               compound_class, compound_subclass,
               source_category, source_organism, source_material,
               sample_code, collection_location,
               gps_coordinates, depth_m, uv_data, ftir_data, cd_data,
               optical_rotation, melting_point, crystallization_method,
               structure_image_path, journal_name, article_title, publication_year,
               volume, issue, pages, doi, ccdc_number,
               molecular_weight, hrms_data, data_source, note,
               created_at, updated_at
        FROM compounds
        ORDER BY id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return enrich_compounds_dataframe(df)

@st.cache_data(show_spinner=False)
def load_compound_row(compound_id):
    conn = get_connection()
    query = """
        SELECT id, trivial_name, iupac_name, molecular_formula,
               smiles, inchi, inchikey,
               compound_class, compound_subclass,
               source_category, source_organism, source_material,
               sample_code, collection_location,
               gps_coordinates, depth_m, uv_data, ftir_data, cd_data,
               optical_rotation, melting_point, crystallization_method,
               structure_image_path, journal_name, article_title, publication_year,
               volume, issue, pages, doi, ccdc_number,
               molecular_weight, hrms_data, data_source, note,
               created_at, updated_at
        FROM compounds
        WHERE id = ?
    """
    df = pd.read_sql_query(query, conn, params=(compound_id,))
    conn.close()
    return enrich_compounds_dataframe(df)

@st.cache_data(show_spinner=False)
def load_proton_data(compound_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, delta_ppm, multiplicity, j_value, proton_count,
               assignment, solvent, instrument_mhz, note
        FROM proton_nmr
        WHERE compound_id = ?
        ORDER BY delta_ppm DESC
    """
    df = pd.read_sql_query(query, conn, params=(compound_id,))
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_all_proton_data():
    conn = get_connection()
    query = """
        SELECT p.id, p.compound_id, c.trivial_name,
               p.delta_ppm, p.multiplicity, p.j_value, p.proton_count,
               p.assignment, p.solvent, p.instrument_mhz, p.note
        FROM proton_nmr p
        LEFT JOIN compounds c ON p.compound_id = c.id
        ORDER BY p.id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_proton_row(proton_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, delta_ppm, multiplicity, j_value, proton_count,
               assignment, solvent, instrument_mhz, note
        FROM proton_nmr
        WHERE id = ?
    """
    df = pd.read_sql_query(query, conn, params=(proton_id,))
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_carbon_data(compound_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, delta_ppm, carbon_type, assignment, solvent,
               instrument_mhz, note
        FROM carbon_nmr
        WHERE compound_id = ?
        ORDER BY delta_ppm DESC
    """
    df = pd.read_sql_query(query, conn, params=(compound_id,))
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_all_carbon_data():
    conn = get_connection()
    query = """
        SELECT n.id, n.compound_id, c.trivial_name,
               n.delta_ppm, n.carbon_type, n.assignment, n.solvent,
               n.instrument_mhz, n.note
        FROM carbon_nmr n
        LEFT JOIN compounds c ON n.compound_id = c.id
        ORDER BY n.id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_carbon_row(carbon_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, delta_ppm, carbon_type, assignment, solvent,
               instrument_mhz, note
        FROM carbon_nmr
        WHERE id = ?
    """
    df = pd.read_sql_query(query, conn, params=(carbon_id,))
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_spectra_files(compound_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, spectrum_type, file_path, note
        FROM spectra_files
        WHERE compound_id = ?
        ORDER BY id ASC
    """
    df = pd.read_sql_query(query, conn, params=(compound_id,))
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_all_spectra_files():
    conn = get_connection()
    query = """
        SELECT s.id, s.compound_id, c.trivial_name,
               s.spectrum_type, s.file_path, s.note
        FROM spectra_files s
        LEFT JOIN compounds c ON s.compound_id = c.id
        ORDER BY s.id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

@st.cache_data(show_spinner=False)
def load_spectrum_file_row(file_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, spectrum_type, file_path, note
        FROM spectra_files
        WHERE id = ?
    """
    df = pd.read_sql_query(query, conn, params=(file_id,))
    conn.close()
    return df


@st.cache_data(show_spinner=False)
def load_bioactivity_data(compound_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, activity_label, target_name, target_category, assay_type,
               potency_type, potency_relation, potency_value, potency_unit, outcome,
               assay_medium, selectivity, assay_source, note
        FROM bioactivity_records
        WHERE compound_id = ?
        ORDER BY id ASC
    """
    df = pd.read_sql_query(query, conn, params=(compound_id,))
    conn.close()
    return df


@st.cache_data(show_spinner=False)
def load_all_bioactivity_data():
    conn = get_connection()
    query = """
        SELECT b.id, b.compound_id, c.trivial_name,
               b.activity_label, b.target_name, b.target_category, b.assay_type,
               b.potency_type, b.potency_relation, b.potency_value, b.potency_unit, b.outcome,
               b.assay_medium, b.selectivity, b.assay_source, b.note
        FROM bioactivity_records b
        LEFT JOIN compounds c ON b.compound_id = c.id
        ORDER BY b.id ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df


@st.cache_data(show_spinner=False)
def load_bioactivity_row(bioactivity_id):
    conn = get_connection()
    query = """
        SELECT id, compound_id, activity_label, target_name, target_category, assay_type,
               potency_type, potency_relation, potency_value, potency_unit, outcome,
               assay_medium, selectivity, assay_source, note
        FROM bioactivity_records
        WHERE id = ?
    """
    df = pd.read_sql_query(query, conn, params=(bioactivity_id,))
    conn.close()
    return df

# =========================
# Insert / update / delete functions
# =========================
def insert_compound_record(
    trivial_name,
    iupac_name,
    molecular_formula,
    compound_class,
    compound_subclass,
    smiles,
    inchi,
    inchikey,
    source_category,
    source_organism,
    source_material,
    sample_code,
    collection_location,
    gps_coordinates,
    depth_m,
    uv_data,
    ftir_data,
    cd_data,
    optical_rotation,
    melting_point,
    crystallization_method,
    structure_image_path,
    journal_name,
    article_title,
    publication_year,
    volume,
    issue,
    pages,
    doi,
    ccdc_number,
    molecular_weight,
    hrms_data,
    data_source,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO compounds (
            trivial_name,
            iupac_name,
            molecular_formula,
            compound_class,
            compound_subclass,
            smiles,
            inchi,
            inchikey,
            source_category,
            source_organism,
            source_material,
            sample_code,
            collection_location,
            gps_coordinates,
            depth_m,
            uv_data,
            ftir_data,
            cd_data,
            optical_rotation,
            melting_point,
            crystallization_method,
            structure_image_path,
            journal_name,
            article_title,
            publication_year,
            volume,
            issue,
            pages,
            doi,
            ccdc_number,
            molecular_weight,
            hrms_data,
            data_source,
            note
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        trivial_name,
        iupac_name,
        molecular_formula,
        compound_class,
        compound_subclass,
        smiles,
        inchi,
        inchikey,
        source_category,
        source_organism,
        source_material,
        sample_code,
        collection_location,
        gps_coordinates,
        depth_m,
        uv_data,
        ftir_data,
        cd_data,
        optical_rotation,
        melting_point,
        crystallization_method,
        structure_image_path,
        journal_name,
        article_title,
        publication_year,
        volume,
        issue,
        pages,
        doi,
        ccdc_number,
        molecular_weight,
        hrms_data,
        data_source,
        note
    ))

    new_id = cursor.lastrowid
    conn.commit()
    invalidate_cached_views()
    conn.close()
    return new_id

def update_compound_record(
    compound_id,
    trivial_name,
    iupac_name,
    molecular_formula,
    compound_class,
    compound_subclass,
    smiles,
    inchi,
    inchikey,
    source_category,
    source_organism,
    source_material,
    sample_code,
    collection_location,
    gps_coordinates,
    depth_m,
    uv_data,
    ftir_data,
    cd_data,
    optical_rotation,
    melting_point,
    crystallization_method,
    structure_image_path,
    journal_name,
    article_title,
    publication_year,
    volume,
    issue,
    pages,
    doi,
    ccdc_number,
    molecular_weight,
    hrms_data,
    data_source,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE compounds
        SET trivial_name = ?,
            iupac_name = ?,
            molecular_formula = ?,
            compound_class = ?,
            compound_subclass = ?,
            smiles = ?,
            inchi = ?,
            inchikey = ?,
            source_category = ?,
            source_organism = ?,
            source_material = ?,
            sample_code = ?,
            collection_location = ?,
            gps_coordinates = ?,
            depth_m = ?,
            uv_data = ?,
            ftir_data = ?,
            cd_data = ?,
            optical_rotation = ?,
            melting_point = ?,
            crystallization_method = ?,
            structure_image_path = ?,
            journal_name = ?,
            article_title = ?,
            publication_year = ?,
            volume = ?,
            issue = ?,
            pages = ?,
            doi = ?,
            ccdc_number = ?,
            molecular_weight = ?,
            hrms_data = ?,
            data_source = ?,
            note = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (
        trivial_name,
        iupac_name,
        molecular_formula,
        compound_class,
        compound_subclass,
        smiles,
        inchi,
        inchikey,
        source_category,
        source_organism,
        source_material,
        sample_code,
        collection_location,
        gps_coordinates,
        depth_m,
        uv_data,
        ftir_data,
        cd_data,
        optical_rotation,
        melting_point,
        crystallization_method,
        structure_image_path,
        journal_name,
        article_title,
        publication_year,
        volume,
        issue,
        pages,
        doi,
        ccdc_number,
        molecular_weight,
        hrms_data,
        data_source,
        note,
        compound_id
    ))

    conn.commit()

    invalidate_cached_views()
    conn.close()

def delete_compound_record(compound_id):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM bioactivity_records WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM proton_nmr WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM carbon_nmr WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM spectra_files WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM compounds WHERE id = ?", (compound_id,))
        conn.commit()
        invalidate_cached_views()
    finally:
        conn.close()

def insert_proton_record(
    compound_id,
    delta_ppm,
    multiplicity,
    j_value,
    proton_count,
    assignment,
    solvent,
    instrument_mhz,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO proton_nmr (
            compound_id,
            delta_ppm,
            multiplicity,
            j_value,
            proton_count,
            assignment,
            solvent,
            instrument_mhz,
            note
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        compound_id,
        delta_ppm,
        multiplicity,
        j_value,
        proton_count,
        assignment,
        solvent,
        instrument_mhz,
        note
    ))

    new_id = cursor.lastrowid
    conn.commit()
    invalidate_cached_views()
    conn.close()
    return new_id

def update_proton_record(
    proton_id,
    compound_id,
    delta_ppm,
    multiplicity,
    j_value,
    proton_count,
    assignment,
    solvent,
    instrument_mhz,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE proton_nmr
        SET compound_id = ?,
            delta_ppm = ?,
            multiplicity = ?,
            j_value = ?,
            proton_count = ?,
            assignment = ?,
            solvent = ?,
            instrument_mhz = ?,
            note = ?
        WHERE id = ?
    """, (
        compound_id,
        delta_ppm,
        multiplicity,
        j_value,
        proton_count,
        assignment,
        solvent,
        instrument_mhz,
        note,
        proton_id
    ))

    conn.commit()

    invalidate_cached_views()
    conn.close()

def delete_proton_record_by_id(proton_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM proton_nmr WHERE id = ?", (proton_id,))
    conn.commit()
    invalidate_cached_views()
    conn.close()

def insert_carbon_record(
    compound_id,
    delta_ppm,
    carbon_type,
    assignment,
    solvent,
    instrument_mhz,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO carbon_nmr (
            compound_id,
            delta_ppm,
            carbon_type,
            assignment,
            solvent,
            instrument_mhz,
            note
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        compound_id,
        delta_ppm,
        carbon_type,
        assignment,
        solvent,
        instrument_mhz,
        note
    ))

    new_id = cursor.lastrowid
    conn.commit()
    invalidate_cached_views()
    conn.close()
    return new_id

def update_carbon_record(
    carbon_id,
    compound_id,
    delta_ppm,
    carbon_type,
    assignment,
    solvent,
    instrument_mhz,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE carbon_nmr
        SET compound_id = ?,
            delta_ppm = ?,
            carbon_type = ?,
            assignment = ?,
            solvent = ?,
            instrument_mhz = ?,
            note = ?
        WHERE id = ?
    """, (
        compound_id,
        delta_ppm,
        carbon_type,
        assignment,
        solvent,
        instrument_mhz,
        note,
        carbon_id
    ))

    conn.commit()

    invalidate_cached_views()
    conn.close()

def delete_carbon_record_by_id(carbon_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM carbon_nmr WHERE id = ?", (carbon_id,))
    conn.commit()
    invalidate_cached_views()
    conn.close()

def insert_spectrum_file_record(
    compound_id,
    spectrum_type,
    file_path,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO spectra_files (
            compound_id,
            spectrum_type,
            file_path,
            note
        ) VALUES (?, ?, ?, ?)
    """, (
        compound_id,
        spectrum_type,
        file_path,
        note
    ))

    new_id = cursor.lastrowid
    conn.commit()
    invalidate_cached_views()
    conn.close()
    return new_id

def update_spectrum_file_record(
    file_id,
    compound_id,
    spectrum_type,
    file_path,
    note
):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE spectra_files
        SET compound_id = ?,
            spectrum_type = ?,
            file_path = ?,
            note = ?
        WHERE id = ?
    """, (
        compound_id,
        spectrum_type,
        file_path,
        note,
        file_id
    ))

    conn.commit()

    invalidate_cached_views()
    conn.close()

def delete_spectrum_file_record_by_id(file_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM spectra_files WHERE id = ?", (file_id,))
    conn.commit()
    invalidate_cached_views()
    conn.close()


def insert_bioactivity_record(
    compound_id,
    activity_label,
    target_name,
    target_category,
    assay_type,
    potency_type,
    potency_relation,
    potency_value,
    potency_unit,
    outcome,
    assay_medium,
    selectivity,
    assay_source,
    note,
):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO bioactivity_records (
            compound_id, activity_label, target_name, target_category, assay_type,
            potency_type, potency_relation, potency_value, potency_unit, outcome,
            assay_medium, selectivity, assay_source, note
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            compound_id,
            activity_label,
            target_name,
            target_category,
            assay_type,
            potency_type,
            potency_relation,
            potency_value,
            potency_unit,
            outcome,
            assay_medium,
            selectivity,
            assay_source,
            note,
        ),
    )
    new_id = cursor.lastrowid
    conn.commit()
    invalidate_cached_views()
    conn.close()
    return new_id


def update_bioactivity_record(
    bioactivity_id,
    compound_id,
    activity_label,
    target_name,
    target_category,
    assay_type,
    potency_type,
    potency_relation,
    potency_value,
    potency_unit,
    outcome,
    assay_medium,
    selectivity,
    assay_source,
    note,
):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE bioactivity_records
        SET compound_id = ?,
            activity_label = ?,
            target_name = ?,
            target_category = ?,
            assay_type = ?,
            potency_type = ?,
            potency_relation = ?,
            potency_value = ?,
            potency_unit = ?,
            outcome = ?,
            assay_medium = ?,
            selectivity = ?,
            assay_source = ?,
            note = ?
        WHERE id = ?
        """,
        (
            compound_id,
            activity_label,
            target_name,
            target_category,
            assay_type,
            potency_type,
            potency_relation,
            potency_value,
            potency_unit,
            outcome,
            assay_medium,
            selectivity,
            assay_source,
            note,
            bioactivity_id,
        ),
    )
    conn.commit()
    invalidate_cached_views()
    conn.close()


def delete_bioactivity_record_by_id(bioactivity_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM bioactivity_records WHERE id = ?", (bioactivity_id,))
    conn.commit()
    invalidate_cached_views()
    conn.close()


def is_template_marker(value) -> bool:
    return maybe_blank(value).upper().startswith("TEMPLATE_")


def normalize_import_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]
    return normalized


def align_import_columns(df: pd.DataFrame, expected_columns: list[str]) -> pd.DataFrame:
    aligned = normalize_import_dataframe(df)
    for column in expected_columns:
        if column not in aligned.columns:
            aligned[column] = ""
    return aligned[expected_columns]


def validate_import_columns(df: pd.DataFrame, required_columns: list[str]) -> list[str]:
    normalized_columns = {str(column).strip() for column in df.columns}
    return [column for column in required_columns if column not in normalized_columns]


def resolve_import_compound_id(row, compounds_df: pd.DataFrame) -> int:
    compound_id_text = maybe_blank(row.get("compound_id"))
    if compound_id_text:
        try:
            compound_id = int(float(compound_id_text))
        except ValueError as exc:
            raise ValueError(f"Invalid compound_id: {compound_id_text}") from exc

        matches = compounds_df[compounds_df["id"] == compound_id]
        if matches.empty:
            raise ValueError(f"compound_id {compound_id} was not found in the current database.")
        return compound_id

    compound_name = maybe_blank(row.get("compound_name")) or maybe_blank(row.get("trivial_name"))
    if not compound_name:
        raise ValueError("Provide either compound_id or compound_name.")
    if is_template_marker(compound_name):
        raise LookupError("Template row skipped.")

    matches = compounds_df[
        compounds_df["trivial_name"].fillna("").astype(str).str.casefold() == compound_name.casefold()
    ]
    if matches.empty:
        raise ValueError(f'Compound "{compound_name}" was not found. Use an exact trivial name or compound_id.')
    if len(matches) > 1:
        raise ValueError(f'Multiple compounds matched "{compound_name}". Use compound_id instead.')
    return int(matches.iloc[0]["id"])


def summarize_import_result(inserted: int, skipped: int, errors: list[str]) -> tuple[str, str]:
    if errors:
        status = "warning"
        headline = f"Imported {inserted} row(s), skipped {skipped}, and found {len(errors)} issue(s)."
    elif inserted:
        status = "success"
        headline = f"Imported {inserted} row(s) successfully."
    else:
        status = "info"
        headline = f"No rows were imported. Skipped {skipped} row(s)."
    return status, headline


def build_batch_import_template_map() -> dict[str, pd.DataFrame]:
    compounds_df = load_all_compounds()
    proton_df = load_all_proton_data()
    carbon_df = load_all_carbon_data()
    spectra_df = load_all_spectra_files()

    compound_example = compounds_df.iloc[0].to_dict() if not compounds_df.empty else {}
    proton_example = proton_df.iloc[0].to_dict() if not proton_df.empty else {}
    carbon_example = carbon_df.iloc[0].to_dict() if not carbon_df.empty else {}
    spectra_example = spectra_df.iloc[0].to_dict() if not spectra_df.empty else {}

    compound_row = {column: "" for column in COMPOUND_IMPORT_COLUMNS}
    for column in COMPOUND_IMPORT_COLUMNS:
        compound_row[column] = maybe_blank(compound_example.get(column))
    compound_row["trivial_name"] = "TEMPLATE_Replace_With_Compound_Name"
    compound_row["source_category"] = compound_row["source_category"] or "Sponge"
    compound_row["source_organism"] = compound_row["source_organism"] or "Stylissa sp."
    compound_row["sample_code"] = compound_row["sample_code"] or "NP-001"
    compound_row["data_source"] = compound_row["data_source"] or "Experimental"
    compound_row["curation_status"] = compound_row["curation_status"] or "imported"
    compound_row["note"] = "Delete or replace this template row before import."

    proton_row = {column: "" for column in PROTON_IMPORT_COLUMNS}
    for column in PROTON_IMPORT_COLUMNS:
        proton_row[column] = maybe_blank(proton_example.get(column))
    proton_row["compound_id"] = ""
    proton_row["compound_name"] = "TEMPLATE_Existing_Compound_Name"
    proton_row["note"] = "Delete or replace this template row before import."

    carbon_row = {column: "" for column in CARBON_IMPORT_COLUMNS}
    for column in CARBON_IMPORT_COLUMNS:
        carbon_row[column] = maybe_blank(carbon_example.get(column))
    carbon_row["compound_id"] = ""
    carbon_row["compound_name"] = "TEMPLATE_Existing_Compound_Name"
    carbon_row["note"] = "Delete or replace this template row before import."

    spectra_row = {column: "" for column in SPECTRA_IMPORT_COLUMNS}
    for column in SPECTRA_IMPORT_COLUMNS:
        spectra_row[column] = maybe_blank(spectra_example.get(column))
    spectra_row["compound_id"] = ""
    spectra_row["compound_name"] = "TEMPLATE_Existing_Compound_Name"
    spectra_row["file_path"] = spectra_row["file_path"] or "data/spectra/example_1H.png"
    spectra_row["spectrum_type"] = spectra_row["spectrum_type"] or "1H"
    spectra_row["note"] = "Delete or replace this template row before import."

    return {
        "compounds_batch_import_template.csv": align_import_columns(pd.DataFrame([compound_row]), COMPOUND_IMPORT_COLUMNS),
        "proton_nmr_batch_import_template.csv": align_import_columns(pd.DataFrame([proton_row]), PROTON_IMPORT_COLUMNS),
        "carbon_nmr_batch_import_template.csv": align_import_columns(pd.DataFrame([carbon_row]), CARBON_IMPORT_COLUMNS),
        "spectra_files_batch_import_template.csv": align_import_columns(pd.DataFrame([spectra_row]), SPECTRA_IMPORT_COLUMNS),
    }


def write_batch_import_templates():
    for filename, template_df in build_batch_import_template_map().items():
        template_path = TEMPLATES_DIR / filename
        template_csv = template_df.to_csv(index=False)
        if template_path.exists():
            try:
                if template_path.read_text(encoding="utf-8") == template_csv:
                    continue
            except UnicodeDecodeError:
                pass
        template_path.write_text(template_csv, encoding="utf-8")


def import_compounds_from_dataframe(df: pd.DataFrame):
    aligned = align_import_columns(df, COMPOUND_IMPORT_COLUMNS)
    existing_df = load_all_compounds()
    existing_keys = {
        (
            maybe_blank(row["trivial_name"]).casefold(),
            maybe_blank(row["sample_code"]).casefold(),
            maybe_blank(row["doi"]).casefold(),
        )
        for _, row in existing_df.iterrows()
    }

    inserted = 0
    skipped = 0
    errors = []

    for row_number, row in aligned.iterrows():
        display_row = row_number + 2
        trivial_name = maybe_blank(row.get("trivial_name"))
        if not trivial_name:
            skipped += 1
            continue
        if is_template_marker(trivial_name):
            skipped += 1
            continue

        dedupe_key = (
            trivial_name.casefold(),
            maybe_blank(row.get("sample_code")).casefold(),
            maybe_blank(row.get("doi")).casefold(),
        )
        if dedupe_key in existing_keys:
            skipped += 1
            errors.append(f"Row {display_row}: skipped because the compound already exists with the same name/sample/DOI.")
            continue

        depth_value = safe_float_or_none(row.get("depth_m"))
        if maybe_blank(row.get("depth_m")) and depth_value is None:
            errors.append(f"Row {display_row}: depth_m must be a valid number.")
            continue

        molecular_weight_value = safe_float_or_none(row.get("molecular_weight"))
        if maybe_blank(row.get("molecular_weight")) and molecular_weight_value is None:
            errors.append(f"Row {display_row}: molecular_weight must be a valid number.")
            continue

        source_category, source_organism, source_material = infer_source_fields(
            row.get("source_category"),
            row.get("source_organism"),
            row.get("source_material"),
        )

        insert_compound_record(
            trivial_name=trivial_name,
            iupac_name=maybe_blank(row.get("iupac_name")),
            molecular_formula=maybe_blank(row.get("molecular_formula")),
            compound_class=maybe_blank(row.get("compound_class")),
            compound_subclass=maybe_blank(row.get("compound_subclass")),
            smiles=maybe_blank(row.get("smiles")),
            inchi=maybe_blank(row.get("inchi")),
            inchikey=maybe_blank(row.get("inchikey")),
            source_category=source_category,
            source_organism=source_organism,
            source_material=source_material,
            sample_code=maybe_blank(row.get("sample_code")),
            collection_location=maybe_blank(row.get("collection_location")),
            gps_coordinates=maybe_blank(row.get("gps_coordinates")),
            depth_m=depth_value,
            uv_data=maybe_blank(row.get("uv_data")),
            ftir_data=maybe_blank(row.get("ftir_data")),
            cd_data=maybe_blank(row.get("cd_data")),
            optical_rotation=maybe_blank(row.get("optical_rotation")),
            melting_point=maybe_blank(row.get("melting_point")),
            crystallization_method=maybe_blank(row.get("crystallization_method")),
            structure_image_path=maybe_blank(row.get("structure_image_path")),
            journal_name=maybe_blank(row.get("journal_name")),
            article_title=maybe_blank(row.get("article_title")),
            publication_year=maybe_blank(row.get("publication_year")),
            volume=maybe_blank(row.get("volume")),
            issue=maybe_blank(row.get("issue")),
            pages=maybe_blank(row.get("pages")),
            doi=maybe_blank(row.get("doi")),
            ccdc_number=maybe_blank(row.get("ccdc_number")),
            molecular_weight=molecular_weight_value,
            hrms_data=maybe_blank(row.get("hrms_data")),
            data_source=maybe_blank(row.get("data_source")),
            curation_status=normalize_curation_status(maybe_blank(row.get("curation_status")), default="imported"),
            note=maybe_blank(row.get("note")),
        )
        existing_keys.add(dedupe_key)
        inserted += 1

    return inserted, skipped, errors


def import_proton_from_dataframe(df: pd.DataFrame):
    aligned = align_import_columns(df, PROTON_IMPORT_COLUMNS)
    compounds_df = load_all_compounds()
    existing_df = load_all_proton_data()
    existing_keys = {
        (
            int(row["compound_id"]),
            round(float(row["delta_ppm"]), 4),
            maybe_blank(row["assignment"]).casefold(),
        )
        for _, row in existing_df.iterrows()
    }

    inserted = 0
    skipped = 0
    errors = []

    for row_number, row in aligned.iterrows():
        display_row = row_number + 2
        if is_template_marker(row.get("compound_name")):
            skipped += 1
            continue

        try:
            compound_id = resolve_import_compound_id(row, compounds_df)
        except LookupError:
            skipped += 1
            continue
        except ValueError as exc:
            errors.append(f"Row {display_row}: {exc}")
            continue

        assignment = maybe_blank(row.get("assignment"))
        delta_text = maybe_blank(row.get("delta_ppm"))
        if not delta_text or not assignment:
            errors.append(f"Row {display_row}: delta_ppm and assignment are required.")
            continue

        delta_value = safe_float_or_none(delta_text)
        if delta_value is None:
            errors.append(f"Row {display_row}: delta_ppm must be a valid number.")
            continue

        instrument_value = safe_float_or_none(row.get("instrument_mhz"))
        if maybe_blank(row.get("instrument_mhz")) and instrument_value is None:
            errors.append(f"Row {display_row}: instrument_mhz must be a valid number.")
            continue

        dedupe_key = (compound_id, round(delta_value, 4), assignment.casefold())
        if dedupe_key in existing_keys:
            skipped += 1
            errors.append(f"Row {display_row}: skipped duplicate 1H peak for the same compound, shift, and assignment.")
            continue

        insert_proton_record(
            compound_id=compound_id,
            delta_ppm=delta_value,
            multiplicity=maybe_blank(row.get("multiplicity")),
            j_value=maybe_blank(row.get("j_value")),
            proton_count=maybe_blank(row.get("proton_count")),
            assignment=assignment,
            solvent=maybe_blank(row.get("solvent")),
            instrument_mhz=instrument_value,
            note=maybe_blank(row.get("note")),
        )
        existing_keys.add(dedupe_key)
        inserted += 1

    return inserted, skipped, errors


def import_carbon_from_dataframe(df: pd.DataFrame):
    aligned = align_import_columns(df, CARBON_IMPORT_COLUMNS)
    compounds_df = load_all_compounds()
    existing_df = load_all_carbon_data()
    existing_keys = {
        (
            int(row["compound_id"]),
            round(float(row["delta_ppm"]), 4),
            maybe_blank(row["assignment"]).casefold(),
        )
        for _, row in existing_df.iterrows()
    }

    inserted = 0
    skipped = 0
    errors = []

    for row_number, row in aligned.iterrows():
        display_row = row_number + 2
        if is_template_marker(row.get("compound_name")):
            skipped += 1
            continue

        try:
            compound_id = resolve_import_compound_id(row, compounds_df)
        except LookupError:
            skipped += 1
            continue
        except ValueError as exc:
            errors.append(f"Row {display_row}: {exc}")
            continue

        assignment = maybe_blank(row.get("assignment"))
        delta_text = maybe_blank(row.get("delta_ppm"))
        if not delta_text or not assignment:
            errors.append(f"Row {display_row}: delta_ppm and assignment are required.")
            continue

        delta_value = safe_float_or_none(delta_text)
        if delta_value is None:
            errors.append(f"Row {display_row}: delta_ppm must be a valid number.")
            continue

        instrument_value = safe_float_or_none(row.get("instrument_mhz"))
        if maybe_blank(row.get("instrument_mhz")) and instrument_value is None:
            errors.append(f"Row {display_row}: instrument_mhz must be a valid number.")
            continue

        dedupe_key = (compound_id, round(delta_value, 4), assignment.casefold())
        if dedupe_key in existing_keys:
            skipped += 1
            errors.append(f"Row {display_row}: skipped duplicate 13C peak for the same compound, shift, and assignment.")
            continue

        insert_carbon_record(
            compound_id=compound_id,
            delta_ppm=delta_value,
            carbon_type=maybe_blank(row.get("carbon_type")),
            assignment=assignment,
            solvent=maybe_blank(row.get("solvent")),
            instrument_mhz=instrument_value,
            note=maybe_blank(row.get("note")),
        )
        existing_keys.add(dedupe_key)
        inserted += 1

    return inserted, skipped, errors


def import_spectra_from_dataframe(df: pd.DataFrame):
    aligned = align_import_columns(df, SPECTRA_IMPORT_COLUMNS)
    compounds_df = load_all_compounds()
    existing_df = load_all_spectra_files()
    existing_keys = {
        (
            int(row["compound_id"]),
            maybe_blank(row["spectrum_type"]).casefold(),
            maybe_blank(row["file_path"]).casefold(),
        )
        for _, row in existing_df.iterrows()
    }

    inserted = 0
    skipped = 0
    errors = []

    for row_number, row in aligned.iterrows():
        display_row = row_number + 2
        if is_template_marker(row.get("compound_name")):
            skipped += 1
            continue

        try:
            compound_id = resolve_import_compound_id(row, compounds_df)
        except LookupError:
            skipped += 1
            continue
        except ValueError as exc:
            errors.append(f"Row {display_row}: {exc}")
            continue

        spectrum_type = maybe_blank(row.get("spectrum_type"))
        file_path = maybe_blank(row.get("file_path"))
        if not spectrum_type or not file_path:
            errors.append(f"Row {display_row}: spectrum_type and file_path are required.")
            continue

        validation_errors, validation_warnings = validate_spectrum_entry(file_path, spectrum_type)
        if validation_errors:
            errors.extend([f"Row {display_row}: {message}" for message in validation_errors])
            continue
        errors.extend([f"Row {display_row}: note - {message}" for message in validation_warnings])

        dedupe_key = (compound_id, spectrum_type.casefold(), file_path.casefold())
        if dedupe_key in existing_keys:
            skipped += 1
            errors.append(f"Row {display_row}: skipped duplicate spectra file entry.")
            continue

        insert_spectrum_file_record(
            compound_id=compound_id,
            spectrum_type=spectrum_type,
            file_path=file_path,
            note=maybe_blank(row.get("note")),
        )
        existing_keys.add(dedupe_key)
        inserted += 1

    return inserted, skipped, errors

# =========================
# Similarity search helpers
# =========================
def get_db_signature():
    if not DB_PATH.exists():
        return 0.0
    return DB_PATH.stat().st_mtime


@st.cache_data(show_spinner=False)
def load_search_index(_db_signature: float):
    conn = get_connection()
    try:
        compounds_df = pd.read_sql_query(
            """
            SELECT id, trivial_name, sample_code, molecular_formula, smiles, inchi, inchikey,
                   source_category, source_organism, source_material,
                   compound_class, compound_subclass, data_source
            FROM compounds
            ORDER BY id ASC
            """,
            conn,
        )
        proton_df = pd.read_sql_query(
            """
            SELECT compound_id, delta_ppm
            FROM proton_nmr
            ORDER BY compound_id ASC, delta_ppm DESC
            """,
            conn,
        )
        carbon_df = pd.read_sql_query(
            """
            SELECT compound_id, delta_ppm
            FROM carbon_nmr
            ORDER BY compound_id ASC, delta_ppm DESC
            """,
            conn,
        )
    finally:
        conn.close()

    compounds_df = enrich_compounds_dataframe(compounds_df)

    proton_groups = {}
    carbon_groups = {}

    if not proton_df.empty:
        proton_groups = proton_df.groupby("compound_id")["delta_ppm"].apply(list).to_dict()
    if not carbon_df.empty:
        carbon_groups = carbon_df.groupby("compound_id")["delta_ppm"].apply(list).to_dict()

    search_index = []
    for _, row in compounds_df.iterrows():
        compound_id = int(row["id"])
        search_index.append(
            {
                "compound_id": compound_id,
                "trivial_name": row["trivial_name"],
                "sample_code": row["sample_code"],
                "molecular_formula": row["molecular_formula"],
                "source_category": row.get("source_category"),
                "source_organism": row.get("source_organism"),
                "source_material": row["source_material"],
                "compound_class": row["compound_class"],
                "compound_subclass": row["compound_subclass"],
                "data_source": row["data_source"],
                "proton_peaks": proton_groups.get(compound_id, []),
                "carbon_peaks": carbon_groups.get(compound_id, []),
            }
        )

    return search_index


def score_peak_matches(query_peaks, db_peaks, tolerance):
    matches = find_best_matches(query_peaks, db_peaks, tolerance)
    match_count = len(matches)
    total_query = len(query_peaks)
    db_peak_count = len(db_peaks)
    query_coverage = (match_count / total_query) if total_query else 0.0
    db_coverage = (match_count / db_peak_count) if db_peak_count else 0.0
    avg_difference = None
    closeness = 0.0

    if matches:
        avg_difference = sum(diff for _, _, diff in matches) / match_count
        if tolerance > 0:
            closeness = max(0.0, 1 - (avg_difference / tolerance))
        elif avg_difference == 0:
            closeness = 1.0

    score = ((query_coverage * 0.65) + (db_coverage * 0.20) + (closeness * 0.15)) * 100

    return {
        "matches": matches,
        "match_count": match_count,
        "total_query": total_query,
        "db_peak_count": db_peak_count,
        "query_coverage": query_coverage,
        "db_coverage": db_coverage,
        "avg_difference": avg_difference,
        "score": score,
    }


def sort_similarity_results(results, score_key="score"):
    return sorted(
        results,
        key=lambda item: (
            item.get(score_key, 0.0),
            item.get("match_count", 0),
            item.get("query_coverage", 0.0),
            -(item.get("avg_difference") if item.get("avg_difference") is not None else 9999),
        ),
        reverse=True,
    )


def search_similarity_13c(query_peaks, tolerance):
    results = []
    search_index = load_search_index(get_db_signature())

    for item in search_index:
        metrics = score_peak_matches(query_peaks, item["carbon_peaks"], tolerance)
        results.append(
            {
                **item,
                "db_peak_count": metrics["db_peak_count"],
                "match_count": metrics["match_count"],
                "total_query": metrics["total_query"],
                "query_coverage": metrics["query_coverage"],
                "db_coverage": metrics["db_coverage"],
                "avg_difference": metrics["avg_difference"],
                "score": metrics["score"],
                "matches": metrics["matches"],
            }
        )

    return sort_similarity_results(results)


def search_similarity_1h(query_peaks, tolerance):
    results = []
    search_index = load_search_index(get_db_signature())

    for item in search_index:
        metrics = score_peak_matches(query_peaks, item["proton_peaks"], tolerance)
        results.append(
            {
                **item,
                "db_peak_count": metrics["db_peak_count"],
                "match_count": metrics["match_count"],
                "total_query": metrics["total_query"],
                "query_coverage": metrics["query_coverage"],
                "db_coverage": metrics["db_coverage"],
                "avg_difference": metrics["avg_difference"],
                "score": metrics["score"],
                "matches": metrics["matches"],
            }
        )

    return sort_similarity_results(results)


def search_similarity_combined(query_protons, proton_tol, query_carbons, carbon_tol):
    results = []
    search_index = load_search_index(get_db_signature())

    for item in search_index:
        proton_metrics = score_peak_matches(query_protons, item["proton_peaks"], proton_tol) if query_protons else None
        carbon_metrics = score_peak_matches(query_carbons, item["carbon_peaks"], carbon_tol) if query_carbons else None

        if proton_metrics and carbon_metrics:
            total_score = (proton_metrics["score"] * 0.5) + (carbon_metrics["score"] * 0.5)
        elif proton_metrics:
            total_score = proton_metrics["score"]
        elif carbon_metrics:
            total_score = carbon_metrics["score"]
        else:
            total_score = 0.0

        avg_differences = [
            metric["avg_difference"]
            for metric in [proton_metrics, carbon_metrics]
            if metric and metric["avg_difference"] is not None
        ]

        results.append(
            {
                **item,
                "db_proton_count": len(item["proton_peaks"]),
                "db_carbon_count": len(item["carbon_peaks"]),
                "proton_match_count": proton_metrics["match_count"] if proton_metrics else 0,
                "carbon_match_count": carbon_metrics["match_count"] if carbon_metrics else 0,
                "proton_total_query": proton_metrics["total_query"] if proton_metrics else len(query_protons),
                "carbon_total_query": carbon_metrics["total_query"] if carbon_metrics else len(query_carbons),
                "proton_score": proton_metrics["score"] if proton_metrics else 0.0,
                "carbon_score": carbon_metrics["score"] if carbon_metrics else 0.0,
                "proton_query_coverage": proton_metrics["query_coverage"] if proton_metrics else 0.0,
                "carbon_query_coverage": carbon_metrics["query_coverage"] if carbon_metrics else 0.0,
                "proton_db_coverage": proton_metrics["db_coverage"] if proton_metrics else 0.0,
                "carbon_db_coverage": carbon_metrics["db_coverage"] if carbon_metrics else 0.0,
                "total_score": total_score,
                "avg_difference": (
                    sum(avg_differences) / len(avg_differences) if avg_differences else None
                ),
                "proton_matches": proton_metrics["matches"] if proton_metrics else [],
                "carbon_matches": carbon_metrics["matches"] if carbon_metrics else [],
            }
        )

    return sorted(
        results,
        key=lambda item: (
            item["total_score"],
            item["proton_match_count"] + item["carbon_match_count"],
            -(item["avg_difference"] if item["avg_difference"] is not None else 9999),
        ),
        reverse=True,
    )

# =========================
# Export helpers
# =========================
def export_name_results(result_df: pd.DataFrame) -> pd.DataFrame:
    export_df = result_df.copy()
    if "source_material" in export_df.columns:
        export_df["source_material"] = export_df.apply(source_summary_from_record, axis=1)
    return export_df.rename(columns={
        "id": "ID",
        "trivial_name": "Trivial Name",
        "iupac_name": "IUPAC Name",
        "molecular_formula": "Molecular Formula",
        "compound_class": "Compound Class",
        "compound_subclass": "Compound Subclass",
        "source_category": "Source Category",
        "source_organism": "Source Organism",
        "source_material": "Source Summary",
        "sample_code": "Sample Code",
        "collection_location": "Collection Location",
        "gps_coordinates": "GPS Coordinates",
        "depth_m": "Depth (m)",
        "uv_data": "UV Data",
        "ftir_data": "FTIR Data",
        "cd_data": "CD / ECD Data",
        "optical_rotation": "Optical Rotation",
        "melting_point": "Melting Point",
        "crystallization_method": "Crystallization Method",
        "structure_image_path": "Structure Image Path",
        "journal_name": "Journal Name",
        "article_title": "Article Title",
        "publication_year": "Publication Year",
        "volume": "Volume",
        "issue": "Issue",
        "pages": "Pages",
        "doi": "DOI",
        "ccdc_number": "CCDC",
        "molecular_weight": "Mr",
        "hrms_data": "HRMS Data",
        "data_source": "Data Source",
        "curation_status": "Curation Status",
        "note": "Note"
    })

def export_similarity_results_13c(results: list) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(results[:10], start=1):
        rows.append({
            "Rank": i,
            "Compound ID": item["compound_id"],
            "Trivial Name": clean_text(item["trivial_name"]),
            "Molecular Formula": clean_text(item["molecular_formula"]),
            "Compound Class": clean_text(item["compound_class"]),
            "Compound Subclass": clean_text(item["compound_subclass"]),
            "Source Category": clean_text(item.get("source_category")),
            "Source Organism": clean_text(item.get("source_organism")),
            "Source Summary": clean_text(source_summary_from_record(item)),
            "Sample Code": clean_text(item["sample_code"]),
            "Matched Peaks": item["match_count"],
            "Query Peaks": item["total_query"],
            "DB Peaks": item["db_peak_count"],
            "Query Coverage (%)": round(item["query_coverage"] * 100, 2),
            "DB Coverage (%)": round(item["db_coverage"] * 100, 2),
            "Average Difference": round(item["avg_difference"], 4) if item["avg_difference"] is not None else "-",
            "Score (%)": round(item["score"], 2),
        })
    return pd.DataFrame(rows)

def export_similarity_results_1h(results: list) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(results[:10], start=1):
        rows.append({
            "Rank": i,
            "Compound ID": item["compound_id"],
            "Trivial Name": clean_text(item["trivial_name"]),
            "Molecular Formula": clean_text(item["molecular_formula"]),
            "Compound Class": clean_text(item["compound_class"]),
            "Compound Subclass": clean_text(item["compound_subclass"]),
            "Source Category": clean_text(item.get("source_category")),
            "Source Organism": clean_text(item.get("source_organism")),
            "Source Summary": clean_text(source_summary_from_record(item)),
            "Sample Code": clean_text(item["sample_code"]),
            "Matched Peaks": item["match_count"],
            "Query Peaks": item["total_query"],
            "DB Peaks": item["db_peak_count"],
            "Query Coverage (%)": round(item["query_coverage"] * 100, 2),
            "DB Coverage (%)": round(item["db_coverage"] * 100, 2),
            "Average Difference": round(item["avg_difference"], 4) if item["avg_difference"] is not None else "-",
            "Score (%)": round(item["score"], 2),
        })
    return pd.DataFrame(rows)

def export_similarity_results_combined(results: list) -> pd.DataFrame:
    rows = []
    for i, item in enumerate(results[:10], start=1):
        rows.append({
            "Rank": i,
            "Compound ID": item["compound_id"],
            "Trivial Name": clean_text(item["trivial_name"]),
            "Molecular Formula": clean_text(item["molecular_formula"]),
            "Compound Class": clean_text(item["compound_class"]),
            "Compound Subclass": clean_text(item["compound_subclass"]),
            "Source Category": clean_text(item.get("source_category")),
            "Source Organism": clean_text(item.get("source_organism")),
            "Source Summary": clean_text(source_summary_from_record(item)),
            "Sample Code": clean_text(item["sample_code"]),
            "1H Matched Peaks": item["proton_match_count"],
            "1H Query Peaks": item["proton_total_query"],
            "1H Query Coverage (%)": round(item["proton_query_coverage"] * 100, 2),
            "1H Score (%)": round(item["proton_score"], 2),
            "13C Matched Peaks": item["carbon_match_count"],
            "13C Query Peaks": item["carbon_total_query"],
            "13C Query Coverage (%)": round(item["carbon_query_coverage"] * 100, 2),
            "13C Score (%)": round(item["carbon_score"], 2),
            "Average Difference": round(item["avg_difference"], 4) if item["avg_difference"] is not None else "-",
            "Total Score (%)": round(item["total_score"], 2),
        })
    return pd.DataFrame(rows)

def build_compound_summary_text(compound_row, proton_df, carbon_df, spectra_df):
    row = compound_row.iloc[0]
    bioactivity_df = load_bioactivity_data(int(row["id"]))

    summary = f"""Natural Products Spectral Database
Compound Summary

Compound ID: {row['id']}
Trivial Name: {clean_text(row['trivial_name'])}
IUPAC Name: {clean_text(row['iupac_name'])}
Molecular Formula: {clean_text(row['molecular_formula'])}
SMILES: {clean_text(row.get('smiles'))}
InChI: {clean_text(row.get('inchi'))}
InChIKey: {clean_text(row.get('inchikey'))}
Compound Class: {clean_text(row['compound_class'])}
Compound Subclass: {clean_text(row['compound_subclass'])}

Source Category: {clean_text(row.get('source_category'))}
Source Organism: {clean_text(row.get('source_organism'))}
Source Summary: {clean_text(source_summary_from_record(row))}
Sample Code: {clean_text(row['sample_code'])}
Collection Location: {clean_text(row['collection_location'])}
GPS Coordinates: {clean_text(row['gps_coordinates'])}
Depth (m): {clean_text(row['depth_m'])}

UV Data: {clean_text(row['uv_data'])}
FTIR Data: {clean_text(row['ftir_data'])}
CD / ECD Data: {clean_text(row.get('cd_data'))}
Optical Rotation: {clean_text(row['optical_rotation'])}
Melting Point: {clean_text(row['melting_point'])}
Crystallization Method: {clean_text(row['crystallization_method'])}

Journal Name: {clean_text(row['journal_name'])}
Article Title: {clean_text(row['article_title'])}
Publication Year: {clean_text(row['publication_year'])}
Volume: {clean_text(row['volume'])}
Issue: {clean_text(row['issue'])}
Pages: {clean_text(row['pages'])}
DOI: {clean_text(row['doi'])}
CCDC: {clean_text(row['ccdc_number'])}
Mr: {clean_text(row['molecular_weight'])}
HRMS Data: {clean_text(row['hrms_data'])}
Data Source: {clean_text(row['data_source'])}
Curation Status: {clean_text(row.get('curation_status')).title()}

Note:
{clean_text(row['note'])}

Data Coverage
-------------
1H NMR Peaks: {len(proton_df)}
13C NMR Peaks: {len(carbon_df)}
Spectra Files: {len(spectra_df)}
Bioactivity Records: {len(bioactivity_df)}
"""
    return add_credit_to_text_bytes(summary.encode("utf-8"))


def _archive_safe_name(value: str, fallback: str = "file") -> str:
    return slugify_value(value, fallback=fallback).replace(" ", "_")


def _asset_bytes_from_source(source_value: str) -> tuple[bytes, str]:
    source_text = maybe_blank(source_value)
    if not source_text:
        return b"", ""
    try:
        if is_external_url(source_text):
            with urllib.request.urlopen(display_asset_url(source_text), timeout=30, context=_supabase_ssl_context()) as response:
                data = response.read()
                suffix = Path(urlparse(source_text).path).suffix or mimetypes.guess_extension(response.headers.get_content_type()) or ".bin"
                return data, suffix
        full_path = get_full_file_path(source_text)
        if full_path is not None and full_path.exists():
            return full_path.read_bytes(), full_path.suffix or ".bin"
    except Exception:
        return b"", ""
    return b"", ""


def _compound_structure_asset(row_data) -> tuple[bytes, str]:
    structure_bytes, suffix = _asset_bytes_from_source(row_data.get("structure_image_path"))
    if structure_bytes:
        return structure_bytes, suffix or ".png"
    smiles_value = maybe_blank(row_data.get("smiles"))
    if smiles_value:
        generated = structure_smiles_png_bytes(smiles_value, size=(900, 650))
        if generated:
            return generated, ".png"
        fallback_url = structure_smiles_image_url(smiles_value)
        return _asset_bytes_from_source(fallback_url)
    return b"", ""


def _dataframes_to_record_workbook(sheets: dict[str, pd.DataFrame]) -> bytes:
    if Alignment is None and Font is None and PatternFill is None:
        raise ModuleNotFoundError("openpyxl is not available")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for raw_name, df in sheets.items():
            sheet_name = raw_name[:31]
            export_df = df.copy()
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 48)
            footer_row = worksheet.max_row + 2
            footer_col_end = max(1, worksheet.max_column)
            worksheet.cell(row=footer_row, column=1, value=OWNER_CREDIT)
            if footer_col_end > 1:
                worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=footer_col_end)
            footer_cell = worksheet.cell(row=footer_row, column=1)
            if Font is not None:
                footer_cell.font = Font(italic=True, size=10, color="4F5B6B")
            if Alignment is not None:
                footer_cell.alignment = Alignment(horizontal="right")
            if PatternFill is not None:
                footer_cell.fill = PatternFill(fill_type="solid", fgColor="F5F8FD")
    output.seek(0)
    return output.getvalue()


def _select_existing_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=columns)
    return df[[column for column in columns if column in df.columns]].copy()


def _important_metadata_dataframe(row_data) -> pd.DataFrame:
    fields = [
        ("Compound ID", row_data.get("id")),
        ("Trivial Name", row_data.get("trivial_name")),
        ("IUPAC Name", row_data.get("iupac_name")),
        ("Molecular Formula", row_data.get("molecular_formula")),
        ("Molecular Weight", row_data.get("molecular_weight")),
        ("SMILES", row_data.get("smiles")),
        ("InChI", row_data.get("inchi")),
        ("InChIKey", row_data.get("inchikey")),
        ("Compound Class", row_data.get("compound_class")),
        ("Compound Subclass", row_data.get("compound_subclass")),
        ("Source Category", row_data.get("source_category")),
        ("Source Organism", row_data.get("source_organism")),
        ("Source Summary", source_summary_from_record(row_data)),
        ("Sample Code", row_data.get("sample_code")),
        ("Collection Location", row_data.get("collection_location")),
        ("GPS Coordinates", row_data.get("gps_coordinates")),
        ("Depth (m)", row_data.get("depth_m")),
        ("UV Data", row_data.get("uv_data")),
        ("FTIR Data", row_data.get("ftir_data")),
        ("CD / ECD Data", row_data.get("cd_data")),
        ("Optical Rotation", row_data.get("optical_rotation")),
        ("Melting Point", row_data.get("melting_point")),
        ("HRMS Data", row_data.get("hrms_data")),
        ("Journal Name", row_data.get("journal_name")),
        ("Article Title", row_data.get("article_title")),
        ("Publication Year", row_data.get("publication_year")),
        ("Volume", row_data.get("volume")),
        ("Issue", row_data.get("issue")),
        ("Pages", row_data.get("pages")),
        ("DOI", row_data.get("doi")),
        ("Data Source", row_data.get("data_source")),
        ("Curation Status", clean_text(normalize_curation_status(row_data.get("curation_status"))).title()),
        ("Structure Image Link", row_data.get("structure_image_path")),
        ("Note", row_data.get("note")),
    ]
    return pd.DataFrame(
        [{"Field": label, "Value": clean_text(value)} for label, value in fields]
    )


def _important_bioactivity_dataframe(bioactivity_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "id",
        "activity_label",
        "target_name",
        "target_category",
        "assay_type",
        "potency_type",
        "potency_relation",
        "potency_value",
        "potency_unit",
        "outcome",
        "assay_source",
        "note",
    ]
    export_df = _select_existing_columns(bioactivity_df, columns)
    return export_df.rename(
        columns={
            "id": "Bioactivity ID",
            "activity_label": "Activity",
            "target_name": "Target",
            "target_category": "Target Category",
            "assay_type": "Assay Type",
            "potency_type": "Metric",
            "potency_relation": "Relation",
            "potency_value": "Value",
            "potency_unit": "Unit",
            "outcome": "Outcome",
            "assay_source": "Assay Source",
            "note": "Note",
        }
    )


def build_compound_record_bundle(compound_row, proton_df, carbon_df, spectra_df, bioactivity_df) -> tuple[bytes, str]:
    row_data = compound_row.iloc[0]
    compound_id = int(row_data["id"])
    record_name = _archive_safe_name(clean_text(row_data["trivial_name"]), fallback=f"compound_{compound_id}")
    bundle_name = f"NPDB_compound_{compound_id}_{record_name}_full_record.zip"
    metadata_df = _important_metadata_dataframe(row_data)
    proton_export_df = _select_existing_columns(
        proton_df,
        ["delta_ppm", "multiplicity", "j_value", "proton_count", "assignment", "solvent", "instrument_mhz", "note"],
    )
    carbon_export_df = _select_existing_columns(
        carbon_df,
        ["delta_ppm", "carbon_type", "assignment", "solvent", "instrument_mhz", "note"],
    )
    spectra_export_df = _select_existing_columns(
        spectra_df,
        ["id", "spectrum_type", "file_path", "note"],
    ).rename(
        columns={
            "id": "Spectra File ID",
            "spectrum_type": "Spectrum Type",
            "file_path": "Cloud Link",
            "note": "Note",
        }
    )
    bioactivity_export_df = _important_bioactivity_dataframe(bioactivity_df)

    workbook_sheets = {
        "Metadata": metadata_df,
        "1H_NMR": proton_export_df,
        "13C_NMR": carbon_export_df,
        "Spectra_Files": spectra_export_df,
        "Bioactivity": bioactivity_export_df,
    }

    manifest = {
        "project": "NPDB: Natural Products Spectral Database",
        "compound_id": compound_id,
        "trivial_name": clean_text(row_data.get("trivial_name")),
        "created_at_utc": datetime.now(UTC).isoformat(),
        "included": {
            "metadata_rows": int(len(metadata_df)),
            "proton_nmr_rows": int(len(proton_df)),
            "carbon_nmr_rows": int(len(carbon_df)),
            "spectra_file_rows": int(len(spectra_df)),
            "bioactivity_rows": int(len(bioactivity_df)),
        },
        "spectra_policy": "Spectra image/raw files are listed as cloud links only and are not embedded in this download package.",
    }

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("README.txt", build_compound_summary_text(compound_row, proton_df, carbon_df, spectra_df))
        archive.writestr("manifest.json", json.dumps(manifest, indent=2))
        archive.writestr("tables/compound_metadata.csv", metadata_df.to_csv(index=False))
        archive.writestr("tables/1H_NMR_chemical_shifts.csv", proton_export_df.to_csv(index=False))
        archive.writestr("tables/13C_NMR_chemical_shifts.csv", carbon_export_df.to_csv(index=False))
        archive.writestr("tables/spectra_files.csv", spectra_export_df.to_csv(index=False))
        archive.writestr("tables/bioactivity_records.csv", bioactivity_export_df.to_csv(index=False))
        try:
            archive.writestr("NPDB_full_record.xlsx", _dataframes_to_record_workbook(workbook_sheets))
        except Exception:
            pass

        structure_bytes, structure_suffix = _compound_structure_asset(row_data)
        if structure_bytes:
            archive.writestr(f"structure/compound_{compound_id}_structure{structure_suffix or '.png'}", structure_bytes)
    output.seek(0)
    return output.getvalue(), bundle_name

# =========================
# Bioactivity helpers
# =========================
def export_bioactivity_results(bioactivity_df: pd.DataFrame) -> pd.DataFrame:
    if bioactivity_df.empty:
        return pd.DataFrame(
            columns=[
                "ID",
                "Compound ID",
                "Trivial Name",
                "Activity",
                "Target",
                "Target Category",
                "Assay Type",
                "Metric",
                "Relation",
                "Value",
                "Unit",
                "Outcome",
                "Assay Medium",
                "Selectivity",
                "Assay Source",
                "Note",
            ]
        )
    export_df = bioactivity_df.copy()
    return export_df.rename(
        columns={
            "id": "ID",
            "compound_id": "Compound ID",
            "trivial_name": "Trivial Name",
            "activity_label": "Activity",
            "target_name": "Target",
            "target_category": "Target Category",
            "assay_type": "Assay Type",
            "potency_type": "Metric",
            "potency_relation": "Relation",
            "potency_value": "Value",
            "potency_unit": "Unit",
            "outcome": "Outcome",
            "assay_medium": "Assay Medium",
            "selectivity": "Selectivity",
            "assay_source": "Assay Source",
            "note": "Note",
        }
    )


def render_bioactivity_table(compound_id: int):
    bioactivity_df = load_bioactivity_data(compound_id)
    section_header(
        "Bioactivity",
        "Reported assay outcomes, targets, potency values, and screening notes linked to this compound.",
    )
    if bioactivity_df.empty:
        st.info("No bioactivity records available for this compound yet.")
        return

    display_df = bioactivity_df.rename(
        columns={
            "id": "ID",
            "activity_label": "Activity",
            "target_name": "Target",
            "target_category": "Target Category",
            "assay_type": "Assay Type",
            "potency_type": "Metric",
            "potency_relation": "Relation",
            "potency_value": "Value",
            "potency_unit": "Unit",
            "outcome": "Outcome",
            "assay_medium": "Assay Medium",
            "selectivity": "Selectivity",
            "assay_source": "Assay Source",
            "note": "Note",
        }
    )
    st.dataframe(display_df, width="stretch", hide_index=True)

    export_df = export_bioactivity_results(load_all_bioactivity_data().query("compound_id == @compound_id"))
    download_dataframe_button(
        label="Download Bioactivity Table as Excel",
        df=export_df,
        file_name=f"compound_{compound_id}_bioactivity.xlsx",
        key=f"download_bioactivity_{compound_id}",
        sheet_name="Bioactivity",
    )

# =========================
# Spectra preview
# =========================
def render_spectra_section(compound_id):
    spectra_df = load_spectra_files(compound_id)

    section_header("Spectra Files", "Registered previews, PDFs, raw-data links, and downloadable files.")

    if spectra_df.empty:
        st.info("No spectra files available.")
        return

    grouped_types = spectra_df["spectrum_type"].fillna("Uncategorized").unique().tolist()

    for spectrum_type in grouped_types:
        sub_df = spectra_df[spectra_df["spectrum_type"].fillna("Uncategorized") == spectrum_type]

        with st.expander(f"{spectrum_type} ({len(sub_df)})", expanded=True):
            for _, row in sub_df.iterrows():
                file_id = row["id"]
                file_path_value = row["file_path"]
                note_value = clean_text(row["note"])
                full_path = get_full_file_path(file_path_value)
                _, file_warnings = validate_spectrum_entry(file_path_value, spectrum_type)

                st.markdown(
                    f"""
                    <div class="panel-card">
                        <div class="result-title">File ID {html_text(file_id)}</div>
                        <div class="badge-row"><strong>Storage:</strong> {html_text(classify_storage_type(file_path_value))}</div>
                        <div class="badge-row"><strong>Path:</strong> {html_text(file_path_value)}</div>
                        <div class="badge-row"><strong>Note:</strong> {html_text(note_value)}</div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                for warning_message in file_warnings:
                    st.caption(warning_message)

                if is_external_url(file_path_value):
                    if can_preview_external_image(file_path_value, spectrum_type):
                        preview_url = google_drive_preview_url(file_path_value) if is_google_drive_url(file_path_value) else display_asset_url(file_path_value)
                        if preview_url:
                            st.image(preview_url, caption=f"{spectrum_type} preview", width="stretch")
                    if is_google_drive_url(file_path_value):
                        external_note = "Google Drive link detected. Preview works when the file is shared with viewer access."
                    else:
                        external_note = "External repository link detected."
                    render_external_link_card("Remote file", display_asset_url(file_path_value), external_note)
                    continue

                if full_path is None or not full_path.exists():
                    st.warning("File not found.")
                    if full_path is not None:
                        st.code(str(full_path))
                    continue

                if is_image_file(full_path):
                    st.image(str(full_path), caption=full_path.name, width="stretch")

                elif is_pdf_file(full_path):
                    with open(full_path, "rb") as f:
                        pdf_bytes = f.read()
                    st.download_button(
                        label=f"Download {full_path.name}",
                        data=pdf_bytes,
                        file_name=full_path.name,
                        mime="application/pdf",
                        key=f"pdf_download_{file_id}"
                    )

                else:
                    with open(full_path, "rb") as f:
                        file_bytes = f.read()
                    st.download_button(
                        label=f"Download {full_path.name}",
                        data=file_bytes,
                        file_name=full_path.name,
                        mime="application/octet-stream",
                        key=f"file_download_{file_id}"
                    )

# =========================
# Compound detail
# =========================
def show_compound_detail(compound_id):
    compounds = load_all_compounds()
    row = compounds[compounds["id"] == compound_id]

    if row.empty:
        st.error("Compound not found.")
        return

    proton_df_raw = load_proton_data(compound_id)
    carbon_df_raw = load_carbon_data(compound_id)
    spectra_df_raw = load_spectra_files(compound_id)
    bioactivity_df_raw = load_bioactivity_data(compound_id)
    row_data = row.iloc[0]

    section_header(clean_text(row_data["trivial_name"]))
    st.markdown('<div class="record-shell">', unsafe_allow_html=True)
    st.markdown(
        f"""
        <div class="record-badge-strip compact">
            <span class="record-badge">Status: {html_text(clean_text(normalize_curation_status(row_data.get('curation_status'))).title())}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="action-strip">', unsafe_allow_html=True)
    is_editor = can_edit_database()
    action_col1, action_col2, action_col3, action_col4, action_col5 = st.columns(5)
    with action_col1:
        bundle_bytes, bundle_name = build_compound_record_bundle(
            row,
            proton_df_raw,
            carbon_df_raw,
            spectra_df_raw,
            bioactivity_df_raw,
        )
        st.download_button(
            label="Download Full Record",
            data=bundle_bytes,
            file_name=bundle_name,
            mime="application/zip",
            key=f"download_full_record_{row_data['id']}"
        )
    with action_col2:
        if is_editor:
            if st.button("Edit This Record", key=f"edit_compound_from_detail_{row_data['id']}", width="stretch"):
                open_compound_editor(int(row_data["id"]))
                st.rerun()
        else:
            if st.button("Open Search", key=f"search_from_detail_{row_data['id']}", width="stretch"):
                navigate_internal("Search & Match")
                st.rerun()
    with action_col3:
        if is_editor:
            if st.button("Open 1H Workspace", key=f"open_1h_from_detail_{row_data['id']}", width="stretch"):
                st.session_state["selected_compound_id"] = int(row_data["id"])
                navigate_internal("1H Peaks")
                st.rerun()
        else:
            if st.button("Open Bioactivity", key=f"bioactivity_from_detail_{row_data['id']}", width="stretch"):
                navigate_internal("Bioactivity")
                st.rerun()
    with action_col4:
        if is_editor:
            if st.button("Open 13C Workspace", key=f"open_13c_from_detail_{row_data['id']}", width="stretch"):
                st.session_state["selected_compound_id"] = int(row_data["id"])
                navigate_internal("13C Peaks")
                st.rerun()
        else:
            if st.button("Open Spectra Browser", key=f"spectra_from_detail_{row_data['id']}", width="stretch"):
                st.session_state["selected_compound_id"] = int(row_data["id"])
                navigate_internal("Spectra Library")
                st.rerun()
    with action_col5:
        if is_editor:
            if st.button("Open Spectra Files", key=f"open_spectra_from_detail_{row_data['id']}", width="stretch"):
                st.session_state["selected_compound_id"] = int(row_data["id"])
                navigate_internal("Spectra Library")
                st.rerun()
        else:
            st.button("Read-only Access", key=f"readonly_detail_{row_data['id']}", disabled=True, width="stretch")
    st.markdown('</div>', unsafe_allow_html=True)

    info_left, info_mid, info_right = st.columns([1.12, 1.12, 0.96])

    with info_left:
        section_header("Structure & Identity")
        render_kv("IUPAC Name", row_data["iupac_name"])
        render_kv("Molecular Formula", row_data["molecular_formula"])
        render_kv("Mr", row_data["molecular_weight"])
        render_kv("SMILES", row_data.get("smiles"))
        render_kv("InChI", row_data.get("inchi"))
        render_kv("InChIKey", row_data.get("inchikey"))
        render_kv("Compound Class", row_data["compound_class"])
        render_kv("Compound Subclass", row_data["compound_subclass"])

    with info_mid:
        section_header("Origin & Reference")
        render_kv("Source Category", row_data.get("source_category"))
        render_kv("Source Organism / Species", row_data.get("source_organism"))
        render_kv("Source Summary", source_summary_from_record(row_data))
        render_kv("Sample Code", row_data["sample_code"])
        render_kv("Collection Location", row_data["collection_location"])
        render_kv("GPS Coordinates", row_data["gps_coordinates"])
        render_kv("Depth (m)", row_data["depth_m"])
        render_kv("Data Source", row_data["data_source"])
        render_kv("Curation Status", clean_text(normalize_curation_status(row_data.get("curation_status"))).title())
        render_kv("Journal Name", row_data["journal_name"])
        render_kv("Article Title", row_data["article_title"])
        render_kv(
            "Publication Year / Volume / Issue / Pages",
            f"{clean_text(row_data['publication_year'])} / {clean_text(row_data['volume'])} / {clean_text(row_data['issue'])} / {clean_text(row_data['pages'])}"
        )
        render_kv("DOI", row_data["doi"])
        render_kv("CCDC", row_data["ccdc_number"])

    with info_right:
        section_header("Structure")
        st.markdown('<div class="structure-card">', unsafe_allow_html=True)
        structure_path = row_data["structure_image_path"]
        if pd.notna(structure_path) and str(structure_path).strip():
            standardized_image = load_standardized_structure_source(structure_path, size=(520, 360))
            if standardized_image is not None:
                st.image(standardized_image, width="stretch")
            else:
                st.warning("Structure image file not found.")
                if is_external_url(str(structure_path).strip()):
                    st.code(str(structure_path).strip())
                else:
                    full_path = get_full_file_path(structure_path)
                    if full_path:
                        st.code(str(full_path))
        else:
            render_structure_preview(row_data.get("smiles"), caption=None, size=(520, 360))
        st.markdown('</div>', unsafe_allow_html=True)

    section_header("Physical, Spectral & Supporting Data")
    spectral_col1, spectral_col2, spectral_col3 = st.columns(3)
    with spectral_col1:
        render_kv("UV Data", row_data["uv_data"])
        render_kv("FTIR Data", row_data["ftir_data"])
        render_kv("CD / ECD Data", row_data.get("cd_data"))
    with spectral_col2:
        render_kv("Optical Rotation", row_data["optical_rotation"])
        render_kv("HRMS", row_data["hrms_data"])
        render_kv("Melting Point", row_data["melting_point"])
    with spectral_col3:
        render_kv("Crystallization Method", row_data["crystallization_method"])
        render_kv("Structure Image Path", row_data["structure_image_path"])
        render_kv("Reference DOI / Journal", f"{clean_text(row_data['doi'])} / {clean_text(row_data['journal_name'])}")

    section_header("1H NMR Table")
    if proton_df_raw.empty:
        st.info("No 1H NMR data available.")
    else:
        proton_df = proton_df_raw.rename(columns={
            "id": "ID",
            "delta_ppm": "δH (ppm)",
            "multiplicity": "Multiplicity",
            "j_value": "J Value",
            "proton_count": "Proton Count",
            "assignment": "Assignment",
            "solvent": "Solvent",
            "instrument_mhz": "Instrument (MHz)",
            "note": "Note"
        })
        st.dataframe(proton_df, width="stretch", hide_index=True)

    section_header("13C NMR Table")
    if carbon_df_raw.empty:
        st.info("No 13C NMR data available.")
    else:
        carbon_df = carbon_df_raw.rename(columns={
            "id": "ID",
            "delta_ppm": "δC (ppm)",
            "carbon_type": "Carbon Type",
            "assignment": "Assignment",
            "solvent": "Solvent",
            "instrument_mhz": "Instrument (MHz)",
            "note": "Note"
        })
        st.dataframe(carbon_df, width="stretch", hide_index=True)

    render_spectra_section(compound_id)
    render_bioactivity_table(compound_id)

    section_header("Notes")
    st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    st.write(clean_text(row_data["note"]))
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def render_best_match_summary(item, mode_label):
    source_summary = html_text(source_summary_from_record(item))
    st.markdown("### Best Match Summary")
    st.markdown(
        f"""
        <div class="best-match-card">
            <div class="result-title">{html_text(item['trivial_name'])}</div>
            <div class="result-subtitle">Compound ID: {html_text(item['compound_id'])}</div>
            <div class="badge-row"><strong>Mode:</strong> {html_text(mode_label)}</div>
            <div class="badge-row"><strong>Molecular Formula:</strong> {html_text(item.get('molecular_formula'))}</div>
            <div class="badge-row"><strong>Compound Class:</strong> {html_text(item.get('compound_class'))}</div>
            <div class="badge-row"><strong>Compound Subclass:</strong> {html_text(item.get('compound_subclass'))}</div>
            <div class="badge-row"><strong>Source:</strong> {source_summary}</div>
            <div class="badge-row"><strong>Sample Code:</strong> {html_text(item.get('sample_code'))}</div>
            <div class="badge-row"><strong>Data Source:</strong> {html_text(item.get('data_source'))}</div>
            <div class="badge-row"><strong>Status:</strong> {html_text(clean_text(normalize_curation_status(item.get('curation_status'))).title())}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

def render_candidate_cards(results, mode="13C", limit=10):
    if not results:
        st.info("No matching candidates found.")
        return

    top = results[0]

    if mode == "13C":
        render_best_match_summary(top, "13C similarity search")
    elif mode == "1H":
        render_best_match_summary(top, "1H similarity search")
    else:
        render_best_match_summary(top, "Combined 1H + 13C similarity search")

    section_header(
        "Candidate Ranking",
        "Ranking blends query coverage, database coverage, and how closely the matched peaks align."
    )

    for i, item in enumerate(results[:limit], start=1):
        title_text = clean_text(item["trivial_name"])
        title = html_text(title_text)
        formula = html_text(item.get("molecular_formula"))
        compound_class = html_text(item.get("compound_class"))
        subclass = html_text(item.get("compound_subclass"))
        source_summary = html_text(source_summary_from_record(item))
        sample_code = html_text(item.get("sample_code"))
        data_source = html_text(item.get("data_source"))

        if mode == "13C":
            subtitle = (
                f"Score: {item['score']:.1f}% | Matched: {item['match_count']}/{item['total_query']} | "
                f"Query Coverage: {item['query_coverage'] * 100:.1f}% | DB Coverage: {item['db_coverage'] * 100:.1f}%"
            )
            progress_value = item["score"] / 100
        elif mode == "1H":
            subtitle = (
                f"Score: {item['score']:.1f}% | Matched: {item['match_count']}/{item['total_query']} | "
                f"Query Coverage: {item['query_coverage'] * 100:.1f}% | DB Coverage: {item['db_coverage'] * 100:.1f}%"
            )
            progress_value = item["score"] / 100
        else:
            subtitle = (
                f"Total Score: {item['total_score']:.1f}% | "
                f"1H: {item['proton_match_count']}/{item['proton_total_query']} ({item['proton_score']:.1f}%) | "
                f"13C: {item['carbon_match_count']}/{item['carbon_total_query']} ({item['carbon_score']:.1f}%)"
            )
            progress_value = item["total_score"] / 100

        subtitle_html = html_text(subtitle)

        with st.expander(f"#{i} · {title_text}", expanded=(i == 1)):
            st.markdown(
                f"""
                <div class="result-card">
                    <div class="result-title">{title}</div>
                    <div class="result-subtitle">{subtitle_html}</div>
                    <div class="badge-row"><strong>Compound ID:</strong> {html_text(item['compound_id'])}</div>
                    <div class="badge-row"><strong>Molecular Formula:</strong> {formula}</div>
                    <div class="badge-row"><strong>Compound Class:</strong> {compound_class}</div>
                    <div class="badge-row"><strong>Compound Subclass:</strong> {subclass}</div>
                    <div class="badge-row"><strong>Source:</strong> {source_summary}</div>
                    <div class="badge-row"><strong>Sample Code:</strong> {sample_code}</div>
                    <div class="badge-row"><strong>Data Source:</strong> {data_source}</div>
                </div>
                """,
                unsafe_allow_html=True
            )

            st.progress(progress_value)

            if mode in ["13C", "1H"]:
                diff_text = (
                    f"{item['avg_difference']:.4f} ppm"
                    if item["avg_difference"] is not None else "-"
                )
                st.caption(
                    f"Average difference: {diff_text} | DB peaks: {item['db_peak_count']}"
                )
            else:
                diff_text = (
                    f"{item['avg_difference']:.4f} ppm"
                    if item["avg_difference"] is not None else "-"
                )
                st.caption(f"Average difference across matched peaks: {diff_text}")

            action_left, action_right = st.columns([1, 1])
            with action_left:
                if st.button(
                    f"Open Record #{i}",
                    key=f"open_detail_{mode}_{item['compound_id']}_{i}"
                ):
                    open_compound_detail(item["compound_id"])
                    st.rerun()

            with action_right:
                if can_edit_database():
                    if st.button(
                        f"Update Metadata #{i}",
                        key=f"edit_compound_{mode}_{item['compound_id']}_{i}"
                    ):
                        open_compound_editor(item["compound_id"])
                        st.rerun()
                else:
                    st.button(
                        f"View Only #{i}",
                        key=f"view_only_{mode}_{item['compound_id']}_{i}",
                        disabled=True,
                    )

            if mode == "13C" and item["matches"]:
                st.markdown("**Matched 13C Peaks**")
                match_df = pd.DataFrame(item["matches"], columns=["Query Peak", "DB Peak", "Difference"])
                st.dataframe(match_df, width="stretch", hide_index=True)

            elif mode == "1H" and item["matches"]:
                st.markdown("**Matched 1H Peaks**")
                match_df = pd.DataFrame(item["matches"], columns=["Query Peak", "DB Peak", "Difference"])
                st.dataframe(match_df, width="stretch", hide_index=True)

            elif mode == "combined":
                if item["proton_matches"]:
                    st.markdown("**Matched 1H Peaks**")
                    proton_df = pd.DataFrame(item["proton_matches"], columns=["Query Peak", "DB Peak", "Difference"])
                    st.dataframe(proton_df, width="stretch", hide_index=True)

                if item["carbon_matches"]:
                    st.markdown("**Matched 13C Peaks**")
                    carbon_df = pd.DataFrame(item["carbon_matches"], columns=["Query Peak", "DB Peak", "Difference"])
                    st.dataframe(carbon_df, width="stretch", hide_index=True)

                    # =========================
# Search page
# =========================
def show_search_page(all_compounds_df):
    section_header("Search & Match")

    if "structure_editor_nonce" not in st.session_state:
        st.session_state["structure_editor_nonce"] = 0

    search_mode = st.radio(
        "Search Mode",
        ["Keyword Search", "Structure Search", "13C Match", "1H Match", "Combined Match"],
        horizontal=True
    )

    with st.sidebar.expander("Search Filters", expanded=True):
        search_class_filter = st.selectbox(
            "Compound Class",
            build_filter_options(all_compounds_df, "compound_class"),
            key="search_class_filter"
        )
        search_source_filter = st.selectbox(
            "Source Category",
            build_filter_options(all_compounds_df, "source_category"),
            key="search_source_filter"
        )
        search_data_source_filter = st.selectbox(
            "Data Source",
            build_filter_options(all_compounds_df, "data_source"),
            key="search_data_source_filter"
        )
        min_similarity_score = st.slider(
            "Minimum similarity score (%)",
            min_value=0,
            max_value=100,
            value=35,
            step=5,
            key="search_min_similarity_score",
        )
        candidate_limit = st.slider(
            "Candidates to display",
            min_value=3,
            max_value=20,
            value=10,
            key="search_candidate_limit",
        )

    if search_mode == "Structure Search":
        filtered_df = apply_dataframe_filters(
            all_compounds_df,
            class_filter=search_class_filter,
            source_filter=search_source_filter,
            data_source_filter=search_data_source_filter
        )

        query_smiles = maybe_blank(st.session_state.get("structure_query_smiles"))

        st.markdown('<div class="panel-card">', unsafe_allow_html=True)
        st.markdown('<div class="structure-search-editor-title">Search by Structure</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="structure-search-editor-subtitle">Draw directly in the embedded editor, or paste SMILES / Molfile when you already have a structure string.</div>',
            unsafe_allow_html=True,
        )
        previous_query_smiles = maybe_blank(st.session_state.get("structure_query_smiles"))
        editor_mode_message = ""
        draw_tab, paste_tab = st.tabs(["Draw Structure", "Paste SMILES / Molfile"])
        with draw_tab:
            if st_ketcher is not None:
                drawn_smiles = st_ketcher(
                    value=previous_query_smiles,
                    height=410,
                    molecule_format="SMILES",
                    key=f"structure_search_editor_primary_{st.session_state['structure_editor_nonce']}",
                )
                drawn_smiles_text = "" if drawn_smiles in {None, 0, "0"} else maybe_blank(drawn_smiles)
                if drawn_smiles_text != previous_query_smiles:
                    st.session_state["structure_query_smiles"] = drawn_smiles_text
                    clear_structure_search_state()
                query_smiles = drawn_smiles_text
            elif streamlit_ketchersa is not None:
                drawn_structure = streamlit_ketchersa(height="410px", key=f"structure_search_editor_full_{st.session_state['structure_editor_nonce']}")
                drawn_structure_text = "" if drawn_structure in {None, 0, "0"} else maybe_blank(drawn_structure)
                if drawn_structure_text != previous_query_smiles:
                    st.session_state["structure_query_smiles"] = drawn_structure_text
                    clear_structure_search_state()
                query_smiles = drawn_structure_text
            else:
                st.warning("The direct drawing editor is not active in this deployment yet.")
                st.caption(f"Editor status: {KETCHER_STATUS}")
                query_smiles = st.text_area(
                    "Query Structure (SMILES or Molfile)",
                    key="structure_query_smiles",
                    height=180,
                    placeholder="Example: C1=CC=CC=C1",
                )
                editor_mode_message = "Fallback mode is active. You can still search by pasting a valid SMILES or Molfile query."
                if maybe_blank(query_smiles) != previous_query_smiles:
                    clear_structure_search_state()
        with paste_tab:
            manual_query = st.text_area(
                "Paste or refine the current query",
                value=query_smiles or previous_query_smiles,
                key="structure_query_paste_area",
                height=140,
                placeholder="Paste a valid SMILES or Molfile here.",
            )
            paste_col1, paste_col2 = st.columns(2)
            with paste_col1:
                if st.button("Use Pasted Query", width="stretch", key="apply_pasted_structure_query"):
                    st.session_state["structure_query_smiles"] = maybe_blank(manual_query)
                    query_smiles = maybe_blank(manual_query)
                    clear_structure_search_state()
                    st.rerun()
            with paste_col2:
                if st.button("Clear Query", width="stretch", key="clear_structure_query"):
                    st.session_state["structure_query_smiles"] = ""
                    st.session_state["structure_query_paste_area"] = ""
                    st.session_state["structure_editor_nonce"] += 1
                    clear_structure_search_state()
                    st.rerun()
        if editor_mode_message:
            st.markdown(f'<div class="structure-editor-note">{editor_mode_message}</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        run_structure_search = False
        mode_col, search_col = st.columns([3.4, 1.25], gap="medium")
        with mode_col:
            structure_search_type = st.radio(
                "Structure search mode",
                ["Identity Search", "Similarity Search", "Substructure Search"],
                horizontal=True,
                label_visibility="collapsed",
                key="structure_search_type",
            )
        with search_col:
            run_structure_search = st.button("Search by Structure", width="stretch", key="run_structure_search")

        if query_smiles:
            if can_edit_database():
                preview_col, admin_col, save_col, spacer_col = st.columns([0.78, 1.28, 1.46, 1.38], gap="small")
            else:
                preview_col, spacer_col = st.columns([0.78, 4.12], gap="small")
                admin_col = save_col = None
            with preview_col:
                render_structure_preview(query_smiles, caption="Current query", empty_message=False, size=(210, 150))
        else:
            preview_col = admin_col = save_col = spacer_col = None

        if query_smiles and can_edit_database():
            target_df = all_compounds_df.copy()
            missing_df = target_df[
                target_df["smiles"].fillna("").astype(str).str.strip().eq("")
                & target_df["inchi"].fillna("").astype(str).str.strip().eq("")
                & target_df["inchikey"].fillna("").astype(str).str.strip().eq("")
            ]
            preferred_df = missing_df if not missing_df.empty else target_df
            preferred_df = preferred_df[["id", "trivial_name"]].copy()
            preferred_df["label"] = preferred_df["id"].astype(str) + " - " + preferred_df["trivial_name"].fillna("Unnamed record").astype(str)
            with admin_col:
                selected_structure_label = st.selectbox(
                    "Save current structure to compound",
                    preferred_df["label"].tolist(),
                    key="structure_link_target_select",
                    label_visibility="collapsed",
                )
            with save_col:
                if st.button("Save Structure IDs to Selected Compound", width="stretch", key="save_structure_ids_from_query"):
                    target_compound_id = int(selected_structure_label.split(" - ")[0])
                    saved, save_message = save_structure_query_to_compound(target_compound_id, query_smiles)
                    if saved:
                        st.success(save_message)
                        st.rerun()
                    else:
                        st.error(save_message)

        if run_structure_search:
            results, error_message = search_by_structure(
                filtered_df,
                query_smiles=query_smiles,
                search_type=structure_search_type,
                similarity_threshold=float(min_similarity_score) / 100.0,
            )
            st.session_state["structure_search_results"] = results
            st.session_state["structure_search_error"] = error_message
            st.session_state["structure_search_mode_label"] = structure_search_type
            st.session_state["structure_search_attempted"] = True

        structure_error = maybe_blank(st.session_state.get("structure_search_error"))
        structure_results = st.session_state.get("structure_search_results", [])
        structure_mode_label = maybe_blank(st.session_state.get("structure_search_mode_label")) or structure_search_type
        structure_attempted = bool(st.session_state.get("structure_search_attempted"))

        if structure_error:
            if structure_error.lower().startswith("no compounds matched") or structure_error.lower().startswith("no searchable structures"):
                st.info(structure_error)
            else:
                st.error(structure_error)
        elif structure_results:
            st.write(f"Found {len(structure_results)} compound(s) for the current structure query.")
            st.markdown('<div class="panel-card">', unsafe_allow_html=True)
            top_score = float(structure_results[0].get("structure_score", 0.0)) if structure_results else 0.0
            st.markdown(
                f'''
                <div class="query-summary-grid">
                    <div class="query-summary-card">
                        <div class="query-summary-label">Search Mode</div>
                        <div class="query-summary-value">{structure_mode_label}</div>
                    </div>
                    <div class="query-summary-card">
                        <div class="query-summary-label">Top Similarity / Match</div>
                        <div class="query-summary-value">{top_score:.1f}%</div>
                    </div>
                    <div class="query-summary-card">
                        <div class="query-summary-label">Candidates Returned</div>
                        <div class="query-summary-value">{len(structure_results)}</div>
                    </div>
                </div>
                ''',
                unsafe_allow_html=True,
            )
            st.caption("Results are ranked against the structures currently available in your own database.")
            st.markdown('</div>', unsafe_allow_html=True)
            render_structure_search_results(structure_results, structure_mode_label, limit=candidate_limit)
        elif structure_attempted:
            st.info("The structure query was submitted, but no result rows were returned.")

    elif search_mode == "Keyword Search":
        filtered_df = apply_dataframe_filters(
            all_compounds_df,
            class_filter=search_class_filter,
            source_filter=search_source_filter,
            data_source_filter=search_data_source_filter
        )

        with st.form("search_by_name_form"):
            left, right = st.columns([1.55, 1])
            with left:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                keyword = st.text_input(
                    "Search expression",
                    key="search_name_keyword",
                    placeholder="Name, formula, DOI, sample code, source, or other indexed metadata",
                )
                st.markdown('</div>', unsafe_allow_html=True)
            with right:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                search_field_label = st.selectbox(
                    "Search Field",
                    list(SEARCH_FIELD_COLUMN_MAP.keys()),
                    key="search_field_label",
                )
                search_match_mode = st.selectbox(
                    "Match Mode",
                    ["All keywords", "Exact phrase", "Starts with"],
                    key="search_match_mode",
                )
                run_name_search = st.form_submit_button("Run Keyword Search", width="stretch")
                st.markdown('</div>', unsafe_allow_html=True)

        if keyword.strip():
            result = filtered_df[field_search_mask(filtered_df, keyword, search_field_label, search_match_mode)].copy()
            st.write(f"Found {len(result)} compound(s).")

            if not result.empty:
                export_df = export_name_results(result)
                download_dataframe_button(
                    label="Download Search Results as Excel",
                    df=export_df,
                    file_name="search_by_name_results.xlsx",
                    key="download_name_xlsx",
                    sheet_name="Keyword Search",
                )
                st.dataframe(export_df, width="stretch", hide_index=True, height=336)

                section_header("Quick Browse")
                for _, row in result.head(candidate_limit).iterrows():
                    c1, c2 = st.columns([5, 1])
                    with c1:
                        render_compound_card(row, show_preview=True)
                    with c2:
                        st.write("")
                        if st.button("Open", key=f"name_open_{row['id']}"):
                            open_compound_detail(int(row["id"]))
                            st.rerun()
            else:
                st.info("No compounds matched all keywords in your query.")
        elif run_name_search:
            st.warning("Please enter at least one keyword.")
        elif not filtered_df.empty:
            st.info("Type one or more keywords to search. The filtered dataset preview is shown below.")
            preview_df = export_name_results(filtered_df)
            st.dataframe(preview_df, width="stretch", hide_index=True, height=336)
        else:
            st.info("No compounds available for the selected filters.")

    elif search_mode == "13C Match":
        with st.form("search_13c_form"):
            left, right = st.columns([1.35, 1])
            with left:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                carbon_text = st.text_area("Enter 13C peaks (comma, space, or newline separated)", height=140)
                carbon_file = st.file_uploader(
                    "Or upload peak text / JCAMP-DX",
                    type=PEAK_UPLOAD_TYPES,
                    key="carbon_peak_upload",
                )
                st.markdown('</div>', unsafe_allow_html=True)
            with right:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                carbon_tol = st.number_input("13C tolerance", min_value=0.0, value=0.5, step=0.1)
                run_13c = st.form_submit_button("Run 13C Match", width="stretch")
                st.markdown('<div class="small-note">Example: 145.2, 122.8, 77.1, 38.5. Peak files can also be pasted from text or uploaded as JCAMP-DX.</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

        if run_13c:
            uploaded_peaks, upload_message = parse_peak_upload(carbon_file)
            query_carbons = parse_peak_input(carbon_text) + uploaded_peaks
            if upload_message:
                st.caption(upload_message)
            if not query_carbons:
                st.warning("Please enter at least one valid 13C peak.")
            else:
                results = search_similarity_13c(query_carbons, carbon_tol)
                filtered_results = filter_similarity_results(
                    results,
                    class_filter=search_class_filter,
                    source_filter=search_source_filter,
                    data_source_filter=search_data_source_filter
                )
                filtered_results = [
                    item for item in filtered_results
                    if item["score"] >= float(min_similarity_score)
                ]

                st.caption(
                    "Similarity score uses query coverage, database coverage, and average peak closeness."
                )
                st.write(f"Found {len(filtered_results)} candidate(s) above the current score threshold.")

                if filtered_results:
                    export_df = export_similarity_results_13c(filtered_results)
                    download_dataframe_button(
                        label="Download 13C Similarity Results as Excel",
                        df=export_df,
                        file_name="search_by_13c_results.xlsx",
                        key="download_13c_xlsx",
                        sheet_name="13C Match",
                    )

                render_candidate_cards(filtered_results, mode="13C", limit=candidate_limit)

    elif search_mode == "1H Match":
        with st.form("search_1h_form"):
            left, right = st.columns([1.35, 1])
            with left:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                proton_text = st.text_area("Enter 1H peaks (comma, space, or newline separated)", height=140)
                proton_file = st.file_uploader(
                    "Or upload peak text / JCAMP-DX",
                    type=PEAK_UPLOAD_TYPES,
                    key="proton_peak_upload",
                )
                st.markdown('</div>', unsafe_allow_html=True)
            with right:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                proton_tol = st.number_input("1H tolerance", min_value=0.0, value=0.05, step=0.01, format="%.2f")
                run_1h = st.form_submit_button("Run 1H Match", width="stretch")
                st.markdown('<div class="small-note">Example: 5.82, 5.35, 3.21, 1.22. Peak files can also be uploaded for quick screening.</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

        if run_1h:
            uploaded_peaks, upload_message = parse_peak_upload(proton_file)
            query_protons = parse_peak_input(proton_text) + uploaded_peaks
            if upload_message:
                st.caption(upload_message)
            if not query_protons:
                st.warning("Please enter at least one valid 1H peak.")
            else:
                results = search_similarity_1h(query_protons, proton_tol)
                filtered_results = filter_similarity_results(
                    results,
                    class_filter=search_class_filter,
                    source_filter=search_source_filter,
                    data_source_filter=search_data_source_filter
                )
                filtered_results = [
                    item for item in filtered_results
                    if item["score"] >= float(min_similarity_score)
                ]

                st.caption(
                    "Similarity score uses query coverage, database coverage, and average peak closeness."
                )
                st.write(f"Found {len(filtered_results)} candidate(s) above the current score threshold.")

                if filtered_results:
                    export_df = export_similarity_results_1h(filtered_results)
                    download_dataframe_button(
                        label="Download 1H Similarity Results as Excel",
                        df=export_df,
                        file_name="search_by_1h_results.xlsx",
                        key="download_1h_xlsx",
                        sheet_name="1H Match",
                    )

                render_candidate_cards(filtered_results, mode="1H", limit=candidate_limit)

    else:
        with st.form("search_combined_form"):
            left, right = st.columns([1.4, 1])
            with left:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                proton_text = st.text_area("Enter 1H peaks (comma, space, or newline separated)", height=120, key="combined_proton_text")
                proton_file = st.file_uploader(
                    "Optional 1H peak upload",
                    type=PEAK_UPLOAD_TYPES,
                    key="combined_proton_file",
                )
                carbon_text = st.text_area("Enter 13C peaks (comma, space, or newline separated)", height=120, key="combined_carbon_text")
                carbon_file = st.file_uploader(
                    "Optional 13C peak upload",
                    type=PEAK_UPLOAD_TYPES,
                    key="combined_carbon_file",
                )
                st.markdown('</div>', unsafe_allow_html=True)
            with right:
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                proton_tol = st.number_input("1H tolerance", min_value=0.0, value=0.05, step=0.01, format="%.2f", key="combined_1h")
                carbon_tol = st.number_input("13C tolerance", min_value=0.0, value=0.5, step=0.1, key="combined_13c")
                run_combined = st.form_submit_button("Run Combined Match", width="stretch")
                st.markdown('<div class="small-note">Use both peak lists for more selective candidate ranking.</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

        if run_combined:
            proton_upload_peaks, proton_upload_message = parse_peak_upload(proton_file)
            carbon_upload_peaks, carbon_upload_message = parse_peak_upload(carbon_file)
            query_protons = parse_peak_input(proton_text) + proton_upload_peaks
            query_carbons = parse_peak_input(carbon_text) + carbon_upload_peaks
            if proton_upload_message:
                st.caption(proton_upload_message)
            if carbon_upload_message:
                st.caption(carbon_upload_message)

            if not query_protons and not query_carbons:
                st.warning("Please enter at least one valid 1H or 13C peak.")
            else:
                results = search_similarity_combined(query_protons, proton_tol, query_carbons, carbon_tol)
                filtered_results = filter_similarity_results(
                    results,
                    class_filter=search_class_filter,
                    source_filter=search_source_filter,
                    data_source_filter=search_data_source_filter
                )
                filtered_results = [
                    item for item in filtered_results
                    if item["total_score"] >= float(min_similarity_score)
                ]

                st.caption(
                    "Combined ranking averages the improved 1H and 13C similarity scores."
                )
                st.write(f"Found {len(filtered_results)} candidate(s) above the current score threshold.")

                if filtered_results:
                    export_df = export_similarity_results_combined(filtered_results)
                    download_dataframe_button(
                        label="Download Combined Similarity Results as Excel",
                        df=export_df,
                        file_name="search_combined_results.xlsx",
                        key="download_combined_xlsx",
                        sheet_name="Combined Match",
                    )

                render_candidate_cards(filtered_results, mode="combined", limit=candidate_limit)


# =========================
# Overview page
# =========================
def show_overview_page(all_compounds_df):
    totals_snapshot = count_database_totals(get_db_signature())
    overall_proton_count = int(totals_snapshot["proton"])
    overall_carbon_count = int(totals_snapshot["carbon"])
    overall_spectra_count = int(totals_snapshot["spectra"])
    overall_bioactivity_count = int(totals_snapshot["bioactivity"])

    with st.sidebar.expander("Dashboard Filters", expanded=True):
        dashboard_class_filter = st.selectbox(
            "Compound Class",
            build_filter_options(all_compounds_df, "compound_class"),
            key="dashboard_class"
        )
        dashboard_subclass_filter = st.selectbox(
            "Compound Subclass",
            build_filter_options(all_compounds_df, "compound_subclass"),
            key="dashboard_subclass"
        )
        dashboard_source_filter = st.selectbox(
            "Source Category",
            build_filter_options(all_compounds_df, "source_category"),
            key="dashboard_source"
        )
        dashboard_data_source_filter = st.selectbox(
            "Data Source",
            build_filter_options(all_compounds_df, "data_source"),
            key="dashboard_data_source"
        )
    filtered_df = apply_dataframe_filters(
        all_compounds_df,
        class_filter=dashboard_class_filter,
        subclass_filter=dashboard_subclass_filter,
        source_filter=dashboard_source_filter,
        data_source_filter=dashboard_data_source_filter
    )

    render_dashboard_showcase(
        all_compounds_df,
        overall_proton_count,
        overall_carbon_count,
        overall_spectra_count,
        overall_bioactivity_count,
    )
    st.markdown('<div class="dashboard-section"></div>', unsafe_allow_html=True)
    section_header("Quick Browse")
    quick_browse_limit = st.select_slider(
        "Quick Browse cards",
        options=[4, 6, 8, 10, 12],
        value=6,
        key="dashboard_quick_browse_limit",
    )
    if filtered_df.empty:
        st.info("No compounds available for the selected filters.")
    else:
        preview_df = filtered_df.sort_values(by="updated_at", ascending=False, na_position="last").head(int(quick_browse_limit))
        for _, row in preview_df.iterrows():
            c1, c2 = st.columns([6.2, 0.8])
            with c1:
                render_compound_card(row, show_preview=True)
            with c2:
                st.write("")
                if st.button("Open", key=f"overview_open_{row['id']}"):
                    open_compound_detail(int(row["id"]))
                    st.rerun()

    st.markdown('<div class="dashboard-section"></div>', unsafe_allow_html=True)
    section_header("Distribution Overview", "Interactive composition charts for the current filtered dataset.")
    left, right = st.columns([1.05, 1.05])

    with left:
        section_header("Compound Distribution")
        if filtered_df.empty:
            st.info("No compounds available for the selected filters.")
        else:
            class_counts = (
                filtered_df["compound_class"]
                .fillna("Uncategorized")
                .replace("", "Uncategorized")
                .value_counts()
                .reset_index()
            )
            class_counts.columns = ["Compound Class", "Count"]
            render_dashboard_pie_chart(
                class_counts,
                names_col="Compound Class",
                values_col="Count",
                color_sequence=["#61D8ED", "#4C8EFF", "#9C63F1", "#7EF0C2", "#F2C66D", "#FF7F6D", "#BFA5FF"],
                top_n=6,
            )

    with right:
        section_header("Source Category Distribution")
        if filtered_df.empty:
            st.info("No compounds available for the selected filters.")
        else:
            source_counts = (
                filtered_df["source_category"]
                .fillna("Uncategorized")
                .replace("", "Uncategorized")
                .value_counts()
                .reset_index()
            )
            source_counts.columns = ["Source Category", "Count"]
            render_dashboard_pie_chart(
                source_counts,
                names_col="Source Category",
                values_col="Count",
                color_sequence=["#9C63F1", "#61D8ED", "#4C8EFF", "#FF7F6D", "#F2C66D", "#7EF0C2"],
                top_n=7,
            )


def show_guide_page():
    section_header("Guide", "Complete usage, submission, storage, and access guidance for this database.")

    intro_left, intro_right = st.columns([1.2, 1])
    with intro_left:
        render_helper_card(
            "What this web app is for",
            "This database is designed to connect compounds, structural metadata, spectra previews, raw-data references, and publication details in one searchable workspace.",
        )
    with intro_right:
        render_helper_card(
            "Who can use it",
            "Researchers can search and compare records. Curators can submit, revise, import, and maintain compounds, peaks, and spectra links.",
        )

    section_header("How To Use")
    use_tabs = st.tabs(["Browse", "Submit", "Spectra & Raw Data", "Storage Layout", "Access & Deployment"])

    with use_tabs[0]:
            st.markdown(
                """
                1. Open `Dashboard` to see coverage, quick browsing, and backup.
                2. Use `Search & Match` for keyword lookup or 1H/13C spectral matching.
                3. Open `Compound Workspace` to inspect full records, references, linked files, and bioactivity tabs.
                4. Use `Bioactivity` to curate assay outcomes, potency values, and target annotations separately from the core compound metadata.
                5. Use `1H Peaks`, `13C Peaks`, and `Spectra Library` when you want to manage sub-records directly.
            """
        )

    with use_tabs[1]:
        st.markdown(
            """
            1. Start in `Compound Workspace` > `New Submission`.
            2. Fill the core identity fields first: trivial name, formula, SMILES/InChI/InChIKey, class, subclass, source category/organism, and structure.
            3. Add publication information, notes, and reference fields.
            4. Save the compound record.
            5. Add 1H peaks, 13C peaks, preview images, PDFs, raw-data links, and bioactivity records from their dedicated sections if needed.
            """
        )

    with use_tabs[2]:
        st.markdown(
            """
            1. Keep lightweight preview images locally or in Google Drive if you want them visible directly in the app.
            2. Store large raw 1H/13C data files in Google Drive to avoid filling the laptop.
            3. Paste the Google Drive sharing link into `Spectra Library` so the database stays metadata-first and device-friendly.
            4. Use spectrum types such as `1H`, `13C`, `COSY`, `HSQC`, or `HMBC` for preview images.
            5. Use `1H Raw Data`, `13C Raw Data`, `JCAMP-DX`, or `MNova` for raw downloadable files.
            6. If a Google Drive link points to an image and sharing is allowed, the spectra image can preview directly inside the web app without opening Google Drive first.
            7. For future structure search, keep `SMILES`, `InChI`, and `InChIKey` filled as consistently as possible.
            """
        )
        st.caption("Preview depends on the Google Drive link being shared with the right viewing permission.")

    with use_tabs[3]:
        st.markdown(
            """
            Recommended local folder layout in `Desktop/NMR_Database_Tyas`:

            1. `database/nmr.db` for the main SQLite metadata database.
            2. `database/backups/` for timestamped backup copies before major edits.
            3. `data/structures/` for lightweight structure images only.
            4. `data/spectra/` for lightweight preview images or PDFs only.
            5. `data/templates/` for batch import CSV templates generated by the app.
            6. `data/submissions/inbox/` for newly received material not yet curated.
            7. `data/submissions/reviewed/` for material already checked but not yet approved.
            8. `data/submissions/approved/` for curated source files that match the published record.
            9. `data/exports/` for CSV exports or reports shared with collaborators.
            """
        )
        st.markdown(
            """
            Recommended Google Drive layout:

            1. `NPDB_Public_Previews/` for shareable image previews.
            2. `NPDB_Raw_Data/Compound_Name_or_ID/1H/`
            3. `NPDB_Raw_Data/Compound_Name_or_ID/13C/`
            4. `NPDB_Raw_Data/Compound_Name_or_ID/JCAMP_DX/`
            5. `NPDB_Raw_Data/Compound_Name_or_ID/MNova/`
            6. `NPDB_Submission_Source/Year/LabMember_or_Paper/`
            """
        )
        st.markdown(
            """
            Recommended naming convention:

            1. Structure preview: `NPDB_<compound_id>_<trivial_name>_structure.png`
            2. Spectra preview: `NPDB_<compound_id>_<trivial_name>_<spectrum_type>_preview.png`
            3. Raw file: `NPDB_<compound_id>_<trivial_name>_<nucleus>_raw.<ext>`
            4. JCAMP-DX: `NPDB_<compound_id>_<trivial_name>_jcamp.dx`
            5. MNova: `NPDB_<compound_id>_<trivial_name>_mnova.mnova`
            """
        )
        st.caption("Keep one canonical file per dataset. If a better version appears later, replace the old file and update the database link instead of making silent duplicates.")

    with use_tabs[4]:
        st.markdown(
            """
            1. `http://localhost:8501` is still your local development address. It only works on your own machine while Streamlit is running there.
            2. A local app can sometimes be opened from another device on the same network using your computer IP, but that is temporary and depends on your network and firewall.
            3. If people need stable access from phone, laptop, Windows, macOS, or Linux, deploy the app to a server or cloud platform and share the public HTTPS URL from there.
            4. The current access gate supports either one shared login with `NPDB_ACCESS_USERNAME` and `NPDB_ACCESS_PASSWORD`, or multiple approved users with `NPDB_APPROVED_USERS`.
            5. After deployment, users should open the public URL, not `localhost`.
            6. Mobile access is possible after deployment, but the best experience still needs responsive visual QA.
            """
        )

    section_header("Important Notes")
    note_left, note_right = st.columns(2)
    with note_left:
        render_helper_card(
            "Storage limits",
            "This app does not enforce its own storage quota. The real limits come from your laptop disk, Google Drive quota, and whichever server or hosting platform you use.",
        )
    with note_right:
        render_helper_card(
            "Stable public access",
            "If you want the same address that people can open anytime from anywhere, you will need deployment. A local `localhost` address will not stay public forever and cannot be your permanent access URL.",
        )


# =========================
# Compound pages
# =========================
def show_compound_pages():
    compound_options = COMPOUND_PAGE_OPTIONS if can_edit_database() else ["Browse Record"]

    current_page = st.session_state.get("compound_page", "Browse Record")
    if current_page not in compound_options:
        current_page = "Browse Record"
        st.session_state["compound_page"] = current_page
        st.session_state["_pending_compound_page_radio"] = current_page

    compound_radio_kwargs = {
        "label": "Compound Workflow",
        "options": compound_options,
        "horizontal": True,
        "key": "compound_page_radio",
    }
    if "compound_page_radio" not in st.session_state:
        compound_radio_kwargs["index"] = compound_options.index(current_page)
    compound_page = st.radio(**compound_radio_kwargs)
    st.session_state["compound_page"] = compound_page

    if not can_edit_database():
        render_read_only_notice("submit, edit, import, or delete compound records")

    if compound_page == "Browse Record":
        section_header("Compound Browser")
        compounds_df = load_all_compounds()

        if compounds_df.empty:
            st.info("No compounds available.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            label_list = options["label"].tolist()

            default_index = 0
            selected_id = st.session_state.get("selected_compound_id")
            if selected_id is not None and selected_id in options["id"].tolist():
                default_index = options.index[options["id"] == selected_id][0]

            selected = st.selectbox("Choose compound record", label_list, index=default_index)
            current_selected_id = int(selected.split(" - ")[0])
            st.session_state["selected_compound_id"] = current_selected_id
            show_compound_detail(current_selected_id)

    elif compound_page == "New Submission":
        section_header("New Submission")

        compounds_df = load_all_compounds()
        spectra_df = load_all_spectra_files()
        persist_wizard_inputs()
        for key in [
            "wizard_trivial_name",
            "wizard_iupac_name",
            "wizard_formula",
            "wizard_molecular_weight",
            "wizard_smiles",
            "wizard_inchi",
            "wizard_inchikey",
            "wizard_compound_class_select",
            "wizard_compound_class_custom",
            "wizard_compound_subclass_select",
            "wizard_compound_subclass_custom",
            "wizard_data_source_select",
            "wizard_data_source_custom",
            "wizard_source_category_select",
            "wizard_source_category_custom",
            "wizard_source_organism",
            "wizard_sample_code",
            "wizard_collection_location",
            "wizard_gps_coordinates",
            "wizard_depth_m",
            "wizard_uv_data",
            "wizard_ftir_data",
            "wizard_cd_data",
            "wizard_optical_rotation",
            "wizard_melting_point",
            "wizard_crystallization_method",
            "wizard_ccdc_number",
            "wizard_hrms_data",
            "wizard_structure_path",
            "wizard_submission_spectrum_type_select",
            "wizard_submission_spectrum_type_custom",
            "wizard_submission_spectra_note",
            "wizard_journal_name",
            "wizard_article_title",
            "wizard_publication_year",
            "wizard_volume",
            "wizard_issue",
            "wizard_pages",
            "wizard_doi",
            "wizard_curation_status",
            "wizard_note",
        ]:
            hydrate_wizard_widget(key)
        wizard_step = st.session_state.get("compound_wizard_step", 1)
        step_labels = {
            1: "Identity",
            2: "Origin",
            3: "Spectral Data & Files",
            4: "Reference & Review",
        }

        st.progress(wizard_step / 4)
        st.caption(f"Step {wizard_step} of 4 · {step_labels[wizard_step]}")

        if wizard_step == 1:
            c1, c2 = st.columns(2)
            with c1:
                st.text_input("Trivial Name", key="wizard_trivial_name")
                st.text_area("IUPAC Name", key="wizard_iupac_name")
                st.text_input("Molecular Formula", key="wizard_formula")
                st.text_input("Mr", key="wizard_molecular_weight")
                st.text_area("SMILES", key="wizard_smiles", placeholder="e.g. C1=CC=CC=C1")
            with c2:
                st.text_area("InChI", key="wizard_inchi", placeholder="e.g. InChI=1S/...")
                st.text_input("InChIKey", key="wizard_inchikey", placeholder="e.g. BSYNRYMUTXBXSQ-UHFFFAOYSA-N")
                class_options = build_existing_options(compounds_df, "compound_class", DEFAULT_CLASS_OPTIONS)
                subclass_options = build_existing_options(compounds_df, "compound_subclass")
                data_source_options = build_existing_options(compounds_df, "data_source", DEFAULT_DATA_SOURCE_OPTIONS)
                select_or_custom(
                    "Compound Class",
                    class_options,
                    "wizard_compound_class",
                    help_text="Choose an existing class or use Custom... to add a new compound class.",
                )
                select_or_custom("Compound Subclass", subclass_options, "wizard_compound_subclass")
                select_or_custom("Data Source", data_source_options, "wizard_data_source", value="Experimental")

        elif wizard_step == 2:
            c1, c2 = st.columns(2)
            with c1:
                source_options = build_existing_options(compounds_df, "source_category", DEFAULT_SOURCE_OPTIONS)
                select_or_custom(
                    "Source Category",
                    source_options,
                    "wizard_source_category",
                    help_text="Choose an existing source category or use Custom... to add a new one.",
                )
                st.text_input(
                    "Source Organism / Species (optional)",
                    key="wizard_source_organism",
                    placeholder="e.g. Halicondria sp. or Unknown sponge",
                )
                st.text_input("Sample Code", key="wizard_sample_code")
                st.text_input("Collection Location", key="wizard_collection_location")
            with c2:
                st.text_input("GPS Coordinates", key="wizard_gps_coordinates")
                st.text_input("Depth (m)", key="wizard_depth_m")
                st.text_area("Notes", key="wizard_note")

        elif wizard_step == 3:
            c1, c2 = st.columns(2)
            with c1:
                st.text_input("UV Data", key="wizard_uv_data")
                st.text_input("FTIR Data", key="wizard_ftir_data")
                st.text_area("Circular Dichroism (CD / ECD)", key="wizard_cd_data")
                st.text_input("Optical Rotation", key="wizard_optical_rotation")
                st.text_input("Melting Point", key="wizard_melting_point")
                st.text_input("Crystallization Method", key="wizard_crystallization_method")
                st.text_area(
                    "HRMS Data",
                    key="wizard_hrms_data",
                    placeholder="e.g. HRMS (ESI) m/z: [M + Na]+ calcd..., found...",
                )
                st.text_input("CCDC", key="wizard_ccdc_number")
            with c2:
                st.text_input(
                    "Structure Image Path (optional)",
                    key="wizard_structure_path",
                    placeholder="e.g. data/structures/example.png",
                )
                st.file_uploader(
                    "Upload Structure Image",
                    type=["png", "jpg", "jpeg", "webp"],
                    key="wizard_structure_upload",
                )
                wizard_spectrum_options = build_existing_options(
                    spectra_df,
                    "spectrum_type",
                    DEFAULT_SPECTRUM_TYPES,
                )
                select_or_custom(
                    "Uploaded Spectra Type",
                    wizard_spectrum_options,
                    "wizard_submission_spectrum_type",
                    value="Supporting Data",
                    help_text="All files uploaded in this step will use the same type label. You can fine-tune them later in the Spectra section.",
                )
                st.file_uploader(
                    "Upload Supporting Spectra Files",
                    accept_multiple_files=True,
                    key="wizard_submission_spectra_uploads",
                )
                st.text_area("Uploaded Spectra Note", key="wizard_submission_spectra_note")
                st.caption("Tip: for large raw 1H/13C datasets, store the raw files in Google Drive and register the share link later from Spectra Library. Keep only lightweight preview files locally when necessary.")

        else:
            c1, c2 = st.columns(2)
            with c1:
                st.text_input("Journal Name", key="wizard_journal_name")
                st.text_input("Article Title", key="wizard_article_title")
                st.text_input("Publication Year", key="wizard_publication_year")
                st.text_input("Volume", key="wizard_volume")
                st.text_input("Issue", key="wizard_issue")
                st.text_input("Pages", key="wizard_pages")
                st.text_input("DOI", key="wizard_doi")
                current_wizard_status = normalize_curation_status(
                    get_wizard_value("wizard_curation_status", "curated")
                )
                st.selectbox(
                    "Curation Status",
                    CURATION_STATUS_OPTIONS,
                    index=CURATION_STATUS_OPTIONS.index(current_wizard_status),
                    key="wizard_curation_status",
                    help="Use curated for owner-reviewed manual submissions, reviewed for checked records, and imported for bulk-ingested records awaiting full curation.",
                )
            with c2:
                draft_row = {
                    "trivial_name": get_wizard_value("wizard_trivial_name", ""),
                    "molecular_formula": get_wizard_value("wizard_formula", ""),
                    "smiles": get_wizard_value("wizard_smiles", ""),
                    "inchi": get_wizard_value("wizard_inchi", ""),
                    "inchikey": get_wizard_value("wizard_inchikey", ""),
                    "compound_class": get_wizard_value("wizard_compound_class_custom", "") or get_wizard_value("wizard_compound_class_select", ""),
                    "source_category": get_wizard_value("wizard_source_category_custom", "") or get_wizard_value("wizard_source_category_select", ""),
                    "source_organism": get_wizard_value("wizard_source_organism", ""),
                    "source_material": source_summary_from_record(
                        {
                            "source_category": get_wizard_value("wizard_source_category_custom", "") or get_wizard_value("wizard_source_category_select", ""),
                            "source_organism": get_wizard_value("wizard_source_organism", ""),
                            "source_material": "",
                        }
                    ),
                    "data_source": get_wizard_value("wizard_data_source_custom", "") or get_wizard_value("wizard_data_source_select", ""),
                    "curation_status": normalize_curation_status(get_wizard_value("wizard_curation_status", "curated")),
                    "hrms_data": get_wizard_value("wizard_hrms_data", ""),
                    "doi": get_wizard_value("wizard_doi", ""),
                    "journal_name": get_wizard_value("wizard_journal_name", ""),
                    "article_title": get_wizard_value("wizard_article_title", ""),
                    "structure_image_path": get_wizard_value("wizard_structure_path", "") or ("uploaded" if get_wizard_value("wizard_structure_upload") else ""),
                }
                completeness_preview = calculate_completeness_score(
                    draft_row,
                    pd.DataFrame(),
                    pd.DataFrame(),
                    pd.DataFrame(),
                )
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                st.write(f"**Draft completeness estimate:** {completeness_preview}%")
                st.write(f"**Trivial Name:** {clean_text(draft_row['trivial_name'])}")
                st.write(f"**Formula:** {clean_text(draft_row['molecular_formula'])}")
                st.write(f"**SMILES:** {clean_text(draft_row['smiles'])}")
                st.write(f"**InChIKey:** {clean_text(draft_row['inchikey'])}")
                st.write(f"**Class:** {clean_text(draft_row['compound_class'])}")
                st.write(f"**Source Category:** {clean_text(draft_row['source_category'])}")
                st.write(f"**Source Organism:** {clean_text(draft_row['source_organism'])}")
                st.write(f"**Source Summary:** {clean_text(draft_row['source_material'])}")
                st.write(f"**Data Source:** {clean_text(draft_row['data_source'])}")
                st.write(f"**Curation Status:** {clean_text(draft_row['curation_status']).title()}")
                st.write(f"**Journal:** {clean_text(get_wizard_value('wizard_journal_name'))}")
                st.write(f"**Article:** {clean_text(get_wizard_value('wizard_article_title'))}")
                st.markdown('</div>', unsafe_allow_html=True)

        nav_left, nav_right = st.columns([1, 1])
        with nav_left:
            if wizard_step > 1 and st.button("Back", width="stretch", key=f"wizard_back_{wizard_step}"):
                persist_wizard_inputs()
                st.session_state["compound_wizard_step"] = wizard_step - 1
                st.rerun()

        with nav_right:
            if wizard_step < 4:
                if st.button("Continue", width="stretch", key=f"wizard_next_{wizard_step}"):
                    persist_wizard_inputs()
                    if wizard_step == 1 and not maybe_blank(get_wizard_value("wizard_trivial_name")):
                        st.error("Trivial Name is required before moving to the next step.")
                    else:
                        st.session_state["compound_wizard_step"] = wizard_step + 1
                        st.rerun()
            else:
                if st.button("Save New Record", width="stretch", key="wizard_submit_compound"):
                    persist_wizard_inputs()
                    trivial_name = maybe_blank(get_wizard_value("wizard_trivial_name"))
                    iupac_name = maybe_blank(get_wizard_value("wizard_iupac_name"))
                    molecular_formula = maybe_blank(get_wizard_value("wizard_formula"))
                    smiles = maybe_blank(get_wizard_value("wizard_smiles"))
                    inchi = maybe_blank(get_wizard_value("wizard_inchi"))
                    inchikey = maybe_blank(get_wizard_value("wizard_inchikey"))
                    compound_class = maybe_blank(get_wizard_value("wizard_compound_class_custom")) or maybe_blank(get_wizard_value("wizard_compound_class_select"))
                    compound_subclass = maybe_blank(get_wizard_value("wizard_compound_subclass_custom")) or maybe_blank(get_wizard_value("wizard_compound_subclass_select"))
                    source_category = maybe_blank(get_wizard_value("wizard_source_category_custom")) or maybe_blank(get_wizard_value("wizard_source_category_select"))
                    source_organism = maybe_blank(get_wizard_value("wizard_source_organism"))
                    _, _, source_material = infer_source_fields(source_category, source_organism, "")
                    sample_code = maybe_blank(get_wizard_value("wizard_sample_code"))
                    collection_location = maybe_blank(get_wizard_value("wizard_collection_location"))
                    gps_coordinates = maybe_blank(get_wizard_value("wizard_gps_coordinates"))
                    depth_m_text = maybe_blank(get_wizard_value("wizard_depth_m"))
                    uv_data = maybe_blank(get_wizard_value("wizard_uv_data"))
                    ftir_data = maybe_blank(get_wizard_value("wizard_ftir_data"))
                    cd_data = maybe_blank(get_wizard_value("wizard_cd_data"))
                    optical_rotation = maybe_blank(get_wizard_value("wizard_optical_rotation"))
                    melting_point = maybe_blank(get_wizard_value("wizard_melting_point"))
                    crystallization_method = maybe_blank(get_wizard_value("wizard_crystallization_method"))
                    structure_image_path = maybe_blank(get_wizard_value("wizard_structure_path"))
                    structure_upload = get_wizard_value("wizard_structure_upload")
                    journal_name = maybe_blank(get_wizard_value("wizard_journal_name"))
                    article_title = maybe_blank(get_wizard_value("wizard_article_title"))
                    publication_year = maybe_blank(get_wizard_value("wizard_publication_year"))
                    volume = maybe_blank(get_wizard_value("wizard_volume"))
                    issue = maybe_blank(get_wizard_value("wizard_issue"))
                    pages = maybe_blank(get_wizard_value("wizard_pages"))
                    doi = maybe_blank(get_wizard_value("wizard_doi"))
                    ccdc_number = maybe_blank(get_wizard_value("wizard_ccdc_number"))
                    molecular_weight_text = maybe_blank(get_wizard_value("wizard_molecular_weight"))
                    hrms_data = maybe_blank(get_wizard_value("wizard_hrms_data"))
                    data_source = maybe_blank(get_wizard_value("wizard_data_source_custom")) or maybe_blank(get_wizard_value("wizard_data_source_select"))
                    curation_status = normalize_curation_status(get_wizard_value("wizard_curation_status", "curated"))
                    note = maybe_blank(get_wizard_value("wizard_note"))
                    uploaded_spectra = get_wizard_value("wizard_submission_spectra_uploads") or []
                    uploaded_spectrum_type = maybe_blank(get_wizard_value("wizard_submission_spectrum_type_custom")) or maybe_blank(get_wizard_value("wizard_submission_spectrum_type_select")) or "Supporting Data"
                    uploaded_spectrum_note = maybe_blank(get_wizard_value("wizard_submission_spectra_note"))

                    if not trivial_name:
                        st.error("Trivial Name is required.")
                        st.stop()

                    depth_value = safe_float_or_none(depth_m_text)
                    if depth_m_text and depth_value is None:
                        st.error("Depth (m) must be a valid number.")
                        st.stop()

                    molecular_weight_value = safe_float_or_none(molecular_weight_text)
                    if molecular_weight_text and molecular_weight_value is None:
                        st.error("Mr must be a valid number.")
                        st.stop()

                    if structure_upload is not None:
                        structure_image_path = save_uploaded_asset(
                            structure_upload,
                            STRUCTURES_DIR,
                            f"{trivial_name}_{sample_code or 'structure'}",
                        )

                    new_id = insert_compound_record(
                        trivial_name=trivial_name,
                        iupac_name=iupac_name,
                        molecular_formula=molecular_formula,
                        compound_class=compound_class,
                        compound_subclass=compound_subclass,
                        smiles=smiles,
                        inchi=inchi,
                        inchikey=inchikey,
                        source_category=source_category,
                        source_organism=source_organism,
                        source_material=source_material,
                        sample_code=sample_code,
                        collection_location=collection_location,
                        gps_coordinates=gps_coordinates,
                        depth_m=depth_value,
                        uv_data=uv_data,
                        ftir_data=ftir_data,
                        cd_data=cd_data,
                        optical_rotation=optical_rotation,
                        melting_point=melting_point,
                        crystallization_method=crystallization_method,
                        structure_image_path=structure_image_path,
                        journal_name=journal_name,
                        article_title=article_title,
                        publication_year=publication_year,
                        volume=volume,
                        issue=issue,
                        pages=pages,
                        doi=doi,
                        ccdc_number=ccdc_number,
                        molecular_weight=molecular_weight_value,
                        hrms_data=hrms_data,
                        data_source=data_source,
                        curation_status=curation_status,
                        note=note,
                    )

                    for uploaded_file in uploaded_spectra:
                        saved_path = save_uploaded_asset(
                            uploaded_file,
                            SPECTRA_DIR,
                            f"compound_{new_id}_{uploaded_spectrum_type}_{uploaded_file.name}",
                        )
                        insert_spectrum_file_record(
                            compound_id=new_id,
                            spectrum_type=uploaded_spectrum_type,
                            file_path=saved_path,
                            note=uploaded_spectrum_note,
                        )

                    st.success(f"Record saved successfully. New Compound ID: {new_id}")
                    reset_compound_wizard()
                    open_compound_detail(new_id)
                    st.rerun()

    elif compound_page == "Batch Import":
        render_batch_import_workspace()

    elif compound_page == "Update Metadata":
        section_header("Update Metadata", "Revise compound metadata, references, and structure links without leaving the database workspace.")
        compounds_df = load_all_compounds()

        if compounds_df.empty:
            st.info("No compounds available.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            label_list = options["label"].tolist()

            default_index = 0
            selected_id = st.session_state.get("selected_compound_id")
            if selected_id is not None and selected_id in options["id"].tolist():
                default_index = options.index[options["id"] == selected_id][0]

            selected_label = st.selectbox(
                "Select record to edit",
                label_list,
                index=default_index,
                key="edit_compound_select"
            )

            edit_compound_id = int(selected_label.split(" - ")[0])
            st.session_state["selected_compound_id"] = edit_compound_id

            row_df = load_compound_row(edit_compound_id)
            if row_df.empty:
                st.error("Record not found.")
            else:
                row = row_df.iloc[0]


                col1, col2 = st.columns(2)

                with col1:
                    trivial_name = st.text_input("Trivial Name", value=maybe_blank(row["trivial_name"]))
                    iupac_name = st.text_area("IUPAC Name", value=maybe_blank(row["iupac_name"]))
                    molecular_formula = st.text_input("Molecular Formula", value=maybe_blank(row["molecular_formula"]))
                    smiles = st.text_area("SMILES", value=maybe_blank(row.get("smiles")))
                    inchi = st.text_area("InChI", value=maybe_blank(row.get("inchi")))
                    inchikey = st.text_input("InChIKey", value=maybe_blank(row.get("inchikey")))
                    compound_class = select_or_custom(
                        "Compound Class",
                        build_existing_options(compounds_df, "compound_class", DEFAULT_CLASS_OPTIONS),
                        f"edit_compound_class_{edit_compound_id}",
                        value=maybe_blank(row["compound_class"]),
                        help_text="Choose an existing class or use Custom... to add a new compound class.",
                    )
                    compound_subclass = select_or_custom(
                        "Compound Subclass",
                        build_existing_options(compounds_df, "compound_subclass"),
                        f"edit_compound_subclass_{edit_compound_id}",
                        value=maybe_blank(row["compound_subclass"]),
                    )
                    source_category = select_or_custom(
                        "Source Category",
                        build_existing_options(compounds_df, "source_category", DEFAULT_SOURCE_OPTIONS),
                        f"edit_source_category_{edit_compound_id}",
                        value=maybe_blank(row.get("source_category")),
                        help_text="Choose an existing source category or use Custom... to add a new one.",
                    )
                    source_organism = st.text_input(
                        "Source Organism / Species (optional)",
                        value=maybe_blank(row.get("source_organism")),
                    )
                    sample_code = st.text_input("Sample Code", value=maybe_blank(row["sample_code"]))
                    collection_location = st.text_input("Collection Location", value=maybe_blank(row["collection_location"]))
                    gps_coordinates = st.text_input("GPS Coordinates", value=maybe_blank(row["gps_coordinates"]))
                    depth_m_text = st.text_input("Depth (m)", value=maybe_blank(row["depth_m"]))

                with col2:
                    uv_data = st.text_input("UV Data", value=maybe_blank(row["uv_data"]))
                    ftir_data = st.text_input("FTIR Data", value=maybe_blank(row["ftir_data"]))
                    cd_data = st.text_area("Circular Dichroism (CD / ECD)", value=maybe_blank(row.get("cd_data")))
                    optical_rotation = st.text_input("Optical Rotation", value=maybe_blank(row["optical_rotation"]))
                    melting_point = st.text_input("Melting Point", value=maybe_blank(row["melting_point"]))
                    crystallization_method = st.text_input("Crystallization Method", value=maybe_blank(row["crystallization_method"]))
                    structure_image_path = st.text_input("Structure Image Path", value=maybe_blank(row["structure_image_path"]))
                    structure_upload = st.file_uploader(
                        "Replace Structure Image",
                        type=["png", "jpg", "jpeg", "webp"],
                        key=f"edit_structure_upload_{edit_compound_id}",
                    )
                    journal_name = st.text_input("Journal Name", value=maybe_blank(row["journal_name"]))
                    article_title = st.text_area("Article Title", value=maybe_blank(row["article_title"]))
                    publication_year = st.text_input("Publication Year", value=maybe_blank(row["publication_year"]))
                    volume = st.text_input("Volume", value=maybe_blank(row["volume"]))
                    issue = st.text_input("Issue / Journal Number", value=maybe_blank(row["issue"]))
                    pages = st.text_input("Pages", value=maybe_blank(row["pages"]))
                    doi = st.text_input("DOI", value=maybe_blank(row["doi"]))
                    ccdc_number = st.text_input("CCDC", value=maybe_blank(row["ccdc_number"]))
                    molecular_weight_text = st.text_input("Mr", value=maybe_blank(row["molecular_weight"]))
                    hrms_data = st.text_area("HRMS Data", value=maybe_blank(row["hrms_data"]))
                    data_source = select_or_custom(
                        "Data Source",
                        build_existing_options(compounds_df, "data_source", DEFAULT_DATA_SOURCE_OPTIONS),
                        f"edit_data_source_{edit_compound_id}",
                        value=maybe_blank(row["data_source"]),
                    )
                    current_curation_status = normalize_curation_status(row.get("curation_status"), default="curated")
                    curation_status = st.selectbox(
                        "Curation Status",
                        CURATION_STATUS_OPTIONS,
                        index=CURATION_STATUS_OPTIONS.index(current_curation_status),
                        key=f"edit_curation_status_{edit_compound_id}",
                        help="Use imported for bulk-ingested records, reviewed for checked records, and curated for records ready as trusted NPDB entries.",
                    )

                note = st.text_area("Note", value=maybe_blank(row["note"]))
                submitted_edit = st.button("Save Changes", key="edit_compound_submit")

                if submitted_edit:
                    if not trivial_name.strip():
                        st.error("Trivial Name is required.")
                    else:
                        depth_value = None
                        if depth_m_text.strip():
                            try:
                                depth_value = float(depth_m_text.strip())
                            except ValueError:
                                st.error("Depth (m) must be a valid number.")
                                st.stop()

                        molecular_weight_value = None
                        if molecular_weight_text.strip():
                            try:
                                molecular_weight_value = float(molecular_weight_text.strip())
                            except ValueError:
                                st.error("Mr must be a valid number.")
                                st.stop()

                        if structure_upload is not None:
                            structure_image_path = save_uploaded_asset(
                                structure_upload,
                                STRUCTURES_DIR,
                                f"{trivial_name}_{sample_code or edit_compound_id}_structure",
                            )

                        source_category, source_organism, source_material = infer_source_fields(
                            source_category.strip(),
                            source_organism.strip(),
                            row.get("source_material"),
                        )

                        update_compound_record(
                            compound_id=edit_compound_id,
                            trivial_name=trivial_name.strip(),
                            iupac_name=iupac_name.strip(),
                            molecular_formula=molecular_formula.strip(),
                            compound_class=compound_class.strip(),
                            compound_subclass=compound_subclass.strip(),
                            smiles=smiles.strip(),
                            inchi=inchi.strip(),
                            inchikey=inchikey.strip(),
                            source_category=source_category.strip(),
                            source_organism=source_organism.strip(),
                            source_material=source_material.strip(),
                            sample_code=sample_code.strip(),
                            collection_location=collection_location.strip(),
                            gps_coordinates=gps_coordinates.strip(),
                            depth_m=depth_value,
                            uv_data=uv_data.strip(),
                            ftir_data=ftir_data.strip(),
                            cd_data=cd_data.strip(),
                            optical_rotation=optical_rotation.strip(),
                            melting_point=melting_point.strip(),
                            crystallization_method=crystallization_method.strip(),
                            structure_image_path=structure_image_path.strip(),
                            journal_name=journal_name.strip(),
                            article_title=article_title.strip(),
                            publication_year=publication_year.strip(),
                            volume=volume.strip(),
                            issue=issue.strip(),
                            pages=pages.strip(),
                            doi=doi.strip(),
                            ccdc_number=ccdc_number.strip(),
                            molecular_weight=molecular_weight_value,
                            hrms_data=hrms_data.strip(),
                            data_source=data_source.strip(),
                            curation_status=normalize_curation_status(curation_status, default="curated"),
                            note=note.strip()
                        )

                        st.success(f"Record ID {edit_compound_id} updated successfully.")

                        left_btn, right_btn = st.columns([1, 1])
                        with left_btn:
                            if st.button("Open Updated Record", key=f"open_updated_compound_{edit_compound_id}"):
                                open_compound_detail(edit_compound_id)
                                st.rerun()
                        with right_btn:
                            if st.button("Refresh Form", key=f"stay_editor_{edit_compound_id}"):
                                st.rerun()

    else:
        section_header("Delete Record", "Delete a compound together with all related spectral records.")
        compounds_df = load_all_compounds()

        if compounds_df.empty:
            st.info("No compounds available.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            selected_label = st.selectbox("Select record to delete", options["label"].tolist(), key="delete_compound_select")
            compound_id = int(selected_label.split(" - ")[0])

            row_df = load_compound_row(compound_id)
            if not row_df.empty:
                row = row_df.iloc[0]
                proton_count = len(load_proton_data(compound_id))
                carbon_count = len(load_carbon_data(compound_id))
                spectra_count = len(load_spectra_files(compound_id))

                st.warning("This action cannot be undone.")
                c1, c2, c3 = st.columns(3)
                render_metric_card("1H records", proton_count, c1)
                render_metric_card("13C records", carbon_count, c2)
                render_metric_card("Spectra records", spectra_count, c3)

                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                st.write(f"**Compound:** {clean_text(row['trivial_name'])}")
                st.write(f"**Compound ID:** {compound_id}")
                st.markdown('</div>', unsafe_allow_html=True)

                with st.form("delete_compound_form"):
                    confirm = st.checkbox("I understand that this will permanently delete the compound record and all related database records.")
                    submitted_delete = st.form_submit_button("Delete Record")

                if submitted_delete:
                    if not confirm:
                        st.error("Please confirm deletion first.")
                    else:
                        delete_compound_record(compound_id)
                        st.success(f"Compound ID {compound_id} and its related records were deleted.")
                        st.session_state["selected_compound_id"] = None


# =========================
# 1H pages
# =========================
def show_proton_pages():
    if not can_edit_database():
        section_header("1H Peak Browser", "Read-only access to proton assignments. Full edit access remains reserved for the database owner.")
        render_read_only_notice("add, edit, or delete 1H peak records")
        compounds_df = load_all_compounds()
        if compounds_df.empty:
            st.info("No compounds available.")
            return
        options = compounds_df[["id", "trivial_name"]].copy()
        options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
        default_index = 0
        selected_id = st.session_state.get("selected_compound_id")
        if selected_id is not None and selected_id in options["id"].tolist():
            default_index = options.index[options["id"] == selected_id][0]
        selected_compound_label = st.selectbox("Select Compound", options["label"].tolist(), index=default_index, key="readonly_proton_compound")
        compound_id = int(selected_compound_label.split(" - ")[0])
        proton_df = load_proton_data(compound_id)
        if proton_df.empty:
            st.info("No 1H NMR data available for this compound.")
        else:
            proton_df = proton_df.rename(columns={
                "id": "ID",
                "delta_ppm": "δH (ppm)",
                "multiplicity": "Multiplicity",
                "j_value": "J Value",
                "proton_count": "Proton Count",
                "assignment": "Assignment",
                "solvent": "Solvent",
                "instrument_mhz": "Instrument (MHz)",
                "note": "Note",
            })
            st.dataframe(proton_df, width="stretch", hide_index=True)
        return

    proton_page = st.radio(
        "1H Peak Tools",
        ["Add Peak", "Edit Peak", "Delete Peak"],
        horizontal=True
    )

    if proton_page == "Add Peak":
        section_header("Add 1H Peak", "Register a single 1H NMR peak for a selected compound.")
        compounds_df = load_all_compounds()

        if compounds_df.empty:
            st.info("No compounds available. Please add a compound first.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            label_list = options["label"].tolist()

            default_index = 0
            selected_id = st.session_state.get("selected_compound_id")
            if selected_id is not None and selected_id in options["id"].tolist():
                default_index = options.index[options["id"] == selected_id][0]

            selected_compound_label = st.selectbox(
                "Select Compound",
                label_list,
                index=default_index,
                key="add1h_compound"
            )

            selected_compound_id = int(selected_compound_label.split(" - ")[0])

            with st.form("add_1h_form", clear_on_submit=False):
                c1, c2 = st.columns(2)

                with c1:
                    delta_ppm_text = st.text_input("δH (ppm)")
                    multiplicity = st.text_input("Multiplicity")
                    j_value = st.text_input("J Value")
                    proton_count = st.text_input("Proton Count", placeholder="e.g. 1H or 3H")
                    assignment = st.text_input("Assignment")

                with c2:
                    solvent = st.text_input("Solvent", value="CDCl3")
                    instrument_mhz_text = st.text_input("Instrument (MHz)", value="500")
                    note = st.text_area("Note")

                submitted_1h = st.form_submit_button("Save 1H Peak")

            if submitted_1h:
                if not delta_ppm_text.strip():
                    st.error("δH (ppm) is required.")
                elif not assignment.strip():
                    st.error("Assignment is required.")
                else:
                    try:
                        delta_ppm_value = float(delta_ppm_text.strip())
                    except ValueError:
                        st.error("δH (ppm) must be a valid number.")
                        st.stop()

                    instrument_mhz_value = None
                    if instrument_mhz_text.strip():
                        try:
                            instrument_mhz_value = float(instrument_mhz_text.strip())
                        except ValueError:
                            st.error("Instrument (MHz) must be a valid number.")
                            st.stop()

                    new_peak_id = insert_proton_record(
                        compound_id=selected_compound_id,
                        delta_ppm=delta_ppm_value,
                        multiplicity=multiplicity.strip(),
                        j_value=j_value.strip(),
                        proton_count=proton_count.strip(),
                        assignment=assignment.strip(),
                        solvent=solvent.strip(),
                        instrument_mhz=instrument_mhz_value,
                        note=note.strip()
                    )

                    st.success(f"1H NMR peak saved successfully. New Peak ID: {new_peak_id}")

                    if st.button("Open Record", key=f"open_detail_after_1h_{new_peak_id}"):
                        open_compound_detail(selected_compound_id)
                        st.rerun()

    elif proton_page == "Edit Peak":
        section_header("Edit 1H Peak", "Update a single 1H record directly from the web interface.")
        proton_df = load_all_proton_data()

        if proton_df.empty:
            st.info("No 1H NMR records available.")
        else:
            proton_df["label"] = (
                proton_df["id"].astype(str)
                + " | "
                + proton_df["trivial_name"].fillna("-").astype(str)
                + " | δH "
                + proton_df["delta_ppm"].astype(str)
                + " | "
                + proton_df["assignment"].fillna("-").astype(str)
            )

            selected_label = st.selectbox(
                "Select 1H NMR Record",
                proton_df["label"].tolist(),
                key="edit_1h_select"
            )

            proton_id = int(selected_label.split(" | ")[0])
            row_df = load_proton_row(proton_id)

            if row_df.empty:
                st.error("1H NMR record not found.")
            else:
                row = row_df.iloc[0]
                compounds_df = load_all_compounds()
                options = compounds_df[["id", "trivial_name"]].copy()
                options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
                label_list = options["label"].tolist()

                default_index = 0
                if row["compound_id"] in options["id"].tolist():
                    default_index = options.index[options["id"] == row["compound_id"]][0]

                with st.form("edit_1h_form", clear_on_submit=False):
                    selected_compound_label = st.selectbox(
                        "Select Compound",
                        label_list,
                        index=default_index,
                        key="edit1h_compound"
                    )

                    c1, c2 = st.columns(2)

                    with c1:
                        delta_ppm_text = st.text_input("δH (ppm)", value=maybe_blank(row["delta_ppm"]))
                        multiplicity = st.text_input("Multiplicity", value=maybe_blank(row["multiplicity"]))
                        j_value = st.text_input("J Value", value=maybe_blank(row["j_value"]))
                        proton_count = st.text_input("Proton Count", value=maybe_blank(row["proton_count"]))
                        assignment = st.text_input("Assignment", value=maybe_blank(row["assignment"]))

                    with c2:
                        solvent = st.text_input("Solvent", value=maybe_blank(row["solvent"]))
                        instrument_mhz_text = st.text_input("Instrument (MHz)", value=maybe_blank(row["instrument_mhz"]))
                        note = st.text_area("Note", value=maybe_blank(row["note"]))

                    submitted_edit_1h = st.form_submit_button("Save Changes")

                if submitted_edit_1h:
                    if not delta_ppm_text.strip():
                        st.error("δH (ppm) is required.")
                    elif not assignment.strip():
                        st.error("Assignment is required.")
                    else:
                        try:
                            delta_ppm_value = float(delta_ppm_text.strip())
                        except ValueError:
                            st.error("δH (ppm) must be a valid number.")
                            st.stop()

                        instrument_mhz_value = None
                        if instrument_mhz_text.strip():
                            try:
                                instrument_mhz_value = float(instrument_mhz_text.strip())
                            except ValueError:
                                st.error("Instrument (MHz) must be a valid number.")
                                st.stop()

                        selected_compound_id = int(selected_compound_label.split(" - ")[0])

                        update_proton_record(
                            proton_id=proton_id,
                            compound_id=selected_compound_id,
                            delta_ppm=delta_ppm_value,
                            multiplicity=multiplicity.strip(),
                            j_value=j_value.strip(),
                            proton_count=proton_count.strip(),
                            assignment=assignment.strip(),
                            solvent=solvent.strip(),
                            instrument_mhz=instrument_mhz_value,
                            note=note.strip()
                        )

                        st.success(f"1H NMR record ID {proton_id} updated successfully.")

                        left_btn, right_btn = st.columns([1, 1])
                        with left_btn:
                            if st.button("Open Record", key=f"open_detail_after_edit_1h_{proton_id}"):
                                open_compound_detail(selected_compound_id)
                                st.rerun()
                        with right_btn:
                            if st.button("Refresh Form", key=f"reload_edit_1h_{proton_id}"):
                                st.rerun()

    else:
        section_header("Delete 1H Peak", "Remove a single 1H NMR record.")
        proton_df = load_all_proton_data()

        if proton_df.empty:
            st.info("No 1H NMR records available.")
        else:
            proton_df["label"] = (
                proton_df["id"].astype(str)
                + " | "
                + proton_df["trivial_name"].fillna("-").astype(str)
                + " | δH "
                + proton_df["delta_ppm"].astype(str)
                + " | "
                + proton_df["assignment"].fillna("-").astype(str)
            )

            selected_label = st.selectbox("Select 1H NMR Record to Delete", proton_df["label"].tolist(), key="delete_1h_select")
            proton_id = int(selected_label.split(" | ")[0])
            row_df = load_proton_row(proton_id)

            if not row_df.empty:
                row = row_df.iloc[0]
                st.warning("This action cannot be undone.")
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                st.write(f"**Record ID:** {proton_id}")
                st.write(f"**Compound ID:** {row['compound_id']}")
                st.write(f"**δH (ppm):** {clean_text(row['delta_ppm'])}")
                st.write(f"**Assignment:** {clean_text(row['assignment'])}")
                st.markdown('</div>', unsafe_allow_html=True)

                with st.form("delete_1h_form"):
                    confirm = st.checkbox("I understand that this will permanently delete this 1H NMR record.")
                    submitted_delete = st.form_submit_button("Delete 1H Record")

                if submitted_delete:
                    if not confirm:
                        st.error("Please confirm deletion first.")
                    else:
                        compound_id = int(row["compound_id"])
                        delete_proton_record_by_id(proton_id)
                        st.success(f"1H NMR record ID {proton_id} was deleted.")
                        if st.button("Open Record", key=f"open_detail_after_delete_1h_{proton_id}"):
                            open_compound_detail(compound_id)
                            st.rerun()


# =========================
# 13C pages
# =========================
def show_carbon_pages():
    if not can_edit_database():
        section_header("13C Peak Browser", "Read-only access to carbon assignments. Full edit access remains reserved for the database owner.")
        render_read_only_notice("add, edit, or delete 13C peak records")
        compounds_df = load_all_compounds()
        if compounds_df.empty:
            st.info("No compounds available.")
            return
        options = compounds_df[["id", "trivial_name"]].copy()
        options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
        default_index = 0
        selected_id = st.session_state.get("selected_compound_id")
        if selected_id is not None and selected_id in options["id"].tolist():
            default_index = options.index[options["id"] == selected_id][0]
        selected_compound_label = st.selectbox("Select Compound", options["label"].tolist(), index=default_index, key="readonly_carbon_compound")
        compound_id = int(selected_compound_label.split(" - ")[0])
        carbon_df = load_carbon_data(compound_id)
        if carbon_df.empty:
            st.info("No 13C NMR data available for this compound.")
        else:
            carbon_df = carbon_df.rename(columns={
                "id": "ID",
                "delta_ppm": "δC (ppm)",
                "carbon_type": "Carbon Type",
                "assignment": "Assignment",
                "solvent": "Solvent",
                "instrument_mhz": "Instrument (MHz)",
                "note": "Note",
            })
            st.dataframe(carbon_df, width="stretch", hide_index=True)
        return

    carbon_page = st.radio(
        "13C Peak Tools",
        ["Add Peak", "Edit Peak", "Delete Peak"],
        horizontal=True
    )

    if carbon_page == "Add Peak":
        section_header("Add 13C Peak", "Register a single 13C NMR peak for a selected compound.")
        compounds_df = load_all_compounds()

        if compounds_df.empty:
            st.info("No compounds available. Please add a compound first.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            label_list = options["label"].tolist()

            default_index = 0
            selected_id = st.session_state.get("selected_compound_id")
            if selected_id is not None and selected_id in options["id"].tolist():
                default_index = options.index[options["id"] == selected_id][0]

            selected_compound_label = st.selectbox(
                "Select Compound",
                label_list,
                index=default_index,
                key="add13c_compound"
            )

            selected_compound_id = int(selected_compound_label.split(" - ")[0])

            with st.form("add_13c_form", clear_on_submit=False):
                c1, c2 = st.columns(2)

                with c1:
                    delta_ppm_text = st.text_input("δC (ppm)")
                    carbon_type = st.text_input("Carbon Type", placeholder="e.g. CH3, CH2, CH, C")
                    assignment = st.text_input("Assignment")

                with c2:
                    solvent = st.text_input("Solvent", value="CDCl3", key="add13c_solvent")
                    instrument_mhz_text = st.text_input("Instrument (MHz)", value="125")
                    note = st.text_area("Note", key="add13c_note")

                submitted_13c = st.form_submit_button("Save 13C Peak")

            if submitted_13c:
                if not delta_ppm_text.strip():
                    st.error("δC (ppm) is required.")
                elif not assignment.strip():
                    st.error("Assignment is required.")
                else:
                    try:
                        delta_ppm_value = float(delta_ppm_text.strip())
                    except ValueError:
                        st.error("δC (ppm) must be a valid number.")
                        st.stop()

                    instrument_mhz_value = None
                    if instrument_mhz_text.strip():
                        try:
                            instrument_mhz_value = float(instrument_mhz_text.strip())
                        except ValueError:
                            st.error("Instrument (MHz) must be a valid number.")
                            st.stop()

                    new_peak_id = insert_carbon_record(
                        compound_id=selected_compound_id,
                        delta_ppm=delta_ppm_value,
                        carbon_type=carbon_type.strip(),
                        assignment=assignment.strip(),
                        solvent=solvent.strip(),
                        instrument_mhz=instrument_mhz_value,
                        note=note.strip()
                    )

                    st.success(f"13C NMR peak saved successfully. New Peak ID: {new_peak_id}")

                    if st.button("Open Record", key=f"open_detail_after_13c_{new_peak_id}"):
                        open_compound_detail(selected_compound_id)
                        st.rerun()

    elif carbon_page == "Edit Peak":
        section_header("Edit 13C Peak", "Update a single 13C record directly from the web interface.")
        carbon_df = load_all_carbon_data()

        if carbon_df.empty:
            st.info("No 13C NMR records available.")
        else:
            carbon_df["label"] = (
                carbon_df["id"].astype(str)
                + " | "
                + carbon_df["trivial_name"].fillna("-").astype(str)
                + " | δC "
                + carbon_df["delta_ppm"].astype(str)
                + " | "
                + carbon_df["assignment"].fillna("-").astype(str)
            )

            selected_label = st.selectbox(
                "Select 13C NMR Record",
                carbon_df["label"].tolist(),
                key="edit_13c_select"
            )

            carbon_id = int(selected_label.split(" | ")[0])
            row_df = load_carbon_row(carbon_id)

            if row_df.empty:
                st.error("13C NMR record not found.")
            else:
                row = row_df.iloc[0]
                compounds_df = load_all_compounds()
                options = compounds_df[["id", "trivial_name"]].copy()
                options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
                label_list = options["label"].tolist()

                default_index = 0
                if row["compound_id"] in options["id"].tolist():
                    default_index = options.index[options["id"] == row["compound_id"]][0]

                with st.form("edit_13c_form", clear_on_submit=False):
                    selected_compound_label = st.selectbox(
                        "Select Compound",
                        label_list,
                        index=default_index,
                        key="edit13c_compound"
                    )

                    c1, c2 = st.columns(2)

                    with c1:
                        delta_ppm_text = st.text_input("δC (ppm)", value=maybe_blank(row["delta_ppm"]))
                        carbon_type = st.text_input("Carbon Type", value=maybe_blank(row["carbon_type"]))
                        assignment = st.text_input("Assignment", value=maybe_blank(row["assignment"]))

                    with c2:
                        solvent = st.text_input("Solvent", value=maybe_blank(row["solvent"]))
                        instrument_mhz_text = st.text_input("Instrument (MHz)", value=maybe_blank(row["instrument_mhz"]))
                        note = st.text_area("Note", value=maybe_blank(row["note"]))

                    submitted_edit_13c = st.form_submit_button("Save Changes")

                if submitted_edit_13c:
                    if not delta_ppm_text.strip():
                        st.error("δC (ppm) is required.")
                    elif not assignment.strip():
                        st.error("Assignment is required.")
                    else:
                        try:
                            delta_ppm_value = float(delta_ppm_text.strip())
                        except ValueError:
                            st.error("δC (ppm) must be a valid number.")
                            st.stop()

                        instrument_mhz_value = None
                        if instrument_mhz_text.strip():
                            try:
                                instrument_mhz_value = float(instrument_mhz_text.strip())
                            except ValueError:
                                st.error("Instrument (MHz) must be a valid number.")
                                st.stop()

                        selected_compound_id = int(selected_compound_label.split(" - ")[0])

                        update_carbon_record(
                            carbon_id=carbon_id,
                            compound_id=selected_compound_id,
                            delta_ppm=delta_ppm_value,
                            carbon_type=carbon_type.strip(),
                            assignment=assignment.strip(),
                            solvent=solvent.strip(),
                            instrument_mhz=instrument_mhz_value,
                            note=note.strip()
                        )

                        st.success(f"13C NMR record ID {carbon_id} updated successfully.")

                        left_btn, right_btn = st.columns([1, 1])
                        with left_btn:
                            if st.button("Open Record", key=f"open_detail_after_edit_13c_{carbon_id}"):
                                open_compound_detail(selected_compound_id)
                                st.rerun()
                        with right_btn:
                            if st.button("Refresh Form", key=f"reload_edit_13c_{carbon_id}"):
                                st.rerun()

    else:
        section_header("Delete 13C Peak", "Remove a single 13C NMR record.")
        carbon_df = load_all_carbon_data()

        if carbon_df.empty:
            st.info("No 13C NMR records available.")
        else:
            carbon_df["label"] = (
                carbon_df["id"].astype(str)
                + " | "
                + carbon_df["trivial_name"].fillna("-").astype(str)
                + " | δC "
                + carbon_df["delta_ppm"].astype(str)
                + " | "
                + carbon_df["assignment"].fillna("-").astype(str)
            )

            selected_label = st.selectbox("Select 13C NMR Record to Delete", carbon_df["label"].tolist(), key="delete_13c_select")
            carbon_id = int(selected_label.split(" | ")[0])
            row_df = load_carbon_row(carbon_id)

            if not row_df.empty:
                row = row_df.iloc[0]
                st.warning("This action cannot be undone.")
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                st.write(f"**Record ID:** {carbon_id}")
                st.write(f"**Compound ID:** {row['compound_id']}")
                st.write(f"**δC (ppm):** {clean_text(row['delta_ppm'])}")
                st.write(f"**Assignment:** {clean_text(row['assignment'])}")
                st.markdown('</div>', unsafe_allow_html=True)

                with st.form("delete_13c_form"):
                    confirm = st.checkbox("I understand that this will permanently delete this 13C NMR record.")
                    submitted_delete = st.form_submit_button("Delete 13C Record")

                if submitted_delete:
                    if not confirm:
                        st.error("Please confirm deletion first.")
                    else:
                        compound_id = int(row["compound_id"])
                        delete_carbon_record_by_id(carbon_id)
                        st.success(f"13C NMR record ID {carbon_id} was deleted.")
                        if st.button("Open Record", key=f"open_detail_after_delete_13c_{carbon_id}"):
                            open_compound_detail(compound_id)
                            st.rerun()


# =========================
# Bioactivity pages
# =========================
def show_bioactivity_pages():
    bioactivity_options = ["Browse Assays", "Add Assay", "Edit Assay", "Delete Assay"] if can_edit_database() else ["Browse Assays"]
    bioactivity_page = st.radio(
        "Bioactivity Tools",
        bioactivity_options,
        horizontal=True,
    )

    bioactivity_df = load_all_bioactivity_data()
    compounds_df = load_all_compounds()
    if not can_edit_database():
        render_read_only_notice("add, edit, or delete bioactivity records")

    if bioactivity_page == "Browse Assays":
        section_header(
            "Bioactivity Browser",
            "Review assay records by activity class, target, potency metric, and linked compound.",
        )
        if bioactivity_df.empty:
            st.info("No bioactivity records available yet.")
            return

        with st.expander("Bioactivity Filters", expanded=True):
            activity_filter = st.selectbox(
                "Activity",
                ["All"] + sorted(set(bioactivity_df["activity_label"].fillna("").astype(str).str.strip()) - {""}),
                key="bioactivity_activity_filter",
            )
            target_category_filter = st.selectbox(
                "Target Category",
                ["All"] + sorted(set(bioactivity_df["target_category"].fillna("").astype(str).str.strip()) - {""}),
                key="bioactivity_target_filter",
            )
            potency_filter = st.selectbox(
                "Potency Metric",
                ["All"] + sorted(set(bioactivity_df["potency_type"].fillna("").astype(str).str.strip()) - {""}),
                key="bioactivity_metric_filter",
            )
            keyword_filter = st.text_input(
                "Keyword",
                key="bioactivity_keyword_filter",
                placeholder="target, organism, compound, assay source...",
            )

        filtered_df = bioactivity_df.copy()
        if activity_filter != "All":
            filtered_df = filtered_df[filtered_df["activity_label"].fillna("").astype(str).str.strip() == activity_filter]
        if target_category_filter != "All":
            filtered_df = filtered_df[filtered_df["target_category"].fillna("").astype(str).str.strip() == target_category_filter]
        if potency_filter != "All":
            filtered_df = filtered_df[filtered_df["potency_type"].fillna("").astype(str).str.strip() == potency_filter]
        if keyword_filter.strip():
            keyword = keyword_filter.strip().lower()
            searchable = filtered_df[
                [
                    "trivial_name",
                    "activity_label",
                    "target_name",
                    "target_category",
                    "assay_type",
                    "assay_source",
                    "note",
                ]
            ].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            filtered_df = filtered_df[searchable.str.contains(re.escape(keyword), regex=True)]

        top1, top2, top3 = st.columns(3)
        render_metric_card("Assay Records", len(filtered_df), top1)
        render_metric_card("Linked Compounds", filtered_df["compound_id"].nunique(), top2)
        active_hits = filtered_df[filtered_df["outcome"].fillna("").astype(str).str.lower().str.contains("active|potent|strong")]
        render_metric_card("Marked Active/Potent", len(active_hits), top3)

        export_df = export_bioactivity_results(filtered_df)
        download_dataframe_button(
            label="Download Bioactivity Browser as Excel",
            df=export_df,
            file_name="bioactivity_browser.xlsx",
            key="download_bioactivity_browser",
            sheet_name="Bioactivity Browser",
        )
        st.dataframe(export_df, width="stretch", hide_index=True)

        section_header("Highlighted Assays")
        for _, row in filtered_df.head(8).iterrows():
            st.markdown('<div class="panel-card">', unsafe_allow_html=True)
            left, right = st.columns([4.5, 1])
            with left:
                st.markdown(f"**{clean_text(row['trivial_name'])}**")
                st.caption(
                    f"{clean_text(row['activity_label'])} | {clean_text(row['target_name'])} | "
                    f"{clean_text(row['potency_type'])} {clean_text(row['potency_relation'])} "
                    f"{clean_text(row['potency_value'])} {clean_text(row['potency_unit'])}"
                )
                st.write(clean_text(row["note"]))
            with right:
                if st.button("Open Record", key=f"bioactivity_open_{row['id']}"):
                    open_compound_detail(int(row["compound_id"]))
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    elif bioactivity_page == "Add Assay":
        section_header(
            "Add Bioactivity Record",
            "Capture reported assay outcomes from marine natural product papers in a flexible but structured format.",
        )
        render_helper_card(
            "Suggested data model",
            "Use Activity for the broad phenotype, Target for the exact cell line / microbe / enzyme, Metric for IC50 or MIC, and Outcome/Note for context such as selectivity or partial inhibition.",
        )

        if compounds_df.empty:
            st.info("No compounds available. Please add a compound first.")
            return

        options = compounds_df[["id", "trivial_name"]].copy()
        options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
        label_list = options["label"].tolist()
        default_index = 0
        selected_id = st.session_state.get("selected_compound_id")
        if selected_id is not None and selected_id in options["id"].tolist():
            default_index = options.index[options["id"] == selected_id][0]


        selected_compound_label = st.selectbox("Select Compound", label_list, index=default_index, key="add_bioactivity_compound")
        c1, c2 = st.columns(2)
        with c1:
            activity_label = select_or_custom(
                "Activity",
                build_existing_options(bioactivity_df, "activity_label", DEFAULT_BIOACTIVITY_CATEGORIES),
                "add_bioactivity_activity",
                help_text="Choose an existing broad activity label or use Custom... for a new one.",
            )
            target_name = st.text_input("Target Name", placeholder="e.g. HCT-116, MRSA, PTP1B")
            target_category = select_or_custom(
                "Target Category",
                build_existing_options(bioactivity_df, "target_category", DEFAULT_TARGET_CATEGORIES),
                "add_bioactivity_target_category",
            )
            assay_type = st.text_input("Assay Type", placeholder="e.g. cytotoxicity assay, antimicrobial assay")
            potency_type = select_or_custom(
                "Potency Metric",
                build_existing_options(bioactivity_df, "potency_type", DEFAULT_POTENCY_TYPES),
                "add_bioactivity_metric",
            )
            potency_relation = st.selectbox("Relation", ["=", "<", "<=", ">", ">=", "~"], index=0)
            potency_value_text = st.text_input("Potency Value", placeholder="e.g. 1.2")
        with c2:
            potency_unit = select_or_custom(
                "Potency Unit",
                build_existing_options(bioactivity_df, "potency_unit", DEFAULT_POTENCY_UNITS),
                "add_bioactivity_unit",
            )
            outcome = st.text_input("Outcome", placeholder="e.g. active, inactive, moderate, selective")
            assay_medium = st.text_input("Assay Medium / Test System", placeholder="e.g. in vitro, broth microdilution")
            selectivity = st.text_input("Selectivity", placeholder="e.g. selective vs normal Vero cells")
            assay_source = st.text_input("Assay Source", placeholder="e.g. J. Am. Chem. Soc. 2006")
            note = st.text_area("Note", placeholder="Any caveat, mechanism note, replicate information, or assay context")
        submitted = st.button("Save Bioactivity Record", key="add_bioactivity_submit")

        if submitted:
            if not activity_label.strip():
                st.error("Activity is required.")
            else:
                potency_value = safe_float_or_none(potency_value_text)
                if potency_value_text.strip() and potency_value is None:
                    st.error("Potency Value must be a valid number.")
                else:
                    selected_compound_id = int(selected_compound_label.split(" - ")[0])
                    new_id = insert_bioactivity_record(
                        compound_id=selected_compound_id,
                        activity_label=activity_label.strip(),
                        target_name=target_name.strip(),
                        target_category=target_category.strip(),
                        assay_type=assay_type.strip(),
                        potency_type=potency_type.strip(),
                        potency_relation=potency_relation.strip(),
                        potency_value=potency_value,
                        potency_unit=potency_unit.strip(),
                        outcome=outcome.strip(),
                        assay_medium=assay_medium.strip(),
                        selectivity=selectivity.strip(),
                        assay_source=assay_source.strip(),
                        note=note.strip(),
                    )
                    st.success(f"Bioactivity record saved successfully. New Assay ID: {new_id}")

    elif bioactivity_page == "Edit Assay":
        section_header("Edit Bioactivity Record", "Update an existing assay entry without touching the parent compound metadata.")
        if bioactivity_df.empty:
            st.info("No bioactivity records available.")
            return

        bioactivity_df["label"] = (
            bioactivity_df["id"].astype(str)
            + " | "
            + bioactivity_df["trivial_name"].fillna("-").astype(str)
            + " | "
            + bioactivity_df["activity_label"].fillna("-").astype(str)
            + " | "
            + bioactivity_df["target_name"].fillna("-").astype(str)
        )
        selected_label = st.selectbox("Select Bioactivity Record", bioactivity_df["label"].tolist(), key="edit_bioactivity_select")
        bioactivity_id = int(selected_label.split(" | ")[0])
        row_df = load_bioactivity_row(bioactivity_id)

        if row_df.empty:
            st.error("Bioactivity record not found.")
            return

        row = row_df.iloc[0]
        options = compounds_df[["id", "trivial_name"]].copy()
        options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
        label_list = options["label"].tolist()
        default_index = 0
        if row["compound_id"] in options["id"].tolist():
            default_index = options.index[options["id"] == row["compound_id"]][0]


        selected_compound_label = st.selectbox("Select Compound", label_list, index=default_index, key="edit_bioactivity_compound")
        c1, c2 = st.columns(2)
        with c1:
            activity_label = select_or_custom(
                "Activity",
                build_existing_options(bioactivity_df, "activity_label", DEFAULT_BIOACTIVITY_CATEGORIES),
                f"edit_bioactivity_activity_{bioactivity_id}",
                value=maybe_blank(row["activity_label"]),
            )
            target_name = st.text_input("Target Name", value=maybe_blank(row["target_name"]))
            target_category = select_or_custom(
                "Target Category",
                build_existing_options(bioactivity_df, "target_category", DEFAULT_TARGET_CATEGORIES),
                f"edit_bioactivity_target_category_{bioactivity_id}",
                value=maybe_blank(row["target_category"]),
            )
            assay_type = st.text_input("Assay Type", value=maybe_blank(row["assay_type"]))
            potency_type = select_or_custom(
                "Potency Metric",
                build_existing_options(bioactivity_df, "potency_type", DEFAULT_POTENCY_TYPES),
                f"edit_bioactivity_metric_{bioactivity_id}",
                value=maybe_blank(row["potency_type"]),
            )
            relation_options = ["=", "<", "<=", ">", ">=", "~"]
            relation_value = maybe_blank(row["potency_relation"]) or "="
            potency_relation = st.selectbox("Relation", relation_options, index=relation_options.index(relation_value) if relation_value in relation_options else 0)
            potency_value_text = st.text_input("Potency Value", value=maybe_blank(row["potency_value"]))
        with c2:
            potency_unit = select_or_custom(
                "Potency Unit",
                build_existing_options(bioactivity_df, "potency_unit", DEFAULT_POTENCY_UNITS),
                f"edit_bioactivity_unit_{bioactivity_id}",
                value=maybe_blank(row["potency_unit"]),
            )
            outcome = st.text_input("Outcome", value=maybe_blank(row["outcome"]))
            assay_medium = st.text_input("Assay Medium / Test System", value=maybe_blank(row["assay_medium"]))
            selectivity = st.text_input("Selectivity", value=maybe_blank(row["selectivity"]))
            assay_source = st.text_input("Assay Source", value=maybe_blank(row["assay_source"]))
            note = st.text_area("Note", value=maybe_blank(row["note"]))
        submitted = st.button("Save Changes", key="edit_bioactivity_submit")

        if submitted:
            if not activity_label.strip():
                st.error("Activity is required.")
            else:
                potency_value = safe_float_or_none(potency_value_text)
                if potency_value_text.strip() and potency_value is None:
                    st.error("Potency Value must be a valid number.")
                else:
                    selected_compound_id = int(selected_compound_label.split(" - ")[0])
                    update_bioactivity_record(
                        bioactivity_id=bioactivity_id,
                        compound_id=selected_compound_id,
                        activity_label=activity_label.strip(),
                        target_name=target_name.strip(),
                        target_category=target_category.strip(),
                        assay_type=assay_type.strip(),
                        potency_type=potency_type.strip(),
                        potency_relation=potency_relation.strip(),
                        potency_value=potency_value,
                        potency_unit=potency_unit.strip(),
                        outcome=outcome.strip(),
                        assay_medium=assay_medium.strip(),
                        selectivity=selectivity.strip(),
                        assay_source=assay_source.strip(),
                        note=note.strip(),
                    )
                    st.success(f"Bioactivity record ID {bioactivity_id} updated successfully.")

    else:
        section_header("Delete Bioactivity Record", "Remove one assay record without deleting the parent compound.")
        if bioactivity_df.empty:
            st.info("No bioactivity records available.")
            return
        bioactivity_df["label"] = (
            bioactivity_df["id"].astype(str)
            + " | "
            + bioactivity_df["trivial_name"].fillna("-").astype(str)
            + " | "
            + bioactivity_df["activity_label"].fillna("-").astype(str)
            + " | "
            + bioactivity_df["target_name"].fillna("-").astype(str)
        )
        selected_label = st.selectbox("Select Bioactivity Record to Delete", bioactivity_df["label"].tolist(), key="delete_bioactivity_select")
        bioactivity_id = int(selected_label.split(" | ")[0])
        row_df = load_bioactivity_row(bioactivity_id)
        if not row_df.empty:
            row = row_df.iloc[0]
            st.warning("This action cannot be undone.")
            st.markdown('<div class="panel-card">', unsafe_allow_html=True)
            st.write(f"**Assay ID:** {bioactivity_id}")
            st.write(f"**Activity:** {clean_text(row['activity_label'])}")
            st.write(f"**Target:** {clean_text(row['target_name'])}")
            st.write(f"**Metric:** {clean_text(row['potency_type'])} {clean_text(row['potency_relation'])} {clean_text(row['potency_value'])} {clean_text(row['potency_unit'])}")
            st.markdown('</div>', unsafe_allow_html=True)

            with st.form("delete_bioactivity_form"):
                confirm = st.checkbox("I understand that this will permanently delete this bioactivity record.")
                submitted_delete = st.form_submit_button("Delete Bioactivity Record")

            if submitted_delete:
                if not confirm:
                    st.error("Please confirm deletion first.")
                else:
                    delete_bioactivity_record_by_id(bioactivity_id)
                    st.success(f"Bioactivity record ID {bioactivity_id} was deleted.")


# =========================
# Spectra pages
# =========================
def show_spectra_library_overview():
    spectra_df = load_all_spectra_files()
    section_header(
        "Spectra Library Overview",
        "Review coverage, storage quality, and quick-access previews before editing individual file records.",
    )
    if spectra_df.empty:
        st.info("No spectra file records available.")
        return

    spectra_df = spectra_df.copy()
    spectra_df["storage_type"] = spectra_df["file_path"].fillna("").astype(str).apply(classify_storage_type)
    spectra_df["is_remote"] = spectra_df["file_path"].fillna("").astype(str).apply(is_external_url)
    spectra_df["exists_locally"] = spectra_df["file_path"].fillna("").astype(str).apply(
        lambda value: True if is_external_url(value) else bool(get_full_file_path(value) and get_full_file_path(value).exists())
    )

    m1, m2, m3, m4 = st.columns(4)
    render_metric_card("Library Records", len(spectra_df), m1)
    render_metric_card("Remote Links", int(spectra_df["is_remote"].sum()), m2)
    render_metric_card("Local Existing Files", int((~spectra_df["is_remote"] & spectra_df["exists_locally"]).sum()), m3)
    render_metric_card("Missing Local Files", int((~spectra_df["is_remote"] & ~spectra_df["exists_locally"]).sum()), m4)

    spectrum_counts = (
        spectra_df["spectrum_type"]
        .fillna("Uncategorized")
        .replace("", "Uncategorized")
        .value_counts()
        .reset_index()
    )
    spectrum_counts.columns = ["Spectrum Type", "Count"]
    render_dashboard_pie_chart(
        spectrum_counts,
        names_col="Spectrum Type",
        values_col="Count",
        color_sequence=["#FF7F6D", "#61D8ED", "#4C8EFF", "#9C63F1", "#F2C66D", "#7EF0C2"],
    )

    st.markdown("**Recent Library Entries**")
    preview_df = spectra_df[
        ["id", "trivial_name", "spectrum_type", "storage_type", "file_path", "note"]
    ].rename(
        columns={
            "id": "ID",
            "trivial_name": "Compound",
            "spectrum_type": "Spectrum Type",
            "storage_type": "Storage",
            "file_path": "Path",
            "note": "Note",
        }
    )
    st.dataframe(preview_df.head(20), width="stretch", hide_index=True)


def show_spectra_pages():
    if not can_edit_database():
        render_read_only_notice("add, edit, or delete spectra file records")
        show_spectra_library_overview()
        return

    spectra_page = st.radio(
        "Spectra Tools",
        ["Library Overview", "Add Files", "Edit Files", "Delete Files"],
        horizontal=True
    )

    if spectra_page == "Library Overview":
        show_spectra_library_overview()

    elif spectra_page == "Add Files":
        section_header("Add Spectra Files", "Upload files directly or register an existing file path for a selected compound.")
        render_helper_card(
            "Tip",
            "Use Google Drive for large raw-data files and keep local uploads for lighter preview images only.",
        )
        compounds_df = load_all_compounds()
        spectra_df = load_all_spectra_files()

        if compounds_df.empty:
            st.info("No compounds available. Please add a compound first.")
        else:
            options = compounds_df[["id", "trivial_name"]].copy()
            options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
            label_list = options["label"].tolist()

            default_index = 0
            selected_id = st.session_state.get("selected_compound_id")
            if selected_id is not None and selected_id in options["id"].tolist():
                default_index = options.index[options["id"] == selected_id][0]

            selected_compound_label = st.selectbox(
                "Select Compound",
                label_list,
                index=default_index,
                key="add_spectra_compound"
            )

            selected_compound_id = int(selected_compound_label.split(" - ")[0])


            spectrum_type = select_or_custom(
                "Spectrum Type",
                build_existing_options(spectra_df, "spectrum_type", DEFAULT_SPECTRUM_TYPES),
                "add_spectrum_type",
                value="Supporting Data",
            )
            file_path = st.text_input("File Path or External URL (optional if uploading)", placeholder="e.g. data/spectra/RU207-C1_1H.png or https://drive.google.com/...")
            uploaded_files = st.file_uploader(
                "Upload Spectra Files",
                accept_multiple_files=True,
                key="add_spectra_uploads",
            )
            note = st.text_area("Note", key="add_spectra_note")
            st.caption("Recommended: raw data types such as 1H Raw Data, 13C Raw Data, JCAMP-DX, and MNova should use Google Drive links.")

            submitted_spectra = st.button("Save Spectra File", key="add_spectra_submit")

            if submitted_spectra:
                if not spectrum_type.strip():
                    st.error("Spectrum Type is required.")
                elif not file_path.strip() and not uploaded_files:
                    st.error("Provide at least one uploaded file or a file path.")
                else:
                    created_records = []

                    if file_path.strip():
                        validation_errors, validation_warnings = validate_spectrum_entry(file_path.strip(), spectrum_type.strip())
                        for warning_message in validation_warnings:
                            st.warning(warning_message)
                        if validation_errors:
                            for error_message in validation_errors:
                                st.error(error_message)
                            st.stop()

                        created_records.append(
                            insert_spectrum_file_record(
                                compound_id=selected_compound_id,
                                spectrum_type=spectrum_type.strip(),
                                file_path=file_path.strip(),
                                note=note.strip()
                            )
                        )

                    for uploaded_file in uploaded_files or []:
                        saved_path = save_uploaded_asset(
                            uploaded_file,
                            SPECTRA_DIR,
                            f"compound_{selected_compound_id}_{spectrum_type}_{uploaded_file.name}",
                        )
                        created_records.append(
                            insert_spectrum_file_record(
                                compound_id=selected_compound_id,
                                spectrum_type=spectrum_type.strip(),
                                file_path=saved_path,
                                note=note.strip()
                            )
                        )

                    st.success(f"Saved {len(created_records)} spectra file record(s).")

                    if st.button("Open Record", key=f"open_detail_after_spectra_{selected_compound_id}"):
                        open_compound_detail(selected_compound_id)
                        st.rerun()

    elif spectra_page == "Edit Files":
        section_header("Edit Spectra Files", "Update a spectra file record and verify its path.")
        render_helper_card(
            "Tip",
            "You can switch a local path to a Google Drive link at any time, especially for large raw files before public deployment.",
        )
        spectra_df = load_all_spectra_files()

        if spectra_df.empty:
            st.info("No spectra file records available.")
        else:
            spectra_df["label"] = (
                spectra_df["id"].astype(str)
                + " | "
                + spectra_df["trivial_name"].fillna("-").astype(str)
                + " | "
                + spectra_df["spectrum_type"].fillna("-").astype(str)
                + " | "
                + spectra_df["file_path"].fillna("-").astype(str)
            )

            selected_label = st.selectbox(
                "Select Spectra File Record",
                spectra_df["label"].tolist(),
                key="edit_spectra_select"
            )

            file_id = int(selected_label.split(" | ")[0])
            row_df = load_spectrum_file_row(file_id)

            if row_df.empty:
                st.error("Spectra file record not found.")
            else:
                row = row_df.iloc[0]
                compounds_df = load_all_compounds()
                options = compounds_df[["id", "trivial_name"]].copy()
                options["label"] = options["id"].astype(str) + " - " + options["trivial_name"].fillna("")
                label_list = options["label"].tolist()

                default_index = 0
                if row["compound_id"] in options["id"].tolist():
                    default_index = options.index[options["id"] == row["compound_id"]][0]


                selected_compound_label = st.selectbox(
                    "Select Compound",
                    label_list,
                    index=default_index,
                    key="edit_spectra_compound"
                )

                spectrum_type = select_or_custom(
                    "Spectrum Type",
                    build_existing_options(spectra_df, "spectrum_type", DEFAULT_SPECTRUM_TYPES),
                    f"edit_spectrum_type_{file_id}",
                    value=maybe_blank(row["spectrum_type"]),
                )
                file_path = st.text_input("File Path or External URL", value=maybe_blank(row["file_path"]))
                replacement_upload = st.file_uploader(
                    "Replace File by Upload",
                    key=f"edit_spectrum_upload_{file_id}",
                )
                note = st.text_area("Note", value=maybe_blank(row["note"]))

                submitted_edit_spectra = st.button("Save Changes", key="edit_spectra_submit")

                if submitted_edit_spectra:
                    if not spectrum_type.strip():
                        st.error("Spectrum Type is required.")
                    elif not file_path.strip() and replacement_upload is None:
                        st.error("File Path is required.")
                    else:
                        selected_compound_id = int(selected_compound_label.split(" - ")[0])

                        if replacement_upload is not None:
                            file_path = save_uploaded_asset(
                                replacement_upload,
                                SPECTRA_DIR,
                                f"compound_{selected_compound_id}_{spectrum_type}_{replacement_upload.name}",
                            )

                        validation_errors, validation_warnings = validate_spectrum_entry(file_path.strip(), spectrum_type.strip())
                        for warning_message in validation_warnings:
                            st.warning(warning_message)
                        if validation_errors:
                            for error_message in validation_errors:
                                st.error(error_message)
                            st.stop()

                        update_spectrum_file_record(
                            file_id=file_id,
                            compound_id=selected_compound_id,
                            spectrum_type=spectrum_type.strip(),
                            file_path=file_path.strip(),
                            note=note.strip()
                        )

                        st.success(f"Spectra file record ID {file_id} updated successfully.")

                        left_btn, right_btn = st.columns([1, 1])
                        with left_btn:
                            if st.button("Open Record", key=f"open_detail_after_edit_spectra_{file_id}"):
                                open_compound_detail(selected_compound_id)
                                st.rerun()
                        with right_btn:
                            if st.button("Refresh Form", key=f"reload_edit_spectra_{file_id}"):
                                st.rerun()

    else:
        section_header("Delete Spectra Files", "Remove a spectra file record from the database.")
        spectra_df = load_all_spectra_files()

        if spectra_df.empty:
            st.info("No spectra file records available.")
        else:
            spectra_df["label"] = (
                spectra_df["id"].astype(str)
                + " | "
                + spectra_df["trivial_name"].fillna("-").astype(str)
                + " | "
                + spectra_df["spectrum_type"].fillna("-").astype(str)
                + " | "
                + spectra_df["file_path"].fillna("-").astype(str)
            )

            selected_label = st.selectbox("Select Spectra File Record to Delete", spectra_df["label"].tolist(), key="delete_spectra_select")
            file_id = int(selected_label.split(" | ")[0])
            row_df = load_spectrum_file_row(file_id)

            if not row_df.empty:
                row = row_df.iloc[0]
                st.warning("This action cannot be undone.")
                st.markdown('<div class="panel-card">', unsafe_allow_html=True)
                st.write(f"**Record ID:** {file_id}")
                st.write(f"**Compound ID:** {row['compound_id']}")
                st.write(f"**Spectrum Type:** {clean_text(row['spectrum_type'])}")
                st.write(f"**File Path:** {clean_text(row['file_path'])}")
                st.markdown('</div>', unsafe_allow_html=True)

                with st.form("delete_spectra_form"):
                    confirm = st.checkbox("I understand that this will permanently delete this spectra file record.")
                    submitted_delete = st.form_submit_button("Delete Spectra File Record")

                if submitted_delete:
                    if not confirm:
                        st.error("Please confirm deletion first.")
                    else:
                        compound_id = int(row["compound_id"])
                        delete_spectrum_file_record_by_id(file_id)
                        st.success(f"Spectra file record ID {file_id} was deleted.")
                        if st.button("Open Record", key=f"open_detail_after_delete_spectra_{file_id}"):
                            open_compound_detail(compound_id)
                            st.rerun()


# =========================
# Supabase-first cloud adapters
# =========================
def get_supabase_url() -> str:
    return get_secret_setting("SUPABASE_URL")


def get_supabase_anon_key() -> str:
    return get_secret_setting("SUPABASE_ANON_KEY")


def get_supabase_service_role_key() -> str:
    return (
        get_secret_setting("SUPABASE_SERVICE_ROLE_KEY")
        or get_secret_setting("SUPABASE_SECRET_KEY")
    )


def get_npdb_read_backend() -> str:
    backend = get_secret_setting("NPDB_READ_BACKEND", "npdb_read_backend").strip().lower()
    if not backend:
        local_secrets = PROJECT_DIR / ".streamlit" / "secrets.toml"
        if local_secrets.exists():
            try:
                match = re.search(r'NPDB_READ_BACKEND\s*=\s*["\']?([A-Za-z0-9_-]+)', local_secrets.read_text())
                if match:
                    backend = match.group(1).strip().lower()
            except Exception:
                backend = ""
    if backend in {"local", "sqlite", "desktop"}:
        return "local"
    if backend in {"supabase", "cloud", "remote"}:
        return "supabase"
    if cloud_backend_is_configured():
        return "supabase"
    return "local" if DB_PATH.exists() else "supabase"


def use_local_read_backend() -> bool:
    return get_npdb_read_backend() == "local" and DB_PATH.exists()


def use_supabase_backend() -> bool:
    if use_local_read_backend():
        return False
    return bool(get_supabase_url() and (get_supabase_service_role_key() or get_supabase_anon_key()))


def use_supabase_write_backend() -> bool:
    return bool(get_supabase_url() and get_supabase_service_role_key())


def _supabase_ssl_context():
    if get_secret_setting("NPDB_SKIP_SSL_VERIFY") == "1":
        return ssl._create_unverified_context()
    return None


def ensure_write_target_ready():
    if use_supabase_backend() and not use_supabase_write_backend():
        raise RuntimeError(
            "Cloud write mode is not configured. To keep Supabase as the single source of truth, editing is blocked until SUPABASE_SERVICE_ROLE_KEY or SUPABASE_SECRET_KEY is available."
        )


def _json_ready(value):
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _supabase_headers(write: bool = False, json_body: bool = True, extra: dict | None = None):
    api_key = get_supabase_service_role_key() if write else (get_supabase_service_role_key() or get_supabase_anon_key())
    if write and not api_key:
        raise RuntimeError("Supabase writes require SUPABASE_SERVICE_ROLE_KEY or SUPABASE_SECRET_KEY in server-side secrets.")
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
    }
    if json_body:
        headers["Content-Type"] = "application/json"
    if extra:
        headers.update(extra)
    return headers


def _supabase_request(method: str, path: str, query: dict | None = None, body=None, write: bool = False, json_body: bool = True, extra_headers: dict | None = None, return_json: bool = True):
    if write and not use_supabase_write_backend():
        raise RuntimeError("Supabase write mode is disabled because SUPABASE_SERVICE_ROLE_KEY or SUPABASE_SECRET_KEY is missing.")
    base = get_supabase_url().rstrip("/")
    url = f"{base}{path}"
    if query:
        query_text = urlencode(query, doseq=True, safe=",().:*+-")
        url = f"{url}?{query_text}"
    payload = None
    if body is not None:
        payload = json.dumps(body).encode("utf-8") if json_body else body
    request = urllib.request.Request(
        url,
        data=payload,
        method=method.upper(),
        headers=_supabase_headers(write=write, json_body=json_body, extra=extra_headers),
    )
    try:
        with urllib.request.urlopen(request, timeout=60, context=_supabase_ssl_context()) as response:
            raw = response.read()
            if not return_json:
                return raw
            if not raw:
                return None
            return json.loads(raw.decode("utf-8"))
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"Supabase request failed ({exc.code} {exc.reason}): {details}") from exc


def _supabase_filter_query(filters: dict | None) -> dict:
    query = {}
    for key, value in (filters or {}).items():
        if isinstance(value, tuple) and len(value) == 2:
            operator, operand = value
            if operator == "in":
                joined = ",".join(str(item) for item in operand)
                query[key] = f"in.({joined})"
            else:
                query[key] = f"{operator}.{operand}"
        else:
            query[key] = f"eq.{value}"
    return query


def supabase_select_df(table: str, columns: str = "*", filters: dict | None = None, order: str | None = None) -> pd.DataFrame:
    if not use_supabase_backend():
        return pd.DataFrame()
    query = {"select": columns}
    query.update(_supabase_filter_query(filters))
    if order:
        query["order"] = order
    rows = []
    start = 0
    while True:
        end = start + SUPABASE_PAGE_SIZE - 1
        page_rows = _supabase_request(
            "GET",
            f"/rest/v1/{table}",
            query=query,
            write=False,
            extra_headers={"Range-Unit": "items", "Range": f"{start}-{end}"},
        ) or []
        if isinstance(page_rows, dict):
            page_rows = [page_rows]
        rows.extend(page_rows)
        if len(page_rows) < SUPABASE_PAGE_SIZE:
            break
        start += SUPABASE_PAGE_SIZE
    return pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def supabase_column_available(table: str, column: str) -> bool:
    if not use_supabase_backend():
        return False
    try:
        _supabase_request(
            "GET",
            f"/rest/v1/{table}",
            query={"select": column, "limit": 1},
            write=False,
        )
        return True
    except RuntimeError as exc:
        message = str(exc).lower()
        if "column" in message and column.lower() in message:
            return False
        raise


def compound_select_columns() -> str:
    columns = [
        "id",
        "trivial_name",
        "iupac_name",
        "molecular_formula",
        "smiles",
        "inchi",
        "inchikey",
        "compound_class",
        "compound_subclass",
        "source_category",
        "source_organism",
        "source_material",
        "sample_code",
        "collection_location",
        "gps_coordinates",
        "depth_m",
        "uv_data",
        "ftir_data",
        "cd_data",
        "optical_rotation",
        "melting_point",
        "crystallization_method",
        "structure_image_path",
        "journal_name",
        "article_title",
        "publication_year",
        "volume",
        "issue",
        "pages",
        "doi",
        "ccdc_number",
        "molecular_weight",
        "hrms_data",
        "data_source",
        "curation_status",
        "note",
        "created_at",
        "updated_at",
    ]
    if use_supabase_backend() and not supabase_column_available("compounds", "curation_status"):
        columns = [column for column in columns if column != "curation_status"]
    return ",".join(columns)


def supabase_insert_row(table: str, row: dict):
    payload = {k: _json_ready(v) for k, v in row.items() if k and v is not None}
    response = _supabase_request(
        "POST",
        f"/rest/v1/{table}",
        query={"select": "id"},
        body=payload,
        write=True,
        extra_headers={"Prefer": "return=representation"},
    ) or []
    if response:
        return response[0]
    return {}


def supabase_update_row(table: str, row_id: int, row: dict):
    payload = {k: _json_ready(v) for k, v in row.items() if k and k != "id"}
    response = _supabase_request(
        "PATCH",
        f"/rest/v1/{table}",
        query={"id": f"eq.{row_id}", "select": "id"},
        body=payload,
        write=True,
        extra_headers={"Prefer": "return=representation"},
    ) or []
    if response:
        return response[0]
    return {}


def supabase_delete_row(table: str, row_id: int):
    _supabase_request(
        "DELETE",
        f"/rest/v1/{table}",
        query={"id": f"eq.{row_id}"},
        body=None,
        write=True,
        extra_headers={"Prefer": "return=minimal"},
    )


def supabase_upload_bytes(bucket: str, object_path: str, data: bytes, content_type: str = "application/octet-stream", public_bucket: bool = True) -> str:
    _supabase_request(
        "POST",
        f"/storage/v1/object/{bucket}/{quote(object_path, safe='/')}"
        ,body=data,
        write=True,
        json_body=False,
        extra_headers={"Content-Type": content_type, "x-upsert": "true"},
        return_json=False,
    )
    if public_bucket:
        return f"{get_supabase_url().rstrip('/')}" + f"/storage/v1/object/public/{bucket}/{quote(object_path, safe='/')}"
    return f"storage://{bucket}/{object_path}"


def _sqlite_dataframe(query: str, params: tuple | list | None = None) -> pd.DataFrame:
    conn = get_connection()
    try:
        return pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()


def _sqlite_columns(table: str) -> list[str]:
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table})")
        return [row[1] for row in cursor.fetchall()]
    finally:
        conn.close()


def _sqlite_upsert_row(table: str, row: dict) -> int | None:
    columns = [column for column in row.keys() if column in _sqlite_columns(table)]
    if not columns:
        return None
    conn = get_connection()
    try:
        cursor = conn.cursor()
        row_id = row.get("id")
        if row_id is not None:
            cursor.execute(f"SELECT 1 FROM {table} WHERE id = ?", (row_id,))
            exists = cursor.fetchone() is not None
        else:
            exists = False
        if exists:
            set_columns = [column for column in columns if column != "id"]
            assignments = ", ".join(f"{column} = ?" for column in set_columns)
            values = [row.get(column) for column in set_columns] + [row_id]
            cursor.execute(f"UPDATE {table} SET {assignments} WHERE id = ?", values)
        else:
            placeholders = ", ".join("?" for _ in columns)
            values = [row.get(column) for column in columns]
            cursor.execute(
                f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
                values,
            )
            row_id = row_id if row_id is not None else cursor.lastrowid
        conn.commit()
        return int(row_id) if row_id is not None else None
    finally:
        conn.close()


def _sqlite_delete_row(table: str, row_id: int):
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM {table} WHERE id = ?", (row_id,))
        conn.commit()
    finally:
        conn.close()


def _local_binary_path(target_dir: Path, base_name: str, suffix: str) -> Path:
    safe_name = slugify_value(base_name, fallback="asset")
    candidate = target_dir / f"{safe_name}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = target_dir / f"{safe_name}_{counter}{suffix}"
        counter += 1
    return candidate


def save_uploaded_asset(uploaded_file, target_dir: Path, base_name: str) -> str:
    suffix = Path(uploaded_file.name).suffix.lower() or ".bin"
    data = uploaded_file.getbuffer().tobytes()
    if use_supabase_write_backend():
        bucket = "exports"
        public_bucket = False
        if target_dir == STRUCTURES_DIR:
            bucket = "structures"
            public_bucket = True
        elif target_dir == SPECTRA_DIR:
            bucket = "spectra"
            public_bucket = True
        upload_time = datetime.now(UTC)
        original_stem = Path(uploaded_file.name).stem
        object_name = f"{slugify_value(base_name + '_' + original_stem, fallback='asset')}_{upload_time.strftime('%H%M%S_%f')}{suffix}"
        object_path = f"{upload_time.strftime('%Y/%m/%d')}/{object_name}"
        content_type = getattr(uploaded_file, "type", None) or mimetypes.guess_type(object_name)[0] or "application/octet-stream"
        try:
            return supabase_upload_bytes(bucket, object_path, data, content_type=content_type, public_bucket=public_bucket)
        except Exception as exc:
            raise RuntimeError(
                f"Cloud asset upload failed for '{uploaded_file.name}'. The file was not committed as a cloud-backed record."
            ) from exc
    candidate = _local_binary_path(target_dir, base_name, suffix)
    candidate.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(candidate, "wb") as output_file:
            output_file.write(data)
    except Exception:
        pass
    return relative_project_path(candidate) if candidate.exists() else str(candidate)


def _merge_compound_names(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    compounds_df = load_all_compounds()[["id", "trivial_name"]].copy()
    compounds_df = compounds_df.rename(columns={"id": "compound_id"})
    return df.merge(compounds_df, on="compound_id", how="left")


def get_db_signature():
    if use_supabase_backend():
        return 1.0
    if not DB_PATH.exists():
        return 0.0
    return DB_PATH.stat().st_mtime


@st.cache_data(show_spinner=False)
def load_all_compounds(source_normalization_version: str = SOURCE_NORMALIZATION_CACHE_VERSION):
    _ = source_normalization_version
    columns = compound_select_columns()
    if use_supabase_backend():
        df = supabase_select_df("compounds", columns=columns, order="id.asc")
        return enrich_compounds_dataframe(df)
    return enrich_compounds_dataframe(_sqlite_dataframe(f"SELECT {columns} FROM compounds ORDER BY id ASC"))


@st.cache_data(show_spinner=False)
def load_compound_row(compound_id, source_normalization_version: str = SOURCE_NORMALIZATION_CACHE_VERSION):
    _ = source_normalization_version
    columns = compound_select_columns()
    if use_supabase_backend():
        df = supabase_select_df("compounds", columns=columns, filters={"id": ("eq", compound_id)})
        return enrich_compounds_dataframe(df)
    return enrich_compounds_dataframe(_sqlite_dataframe(f"SELECT {columns} FROM compounds WHERE id = ?", (compound_id,)))


@st.cache_data(show_spinner=False)
def load_proton_data(compound_id):
    columns = "id,compound_id,delta_ppm,multiplicity,j_value,proton_count,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        return supabase_select_df("proton_nmr", columns=columns, filters={"compound_id": ("eq", compound_id)}, order="delta_ppm.desc")
    return _sqlite_dataframe(f"SELECT {columns} FROM proton_nmr WHERE compound_id = ? ORDER BY delta_ppm DESC", (compound_id,))


@st.cache_data(show_spinner=False)
def load_all_proton_data():
    columns = "id,compound_id,delta_ppm,multiplicity,j_value,proton_count,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        df = supabase_select_df("proton_nmr", columns=columns, order="id.asc")
        return _merge_compound_names(df)
    return _sqlite_dataframe("SELECT p.id, p.compound_id, c.trivial_name, p.delta_ppm, p.multiplicity, p.j_value, p.proton_count, p.assignment, p.solvent, p.instrument_mhz, p.note FROM proton_nmr p LEFT JOIN compounds c ON p.compound_id = c.id ORDER BY p.id ASC")


@st.cache_data(show_spinner=False)
def load_proton_row(proton_id):
    columns = "id,compound_id,delta_ppm,multiplicity,j_value,proton_count,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        return supabase_select_df("proton_nmr", columns=columns, filters={"id": ("eq", proton_id)})
    return _sqlite_dataframe(f"SELECT {columns} FROM proton_nmr WHERE id = ?", (proton_id,))


@st.cache_data(show_spinner=False)
def load_carbon_data(compound_id):
    columns = "id,compound_id,delta_ppm,carbon_type,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        return supabase_select_df("carbon_nmr", columns=columns, filters={"compound_id": ("eq", compound_id)}, order="delta_ppm.desc")
    return _sqlite_dataframe(f"SELECT {columns} FROM carbon_nmr WHERE compound_id = ? ORDER BY delta_ppm DESC", (compound_id,))


@st.cache_data(show_spinner=False)
def load_all_carbon_data():
    columns = "id,compound_id,delta_ppm,carbon_type,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        df = supabase_select_df("carbon_nmr", columns=columns, order="id.asc")
        return _merge_compound_names(df)
    return _sqlite_dataframe("SELECT c.id, c.compound_id, cp.trivial_name, c.delta_ppm, c.carbon_type, c.assignment, c.solvent, c.instrument_mhz, c.note FROM carbon_nmr c LEFT JOIN compounds cp ON c.compound_id = cp.id ORDER BY c.id ASC")


@st.cache_data(show_spinner=False)
def load_carbon_row(carbon_id):
    columns = "id,compound_id,delta_ppm,carbon_type,assignment,solvent,instrument_mhz,note"
    if use_supabase_backend():
        return supabase_select_df("carbon_nmr", columns=columns, filters={"id": ("eq", carbon_id)})
    return _sqlite_dataframe(f"SELECT {columns} FROM carbon_nmr WHERE id = ?", (carbon_id,))


@st.cache_data(show_spinner=False)
def load_spectra_files(compound_id):
    columns = "id,compound_id,spectrum_type,file_path,note"
    if use_supabase_backend():
        return supabase_select_df("spectra_files", columns=columns, filters={"compound_id": ("eq", compound_id)}, order="id.asc")
    return _sqlite_dataframe(f"SELECT {columns} FROM spectra_files WHERE compound_id = ? ORDER BY id ASC", (compound_id,))


@st.cache_data(show_spinner=False)
def load_all_spectra_files():
    columns = "id,compound_id,spectrum_type,file_path,note"
    if use_supabase_backend():
        df = supabase_select_df("spectra_files", columns=columns, order="id.asc")
        return _merge_compound_names(df)
    return _sqlite_dataframe("SELECT s.id, s.compound_id, c.trivial_name, s.spectrum_type, s.file_path, s.note FROM spectra_files s LEFT JOIN compounds c ON s.compound_id = c.id ORDER BY s.id ASC")


@st.cache_data(show_spinner=False)
def load_spectrum_file_row(file_id):
    columns = "id,compound_id,spectrum_type,file_path,note"
    if use_supabase_backend():
        return supabase_select_df("spectra_files", columns=columns, filters={"id": ("eq", file_id)})
    return _sqlite_dataframe(f"SELECT {columns} FROM spectra_files WHERE id = ?", (file_id,))


@st.cache_data(show_spinner=False)
def load_bioactivity_data(compound_id):
    columns = "id,compound_id,activity_label,target_name,target_category,assay_type,potency_type,potency_relation,potency_value,potency_unit,outcome,assay_medium,selectivity,assay_source,note"
    if use_supabase_backend():
        return supabase_select_df("bioactivity_records", columns=columns, filters={"compound_id": ("eq", compound_id)}, order="id.asc")
    return _sqlite_dataframe(f"SELECT {columns} FROM bioactivity_records WHERE compound_id = ? ORDER BY id ASC", (compound_id,))


@st.cache_data(show_spinner=False)
def load_all_bioactivity_data():
    columns = "id,compound_id,activity_label,target_name,target_category,assay_type,potency_type,potency_relation,potency_value,potency_unit,outcome,assay_medium,selectivity,assay_source,note"
    if use_supabase_backend():
        df = supabase_select_df("bioactivity_records", columns=columns, order="id.asc")
        return _merge_compound_names(df)
    return _sqlite_dataframe("SELECT b.id, b.compound_id, c.trivial_name, b.activity_label, b.target_name, b.target_category, b.assay_type, b.potency_type, b.potency_relation, b.potency_value, b.potency_unit, b.outcome, b.assay_medium, b.selectivity, b.assay_source, b.note FROM bioactivity_records b LEFT JOIN compounds c ON b.compound_id = c.id ORDER BY b.id ASC")


@st.cache_data(show_spinner=False)
def load_bioactivity_row(bioactivity_id):
    columns = "id,compound_id,activity_label,target_name,target_category,assay_type,potency_type,potency_relation,potency_value,potency_unit,outcome,assay_medium,selectivity,assay_source,note"
    if use_supabase_backend():
        return supabase_select_df("bioactivity_records", columns=columns, filters={"id": ("eq", bioactivity_id)})
    return _sqlite_dataframe(f"SELECT {columns} FROM bioactivity_records WHERE id = ?", (bioactivity_id,))


def count_related_records(filtered_ids):
    filtered_ids = [int(item) for item in filtered_ids if str(item).strip()]
    if not filtered_ids:
        return 0, 0, 0
    if use_supabase_backend():
        filtered_id_set = set(filtered_ids)
        proton_df = load_all_proton_data()
        carbon_df = load_all_carbon_data()
        spectra_df = load_all_spectra_files()
        proton_count = int(proton_df["compound_id"].isin(filtered_id_set).sum()) if "compound_id" in proton_df else 0
        carbon_count = int(carbon_df["compound_id"].isin(filtered_id_set).sum()) if "compound_id" in carbon_df else 0
        spectra_count = int(spectra_df["compound_id"].isin(filtered_id_set).sum()) if "compound_id" in spectra_df else 0
        return proton_count, carbon_count, spectra_count

    conn = get_connection()
    try:
        placeholders = ",".join("?" * len(filtered_ids))
        proton_query = f"SELECT COUNT(*) AS n FROM proton_nmr WHERE compound_id IN ({placeholders})"
        carbon_query = f"SELECT COUNT(*) AS n FROM carbon_nmr WHERE compound_id IN ({placeholders})"
        spectra_query = f"SELECT COUNT(*) AS n FROM spectra_files WHERE compound_id IN ({placeholders})"
        proton_count = int(pd.read_sql_query(proton_query, conn, params=filtered_ids)["n"][0])
        carbon_count = int(pd.read_sql_query(carbon_query, conn, params=filtered_ids)["n"][0])
        spectra_count = int(pd.read_sql_query(spectra_query, conn, params=filtered_ids)["n"][0])
        return proton_count, carbon_count, spectra_count
    finally:
        conn.close()


def count_bioactivity_records(filtered_ids):
    filtered_ids = [int(item) for item in filtered_ids if str(item).strip()]
    if not filtered_ids:
        return 0
    if use_supabase_backend():
        filtered_id_set = set(filtered_ids)
        bioactivity_df = load_all_bioactivity_data()
        return int(bioactivity_df["compound_id"].isin(filtered_id_set).sum()) if "compound_id" in bioactivity_df else 0

    conn = get_connection()
    try:
        placeholders = ",".join("?" * len(filtered_ids))
        query = f"SELECT COUNT(*) AS n FROM bioactivity_records WHERE compound_id IN ({placeholders})"
        return int(pd.read_sql_query(query, conn, params=filtered_ids)["n"][0])
    finally:
        conn.close()


@st.cache_data(show_spinner=False)
def load_search_index(_db_signature: float):
    compounds_df = load_all_compounds()
    all_proton_df = load_all_proton_data()
    all_carbon_df = load_all_carbon_data()
    proton_df = all_proton_df[["compound_id", "delta_ppm"]] if not all_proton_df.empty else pd.DataFrame(columns=["compound_id", "delta_ppm"])
    carbon_df = all_carbon_df[["compound_id", "delta_ppm"]] if not all_carbon_df.empty else pd.DataFrame(columns=["compound_id", "delta_ppm"])
    proton_groups = proton_df.groupby("compound_id")["delta_ppm"].apply(list).to_dict() if not proton_df.empty else {}
    carbon_groups = carbon_df.groupby("compound_id")["delta_ppm"].apply(list).to_dict() if not carbon_df.empty else {}
    search_index = []
    for _, row in compounds_df.iterrows():
        compound_id = int(row["id"])
        search_index.append(
            {
                "compound_id": compound_id,
                "trivial_name": row.get("trivial_name"),
                "sample_code": row.get("sample_code"),
                "molecular_formula": row.get("molecular_formula"),
                "source_category": row.get("source_category"),
                "source_organism": row.get("source_organism"),
                "source_material": row.get("source_material"),
                "compound_class": row.get("compound_class"),
                "compound_subclass": row.get("compound_subclass"),
                "data_source": row.get("data_source"),
                "proton_peaks": proton_groups.get(compound_id, []),
                "carbon_peaks": carbon_groups.get(compound_id, []),
            }
        )
    return search_index


def _upsert_compound_local(row: dict):
    return _sqlite_upsert_row("compounds", row)


def insert_compound_record(trivial_name, iupac_name, molecular_formula, compound_class, compound_subclass, smiles, inchi, inchikey, source_category, source_organism, source_material, sample_code, collection_location, gps_coordinates, depth_m, uv_data, ftir_data, cd_data, optical_rotation, melting_point, crystallization_method, structure_image_path, journal_name, article_title, publication_year, volume, issue, pages, doi, ccdc_number, molecular_weight, hrms_data, data_source, curation_status, note):
    ensure_write_target_ready()
    row = {
        "trivial_name": trivial_name,
        "iupac_name": iupac_name,
        "molecular_formula": molecular_formula,
        "compound_class": compound_class,
        "compound_subclass": compound_subclass,
        "smiles": smiles,
        "inchi": inchi,
        "inchikey": inchikey,
        "source_category": source_category,
        "source_organism": source_organism,
        "source_material": source_material,
        "sample_code": sample_code,
        "collection_location": collection_location,
        "gps_coordinates": gps_coordinates,
        "depth_m": depth_m,
        "uv_data": uv_data,
        "ftir_data": ftir_data,
        "cd_data": cd_data,
        "optical_rotation": optical_rotation,
        "melting_point": melting_point,
        "crystallization_method": crystallization_method,
        "structure_image_path": structure_image_path,
        "journal_name": journal_name,
        "article_title": article_title,
        "publication_year": publication_year,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "doi": doi,
        "ccdc_number": ccdc_number,
        "molecular_weight": molecular_weight,
        "hrms_data": hrms_data,
        "data_source": data_source,
        "curation_status": normalize_curation_status(curation_status, default="curated"),
        "note": note,
        "updated_at": datetime.now(UTC).isoformat(),
    }
    if use_supabase_write_backend():
        if not supabase_column_available("compounds", "curation_status"):
            row.pop("curation_status", None)
        inserted = supabase_insert_row("compounds", row)
        row_id = int(inserted.get("id")) if inserted and inserted.get("id") is not None else None
        if row_id is None:
            raise RuntimeError("Supabase insert for the compound did not return an ID, so the local mirror was not updated.")
        if use_local_read_backend():
            row["id"] = row_id
            _upsert_compound_local(row)
        invalidate_cached_views()
        return row_id
    row_id = _sqlite_upsert_row("compounds", row)
    invalidate_cached_views()
    return row_id


def update_compound_record(compound_id, trivial_name, iupac_name, molecular_formula, compound_class, compound_subclass, smiles, inchi, inchikey, source_category, source_organism, source_material, sample_code, collection_location, gps_coordinates, depth_m, uv_data, ftir_data, cd_data, optical_rotation, melting_point, crystallization_method, structure_image_path, journal_name, article_title, publication_year, volume, issue, pages, doi, ccdc_number, molecular_weight, hrms_data, data_source, curation_status, note):
    ensure_write_target_ready()
    row = {
        "id": compound_id,
        "trivial_name": trivial_name,
        "iupac_name": iupac_name,
        "molecular_formula": molecular_formula,
        "compound_class": compound_class,
        "compound_subclass": compound_subclass,
        "smiles": smiles,
        "inchi": inchi,
        "inchikey": inchikey,
        "source_category": source_category,
        "source_organism": source_organism,
        "source_material": source_material,
        "sample_code": sample_code,
        "collection_location": collection_location,
        "gps_coordinates": gps_coordinates,
        "depth_m": depth_m,
        "uv_data": uv_data,
        "ftir_data": ftir_data,
        "cd_data": cd_data,
        "optical_rotation": optical_rotation,
        "melting_point": melting_point,
        "crystallization_method": crystallization_method,
        "structure_image_path": structure_image_path,
        "journal_name": journal_name,
        "article_title": article_title,
        "publication_year": publication_year,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "doi": doi,
        "ccdc_number": ccdc_number,
        "molecular_weight": molecular_weight,
        "hrms_data": hrms_data,
        "data_source": data_source,
        "curation_status": normalize_curation_status(curation_status, default="curated"),
        "note": note,
        "updated_at": datetime.now(UTC).isoformat(),
    }
    if use_supabase_write_backend():
        if not supabase_column_available("compounds", "curation_status"):
            row.pop("curation_status", None)
        supabase_update_row("compounds", compound_id, row)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    _upsert_compound_local(row)
    invalidate_cached_views()


def delete_compound_record(compound_id):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        supabase_delete_row("compounds", compound_id)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM bioactivity_records WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM proton_nmr WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM carbon_nmr WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM spectra_files WHERE compound_id = ?", (compound_id,))
        cursor.execute("DELETE FROM compounds WHERE id = ?", (compound_id,))
        conn.commit()
    finally:
        conn.close()
    invalidate_cached_views()


def _write_child_row(table: str, row: dict, row_id: int | None = None):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        if row_id is None:
            inserted = supabase_insert_row(table, row)
            inserted_id = int(inserted.get("id")) if inserted and inserted.get("id") is not None else None
            if inserted_id is None:
                raise RuntimeError(f"Supabase insert for '{table}' did not return an ID, so the local mirror was not updated.")
            row["id"] = inserted_id
        else:
            supabase_update_row(table, row_id, row)
            row["id"] = row_id
        if not use_local_read_backend():
            return int(row["id"]) if row.get("id") is not None else None
    return _sqlite_upsert_row(table, row)


def insert_proton_record(compound_id, delta_ppm, multiplicity, j_value, proton_count, assignment, solvent, instrument_mhz, note):
    row = {"compound_id": compound_id, "delta_ppm": delta_ppm, "multiplicity": multiplicity, "j_value": j_value, "proton_count": proton_count, "assignment": assignment, "solvent": solvent, "instrument_mhz": instrument_mhz, "note": note}
    row_id = _write_child_row("proton_nmr", row)
    invalidate_cached_views()
    return row_id


def update_proton_record(proton_id, compound_id, delta_ppm, multiplicity, j_value, proton_count, assignment, solvent, instrument_mhz, note):
    row = {"compound_id": compound_id, "delta_ppm": delta_ppm, "multiplicity": multiplicity, "j_value": j_value, "proton_count": proton_count, "assignment": assignment, "solvent": solvent, "instrument_mhz": instrument_mhz, "note": note}
    _write_child_row("proton_nmr", row, row_id=proton_id)
    invalidate_cached_views()


def delete_proton_record_by_id(proton_id):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        supabase_delete_row("proton_nmr", proton_id)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    _sqlite_delete_row("proton_nmr", proton_id)
    invalidate_cached_views()


def insert_carbon_record(compound_id, delta_ppm, carbon_type, assignment, solvent, instrument_mhz, note):
    row = {"compound_id": compound_id, "delta_ppm": delta_ppm, "carbon_type": carbon_type, "assignment": assignment, "solvent": solvent, "instrument_mhz": instrument_mhz, "note": note}
    row_id = _write_child_row("carbon_nmr", row)
    invalidate_cached_views()
    return row_id


def update_carbon_record(carbon_id, compound_id, delta_ppm, carbon_type, assignment, solvent, instrument_mhz, note):
    row = {"compound_id": compound_id, "delta_ppm": delta_ppm, "carbon_type": carbon_type, "assignment": assignment, "solvent": solvent, "instrument_mhz": instrument_mhz, "note": note}
    _write_child_row("carbon_nmr", row, row_id=carbon_id)
    invalidate_cached_views()


def delete_carbon_record_by_id(carbon_id):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        supabase_delete_row("carbon_nmr", carbon_id)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    _sqlite_delete_row("carbon_nmr", carbon_id)
    invalidate_cached_views()


def insert_spectrum_file_record(compound_id, spectrum_type, file_path, note):
    row = {"compound_id": compound_id, "spectrum_type": spectrum_type, "file_path": file_path, "note": note}
    row_id = _write_child_row("spectra_files", row)
    invalidate_cached_views()
    return row_id


def update_spectrum_file_record(file_id, compound_id, spectrum_type, file_path, note):
    row = {"compound_id": compound_id, "spectrum_type": spectrum_type, "file_path": file_path, "note": note}
    _write_child_row("spectra_files", row, row_id=file_id)
    invalidate_cached_views()


def delete_spectrum_file_record_by_id(file_id):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        supabase_delete_row("spectra_files", file_id)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    _sqlite_delete_row("spectra_files", file_id)
    invalidate_cached_views()


def insert_bioactivity_record(compound_id, activity_label, target_name, target_category, assay_type, potency_type, potency_relation, potency_value, potency_unit, outcome, assay_medium, selectivity, assay_source, note):
    row = {"compound_id": compound_id, "activity_label": activity_label, "target_name": target_name, "target_category": target_category, "assay_type": assay_type, "potency_type": potency_type, "potency_relation": potency_relation, "potency_value": potency_value, "potency_unit": potency_unit, "outcome": outcome, "assay_medium": assay_medium, "selectivity": selectivity, "assay_source": assay_source, "note": note}
    row_id = _write_child_row("bioactivity_records", row)
    invalidate_cached_views()
    return row_id


def update_bioactivity_record(bioactivity_id, compound_id, activity_label, target_name, target_category, assay_type, potency_type, potency_relation, potency_value, potency_unit, outcome, assay_medium, selectivity, assay_source, note):
    row = {"compound_id": compound_id, "activity_label": activity_label, "target_name": target_name, "target_category": target_category, "assay_type": assay_type, "potency_type": potency_type, "potency_relation": potency_relation, "potency_value": potency_value, "potency_unit": potency_unit, "outcome": outcome, "assay_medium": assay_medium, "selectivity": selectivity, "assay_source": assay_source, "note": note}
    _write_child_row("bioactivity_records", row, row_id=bioactivity_id)
    invalidate_cached_views()


def delete_bioactivity_record_by_id(bioactivity_id):
    ensure_write_target_ready()
    if use_supabase_write_backend():
        supabase_delete_row("bioactivity_records", bioactivity_id)
        if not use_local_read_backend():
            invalidate_cached_views()
            return
    _sqlite_delete_row("bioactivity_records", bioactivity_id)
    invalidate_cached_views()


def derive_structure_identifiers(structure_text: str) -> dict | None:
    if not is_structure_backend_available():
        return None
    mol = structure_text_to_mol(structure_text)
    if mol is None:
        return None
    smiles_value = maybe_blank(Chem.MolToSmiles(mol, canonical=True)) if Chem is not None else ""
    inchi_value = ""
    inchikey_value = ""
    if Chem is not None:
        try:
            inchi_value = maybe_blank(Chem.MolToInchi(mol))
        except Exception:
            inchi_value = ""
        try:
            inchikey_value = maybe_blank(Chem.InchiToInchiKey(inchi_value)) if inchi_value else ""
        except Exception:
            inchikey_value = ""
    return {"mol": mol, "smiles": smiles_value, "inchi": inchi_value, "inchikey": inchikey_value}


def _save_generated_structure_image(compound_id: int, mol) -> str:
    if Draw is None or Image is None or mol is None:
        return ""
    image = normalize_structure_image(Draw.MolToImage(mol, size=(720, 540)), size=(720, 540))
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    data = output.getvalue()
    file_name = f"compound_{compound_id}_structure_{datetime.now(UTC).strftime('%H%M%S_%f')}.png"

    if use_supabase_write_backend():
        try:
            return supabase_upload_bytes("structures", f"generated/{file_name}", data, content_type="image/png")
        except Exception as exc:
            raise RuntimeError(
                f"Cloud structure image upload failed for compound ID {compound_id}. The structure metadata was not saved."
            ) from exc

    candidate = _local_binary_path(STRUCTURES_DIR, f"compound_{compound_id}_structure", ".png")
    candidate.write_bytes(data)
    return str(candidate.relative_to(PROJECT_DIR))


def save_structure_query_to_compound(compound_id: int, query_text: str) -> tuple[bool, str]:
    ensure_write_target_ready()
    identifiers = derive_structure_identifiers(query_text)
    if not identifiers:
        return False, "The current query could not be converted into searchable structure identifiers."
    structure_image_path = _save_generated_structure_image(compound_id, identifiers.get("mol"))
    payload = {
        "smiles": identifiers.get("smiles"),
        "inchi": identifiers.get("inchi"),
        "inchikey": identifiers.get("inchikey"),
        "updated_at": datetime.now(UTC).isoformat(),
    }
    if structure_image_path:
        payload["structure_image_path"] = structure_image_path
    if use_supabase_write_backend():
        supabase_update_row("compounds", compound_id, payload)
    if use_local_read_backend():
        conn = get_connection()
        try:
            cursor = conn.cursor()
            assignments = ", ".join(f"{key} = ?" for key in payload.keys())
            values = list(payload.values()) + [compound_id]
            cursor.execute(f"UPDATE compounds SET {assignments} WHERE id = ?", values)
            conn.commit()
        finally:
            conn.close()
    invalidate_cached_views()
    return True, f"Structure identifiers were saved to compound ID {compound_id}."


# =========================
# App boot
# =========================
apply_navigation_query_params()
all_compounds_df = load_all_compounds()
render_cloud_sync_notice()
if use_supabase_backend() and all_compounds_df.empty:
    try:
        load_all_compounds.clear()
    except Exception:
        pass
    all_compounds_df = load_all_compounds()
write_batch_import_templates()


# =========================
# Sidebar navigation
# =========================
with st.sidebar:
    active_section = st.session_state.get("main_section_radio", st.session_state.get("nav_section", "Dashboard"))
    render_sidebar_workspace_summary(active_section, all_compounds_df)
    render_sidebar_navigation()

main_section = st.session_state.get("nav_section", "Dashboard")
render_workspace_headbar(main_section)
# =========================
# Main routing
# =========================
if main_section == "Dashboard":
    show_overview_page(all_compounds_df)

elif main_section == "Search & Match":
    show_search_page(all_compounds_df)

elif main_section == "Compound Workspace":
    show_compound_pages()

elif main_section == "Bioactivity":
    show_bioactivity_pages()

elif main_section == "1H Peaks":
    show_proton_pages()

elif main_section == "13C Peaks":
    show_carbon_pages()

elif main_section == "Spectra Library":
    show_spectra_pages()

elif main_section == "Guide":
    show_guide_page()

with st.sidebar:
    render_sidebar_session_controls()
    render_sidebar_credit()
